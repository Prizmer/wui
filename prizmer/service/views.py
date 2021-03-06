# coding -*- coding: utf-8 -*-

from __future__ import unicode_literals
from django.shortcuts import render
from django.shortcuts import render_to_response, HttpResponse
from django.http import HttpResponseRedirect
from django.shortcuts import redirect
from django import forms
from django.core.context_processors import csrf
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from django.db import connection
#from general.models import Objects, Abonents, TypesAbonents, Meters, MonthlyValues, DailyValues, CurrentValues, VariousValues, TypesParams, Params, TakenParams, LinkAbonentsTakenParams, Resources, TypesMeters, Measurement, NamesParams, BalanceGroups, LinkMetersComportSettings, LinkMetersTcpipSettings, ComportSettings, TcpipSettings, LinkBalanceGroupsMeters, Groups80020, LinkGroups80020Meters
from general.models import  Objects, Abonents, TcpipSettings, TypesAbonents, Meters, TypesMeters,LinkAbonentsTakenParams,LinkMetersComportSettings, LinkMetersTcpipSettings, ComportSettings,  TakenParams
from django.db.models.signals import pre_save
from django.db.models.signals import post_save
from django.db.models import signals
import datetime

cfg_excel_name=""
cfg_sheet_name=""

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
# Create your views here.


    
    
class UploadFileForm(forms.Form):
    #title = forms.CharField(max_length=150)
    path  = forms.FileField()

def MakeSheet(request):
    args={}
    fileName=""
    sheets=""
    if request.is_ajax():
        if request.method == 'GET':
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            #print fileName
            directory=os.path.join(BASE_DIR,'static/cfg/')
            wb=load_workbook(directory+fileName)
            sheets=wb.sheetnames

    args['sheets']=sheets
    return render_to_response("service/service_sheets_excel.html", args)

def writeToLog(msg):
    #msg=unicode(msg)
#    directory=os.path.join(BASE_DIR,'static\\log\\')
#    if  not(os.path.exists(directory)):
#        os.mkdir(directory)
#    dir_date=datetime.datetime.now().strftime("%d-%m-%Y")        
#    if  not(os.path.exists(directory+dir_date)):
#        os.mkdir(directory+dir_date)  
#        
#    path=directory+dir_date+'\log.txt'
#     
#    f = open(path, 'w')
#    f.write(msg)
#    f.close()
    pass

def choose_service(request):
    args={}
    directory=os.path.join(BASE_DIR,'static\\cfg\\')
    if  not(os.path.exists(directory)):
        os.mkdir(directory)

    files = os.listdir(directory) 
    
    args['filesFF']= files
    return render_to_response("choose_service.html", args)

@csrf_exempt
def service_electric(request):
    args={}
    return render_to_response("service/service_electric.html", args)


def service_file(request):
    args={}
    args.update(csrf(request))    
    data_table=[]
    status='file not loaded'
    args['data_table'] = data_table
    args['status']=status

    return render_to_response("service/service_file.html", args)
    
def service_file_loading(request):
    args={}
    data_table=[]
    status='file not loaded'
    sPath=""
    if request.method == 'POST':        
        form = UploadFileForm(request.POST, request.FILES)
        #print form.as_table()
        #print form.is_valid()
        
        #print sPath
        if form.is_valid():
            sPath=os.path.join(BASE_DIR,'static/cfg/'+request.FILES['path'].name)
            handle_uploaded_file(request.FILES['path'])
            status= u'Файл загружен'
    else:
        form = UploadFileForm()
        
    args['data_table'] = data_table
    args['status']=status
    args['sPath']=sPath
    #print status
    return render_to_response("choose_service.html", args)

    
def service_electric_load(request):
    args={}
    data_table=[]
    status='file not loaded'

    if request.method == 'POST':

        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            handle_uploaded_file(request.FILES['path'])
            status='file loaded'
    else:
        form = UploadFileForm()
        
    args['data_table'] = data_table
    args['status']=status
    return render_to_response("service/service_electric.html", args)
    #return render_to_response("service/service_electric_load.html", args)
    
def handle_uploaded_file(f):

    destination = open(os.path.join(BASE_DIR,'static/cfg/'+f.name), 'wb+')
    for chunk in f.chunks():
        destination.write(chunk)
    #print 'file load'
    destination.close()
    
def load_port(request):
    args={}
    #print '!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!'
    fileName=""
    sheet    = ""
    tcp_ip_status    = ""
    object_status    = ""
    counter_status    = ""
    result=""
    if request.is_ajax():
        if request.method == 'GET':
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
            request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
            request.session["object_status"]    = object_status    = request.GET['object_status']
            request.session["counter_status"]    = counter_status    = request.GET['counter_status']
            
            #print fileName
            directory=os.path.join(BASE_DIR,'static/cfg/')
            sPath=directory+fileName
            #print sPath, sheet
            result=load_tcp_ip_or_com_ports_from_excel(sPath, sheet)
    writeToLog(result)
    if result:
        tcp_ip_status=u"Порт/ы был успешно добавлен"
    else:
        tcp_ip_status=u"Порт не был загружен, он уже существует в БД"
    
    
    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["tcp_ip_status"]=tcp_ip_status
    args["object_status"]=object_status
    args["counter_status"]=counter_status
    return render_to_response("service/service_electric.html", args)

def checkPortIsExist(ip_adr,ip_port):
    dt_ports=[]
    cursor = connection.cursor()
    sQuery="""
    SELECT guid, ip_address, ip_port, write_timeout, read_timeout, attempts, 
       delay_between_sending
  FROM tcpip_settings
  where ip_address='%s' and  ip_port='%s'"""%(unicode(ip_adr).rstrip(),unicode(ip_port).rstrip())
    #print sQuery
    cursor.execute(sQuery)
    dt_ports = cursor.fetchall()
    #print dt_ports
    if len(dt_ports):  
        return False
    else: 
        return True

def load_tcp_ip_or_com_ports_from_excel(sPath, sSheet):
    #Добавление tcp_ip портов
    global cfg_excel_name
    cfg_excel_name=sPath
    global cfg_sheet_name
    cfg_sheet_name=sSheet
    wb = load_workbook(filename = sPath)
    sheet_ranges = wb[sSheet]
    row = 2
    result=""
    IsAdded=False
    portType=sheet_ranges[u'M1'].value
    while (bool(sheet_ranges[u'G%s'%(row)].value)):
        if sheet_ranges[u'G%s'%(row)].value is not None:
            writeToLog(u'Обрабатываем строку ' + str(u'G%s '%(row)) + str(sheet_ranges[u'G%s'%(row)].value))
            ip_adr=sheet_ranges[u'K%s'%(row)].value
            ip_port=sheet_ranges[u'L%s'%(row)].value
            com_port=sheet_ranges[u'M%s'%(row)].value
            if portType==u'Com-port': #добавление com-порта
                writeToLog(com_port)
                if not com_port or com_port==None: 
                    result+="Отсутствует значение для com-порта в строке"+str(row)+". Заполните все ячейки excel таблицы."
                    break
                if not (SimpleCheckIfExist('comport_settings','name', com_port, "", "", "")):
                    add_port=ComportSettings(name=unicode(com_port).rstrip(),baudrate=9600,data_bits=8,parity=0,stop_bits=1, write_timeout=100, read_timeout=100, attempts=2, delay_between_sending=100)
                    add_port.save()
                    result+=u"Новый com-порт добавлен"
                    IsAdded=True
                else: result= u'Порт '+str(com_port)+u" уже существует"
            else:
                # проверка есть ли уже такой порт, запрос в БД с адресом и портом, если ответ пустой-добавляем, в противном случае continue
                if not ip_adr or not ip_port or ip_adr==None or ip_port==None: 
                    result+="Отсутствует значение/я для tcp/ip-порта в строке"+str(row)+". Заполните все ячейки excel таблицы."
                    break
                else:
                    if (checkPortIsExist(ip_adr,ip_port)):
                        add_port=TcpipSettings(ip_address = unicode(ip_adr).rstrip(), ip_port =int(ip_port), write_timeout =300 , read_timeout =700 , attempts =3 , delay_between_sending =400)
                        add_port.save()
                        result =u'Новый tcp/ip порт добавлен'
                        IsAdded=True
    #                add_meter = Meters(name = unicode(sheet_ranges[u'F%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'E%s'%(row)].value), address = unicode(sheet_ranges[u'E%s'%(row)].value),  factory_number_manual = unicode(sheet_ranges[u'E%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"7cd88751-d232-410c-a0ef-6354a79112f1") )
    #                add_meter.save()
                    else: result+= u'Порт '+str(ip_adr)+": "+str(ip_port)+u" уже существует"
        writeToLog( result)
        row+=1
    return IsAdded

def SimpleCheckIfExist(table1,fieldName1, value1, table2, fieldName2, value2):
    dt=[]
    cursor = connection.cursor()
    if len(table2)==0: #проверка для одной таблицы
        sQuery="""
        Select *
        from %s
        where %s.%s='%s'"""%(table1, table1, fieldName1, value1)
    else:#проверка для двух сводных таблиц
        sQuery="""
        Select *
        from %s, %s
        where %s.guid_%s=%s.guid and
        %s.%s='%s' and
        %s.%s='%s'
        """%(table1,table2, table2, table1,table1, table1, fieldName1, value1,table2, fieldName2, value2)
    #print sQuery
    #print bool(dt)
    cursor.execute(sQuery)
    dt = cursor.fetchall()

    if not dt:  
        return False
    else: 
        return True
    
def GetSimpleTable(table,fieldName,value):
    dt=[]
    cursor = connection.cursor()
    sQuery="""
        Select *
        from %s
        where %s.%s='%s'"""%(table, table, fieldName, value)
    #print sQuery
    cursor.execute(sQuery)
    dt = cursor.fetchall()
    return dt
    

def GetTableFromExcel(sPath,sSheet):
    wb = load_workbook(filename = sPath)
    ws = wb[sSheet]
    row = 1
    dt=[]
    while (bool(ws[u'A%s'%(row)].value)):
        A=ws[u'A%s'%(row)].value
        B=ws[u'b%s'%(row)].value
        C=ws[u'c%s'%(row)].value
        D=ws[u'd%s'%(row)].value
        E=ws[u'e%s'%(row)].value
        F=ws[u'f%s'%(row)].value
        G=ws[u'g%s'%(row)].value
        H=ws[u'h%s'%(row)].value
        I=ws[u'i%s'%(row)].value
        J=ws[u'j%s'%(row)].value
        K=ws[u'k%s'%(row)].value
        L=ws[u'l%s'%(row)].value
        M=ws[u'm%s'%(row)].value
        N=ws[u'n%s'%(row)].value
        
        vals =[A,B,C,D,E,F,G,H,I,J,K,L,M,N]
        dt.append(vals)
        row+=1
    return dt
    
def LoadObjectsAndAbons(sPath, sSheet):
    #Добавление объектов
    global cfg_excel_name
    cfg_excel_name=sPath
    global cfg_sheet_name
    cfg_sheet_name=sSheet
    result="Объекты не загружены"

    dtAll=GetTableFromExcel(sPath,sSheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    
    for i in range(1,len(dtAll)):
        writeToLog( u'Обрабатываем строку ' + dtAll[i][2]+' - '+dtAll[i][3])
        obj_l0=dtAll[i][0]
        writeToLog( obj_l0)
        obj_l1=dtAll[i][1]
        writeToLog(obj_l1)
        obj_l2=dtAll[i][2]
        writeToLog(obj_l2)
        abon=dtAll[i][3]
        writeToLog(abon)
        account_1=dtAll[i][4]
        writeToLog(account_1)
        account_2=dtAll[i][5]
        writeToLog(account_2)
        isNewObj_l0=SimpleCheckIfExist('objects','name',obj_l0,"","","")
        isNewObj_l1=SimpleCheckIfExist('objects','name',obj_l1,"","","")
        isNewObj_l2=SimpleCheckIfExist('objects','name',obj_l2,"","","")
        isNewAbon=SimpleCheckIfExist('objects','name', obj_l2,'abonents', 'name', abon)
        kv=0
        if not (isNewObj_l0):
            writeToLog('create object '+obj_l0)
            add_parent_object = Objects( name=obj_l0, level=0)
            add_parent_object.save()
            writeToLog('create object '+obj_l1)
            #print add_parent_object
            add_object1=Objects(name=obj_l1, level=1, guid_parent = add_parent_object)
            add_object1.save()
            writeToLog('create object '+obj_l2)
            add_object2=Objects(name=obj_l2, level=2, guid_parent = add_object1)
            add_object2.save()
            
            writeToLog('create abonent '+abon)
            add_abonent = Abonents(name = abon, account_1 =unicode(account_1), account_2 =unicode(account_2), guid_objects =add_object2, guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
            add_abonent.save()
            result=u"Объекты: "+obj_l0+", "+obj_l1+u", "+obj_l2+u","+abon+u" созданы"
            continue
        if not (isNewObj_l1):
            writeToLog('create object '+obj_l1)
            dtParent=GetSimpleTable('objects','name',obj_l0)
            if dtParent: #родительский объект есть
                guid_parent=dtParent[0][0]
                add_object1=Objects(name=obj_l1, level=1, guid_parent = Objects.objects.get(guid=guid_parent))
                add_object1.save()                
                add_object2=Objects(name=obj_l2, level=2, guid_parent = add_object1)
                add_object2.save()
                writeToLog('create abonent '+abon)
                add_abonent = Abonents(name = abon, account_1 =unicode(account_1), account_2 =unicode(account_2), guid_objects =add_object2, guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                add_abonent.save()
                result+=u"Объекты: "+obj_l1+u", "+obj_l2+u","+abon+u" созданы"
                continue
        if not (isNewObj_l2):
            writeToLog('create object '+obj_l2)
            dtParent=GetSimpleTable('objects','name',obj_l1)
            if dtParent: #родительский объект есть
                guid_parent=dtParent[0][0]                
                add_object = Objects(name=obj_l2, level=2, guid_parent = Objects.objects.get(guid=guid_parent))
                add_object.save()
                result+=u"Объект: "+obj_l2+u" создан"
        if not (isNewAbon):
            writeToLog('create abonent '+ abon)
            dtObj=GetSimpleTable('objects','name',obj_l2)
            if dtObj: #родительский объект есть
                guid_object=dtObj[0][0]
                add_abonent = Abonents(name = abon, account_1 =unicode(account_1), account_2 =unicode(account_2), guid_objects = Objects.objects.get(guid=guid_object), guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                add_abonent.save()
                kv+=1

    result+=u" Прогружено "+str(kv)+u" абонентов"

    return result

def load_electric_objects(request):
    args={}
    fileName=""
    sheet    = ""
    tcp_ip_status    = ""
    object_status    = ""
    counter_status    = ""
    result="Не загружено"
    writeToLog('test1')
    if request.is_ajax():
        if request.method == 'GET':
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
            request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
            request.session["object_status"]    = object_status    = request.GET['object_status']
            request.session["counter_status"]    = counter_status    = request.GET['counter_status']
            
            directory=os.path.join(BASE_DIR,'static/cfg/')
            sPath=directory+fileName
            writeToLog(sPath)
            #print 'Path:_____',sPath, sheet
            result=LoadObjectsAndAbons(sPath, sheet)
    
    object_status=result#"Загрузка объектов условно прошла"

    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["tcp_ip_status"]=tcp_ip_status
    args["object_status"]=object_status
    args["counter_status"]=counter_status
    return render_to_response("service/service_electric.html", args)
    
def LoadElectricMeters(sPath, sSheet):
    global cfg_excel_name
    cfg_excel_name=sPath
    global cfg_sheet_name
    cfg_sheet_name=sSheet
    result=u"Счётчики не загружены"
    #print type(sPath), sPath, type(sSheet), sSheet
    dtAll=GetTableFromExcel(sPath,sSheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    met=0
    #print 'load dt - ok'
    for i in range(1,len(dtAll)):
        writeToLog(u'Обрабатываем строку ' + unicode(dtAll[i][3])+' - '+unicode(dtAll[i][6]))
        obj_l2=unicode(dtAll[i][2]) #корпус
        abon=unicode(dtAll[i][3]) #квартира
        meter=unicode(dtAll[i][6]) #номер счётчика
        adr=unicode(dtAll[i][7]) #номер в сети
        type_meter=unicode(dtAll[i][8]) #тип счётчика
        NumLic=unicode(dtAll[i][5]) #номер лицевого счёта, тут используется как пароль для м-230-ум
        Group=unicode(dtAll[i][12])
        attr1=unicode(dtAll[i][13])
#        print obj_l2
#        print abon
#        print meter
#        print adr
#        print type_meter
        isNewMeter=SimpleCheckIfExist('meters','factory_number_manual',meter,"","","")
        isNewAbon=SimpleCheckIfExist('objects','name', obj_l2,'abonents', 'name', abon)        
        
        #writeToLog( u'счётчик существует ', isNewMeter)
        if not (isNewAbon):
            return u"Сначала создайте стурктуру объектов и абонентов"
        if not (isNewMeter):
            
            #writeToLog('create meter '+meter +" adress: "+adr)
            
            if unicode(type_meter) == u'М-200':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"6224d20b-1781-4c39-8799-b1446b60774d") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'М-200')
                
                
            elif unicode(type_meter) == u'М-230':
                writeToLog('m-230')
#                print unicode(type_meter)
#                print unicode(meter)
#                print unicode(adr)
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), password = 111111 , factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"423b33a7-2d68-47b6-b4f6-5b470aedc4f4") )
#                print add_meter
#                print 'bryak'
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'М-230')
                
            elif unicode(type_meter) == u'М-230-УМ':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), password = unicode(NumLic) , factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"20e4767a-49e5-4f84-890c-25e311339c28") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'М-230-УМ')
                
            elif unicode(type_meter) == u'Эльф 1.08':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"1c5a8a80-1c51-4733-8332-4ed8d510a650") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Эльф 1.08')
            elif unicode(type_meter) == u'СПГ762-1':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"c3ec5c22-d184-41c5-b6bf-66fa30215a41") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'СПГ762-1')
                
            elif unicode(type_meter) == u'СПГ762-2':
                add_meter = Meters(name=unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"5eb7dd59-faf9-4ead-8654-4f3de74de2b0") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'СПГ762-2')
            elif unicode(type_meter) == u'СПГ762-3':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"e4fb7950-a44f-41f0-a6ff-af5e30d9d562") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'СПГ762-3')
            elif unicode(type_meter) == u'Sayany':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"5429b439-233e-4944-b91b-4b521a10f77b") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Sayany')
            elif unicode(type_meter) == u'Tekon_hvs':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), password = unicode(Group), guid_types_meters = TypesMeters.objects.get(guid = u"8398e7d6-39f7-45d2-9c45-a1c48e751b61") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Tekon_gvs')
            elif unicode(type_meter) == u'Tekon_hvs':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), password = unicode(Group), guid_types_meters = TypesMeters.objects.get(guid = u"64f02a2c-41e1-48b2-bc72-7873ea9b6431") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Tekon_gvs')

            elif unicode(type_meter) == u'Tekon_heat':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), password = unicode(Group), guid_types_meters = TypesMeters.objects.get(guid = u"b53173f2-2307-4b70-b84c-61b634521e87") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Tekon_heat')
            elif unicode(type_meter) == u'Пульсар ХВС':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), attr1 = unicode(attr1), guid_types_meters = TypesMeters.objects.get(guid = u"f1789bb7-7fcd-4124-8432-40320559890f") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Пульсар ХВС')
            
            elif unicode(type_meter) == u'Пульсар ГВС':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), attr1 = unicode(attr1), guid_types_meters = TypesMeters.objects.get(guid = u"a1a349ba-e070-4ec9-975d-9f39e61c34da") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Пульсар ГВС')

            elif unicode(type_meter) == u'Пульсар Теплосчётчик':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"82b96b1c-31cf-4753-9d64-d22e2f4d036e") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Пульсар Теплосчётчик')
            else:
                writeToLog(u'Не найдено совпадение с существующим типом прибора')
                met-=1
            met+=1
            
    result=u" Загружено счётчиков "+str(met)
    
    return result


def load_electric_counters(request):
    args={}
    fileName=""
    sheet    = ""
    tcp_ip_status    = ""
    object_status    = ""
    counter_status    = ""
    result=""
    if request.is_ajax():
        if request.method == 'GET':
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
            request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
            request.session["object_status"]    = object_status    = request.GET['object_status']
            request.session["counter_status"]    = counter_status    = request.GET['counter_status']
            directory=os.path.join(BASE_DIR,'static/cfg/')
            sPath=directory+fileName
            result=LoadElectricMeters(sPath, sheet)
    counter_status=result#"Загрузка счётчиков условно прошла"
        
    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["tcp_ip_status"]=tcp_ip_status
    args["object_status"]=object_status
    args["counter_status"]=counter_status
    return render_to_response("service/service_electric.html", args)


@csrf_exempt
def service_water(request):
    args={}
    return render_to_response("service/service_water.html", args)
    
def add_link_meter(sender, instance, created, **kwargs):
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    writeToLog( unicode(dtAll[1][1]))
    if (dtAll[1][1] == u'Объект'): #вода
        writeToLog(u'Добавляем связь портов по воде')
        add_link_meter_port_from_excel_cfg_water(sender, instance, created, **kwargs)
    else:# электрика
        writeToLog(u'Добавляем связь портов по электрике')
        add_link_meter_port_from_excel_cfg_electric(sender, instance, created, **kwargs)

def add_link_meter_port_from_excel_cfg_water(sender, instance, created, **kwargs):
    """Делаем привязку счётчика к порту по excel файлу ведомости"""
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    i=3
    ip_adr=unicode(dtAll[i][7]).strip()
    ip_port=unicode(dtAll[i][8]).strip()
# Привязка к tpc порту
    guid_ip_port_from_excel = connection.cursor()
    sQuery="""SELECT 
                                      tcpip_settings.guid
                                    FROM 
                                      public.tcpip_settings
                                    WHERE 
                                      tcpip_settings.ip_address = '%s' AND 
                                      tcpip_settings.ip_port = '%s';"""%(unicode(ip_adr), unicode(ip_port))
    #print sQuery
    guid_ip_port_from_excel.execute(sQuery)
    guid_ip_port_from_excel = guid_ip_port_from_excel.fetchall()

    if guid_ip_port_from_excel:
        guid_ip_port = TcpipSettings.objects.get(guid=guid_ip_port_from_excel[0][0])
        add_ip_port_link = LinkMetersTcpipSettings(guid_meters = instance, guid_tcpip_settings = guid_ip_port)            
        add_ip_port_link.save()
    else: writeToLog(u'Нет tcp-ip порта, создайте его!')

def add_link_meter_port_from_excel_cfg_electric(sender, instance, created, **kwargs):
    """Делаем привязку счётчика к порту по excel файлу ведомости"""    
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    
    for i in range(1,len(dtAll)):
        writeToLog(u'Обрабатываем строку ' + unicode(dtAll[i][6])+' - '+unicode(dtAll[i][7]))
        meter=dtAll[i][6] #счётчик
        #print dtAll[0][11], dtAll[0][12]
        PortType=dtAll[0][11] # com или tcp-ip
        #print 'i=',i,' len=', len(dtAll)
        ip_adr=unicode(dtAll[i][10]).strip()
        ip_port=unicode(dtAll[i][11]).strip()
        # Привязка к tpc порту
        if meter is not None:
            if unicode(meter) == instance.factory_number_manual :
                if unicode(PortType) == u'Com-port':
                    guid_com_port_from_excel = connection.cursor()
                    guid_com_port_from_excel.execute("""SELECT 
                                                      comport_settings.guid
                                                    FROM 
                                                      public.comport_settings
                                                    WHERE 
                                                      comport_settings.name = '%s';"""%(unicode(dtAll[i][12])))
                    guid_com_port_from_excel = guid_com_port_from_excel.fetchall()
            
                    guid_com_port = ComportSettings.objects.get(guid=guid_com_port_from_excel[0][0])
                    add_com_port_link = LinkMetersComportSettings(guid_meters = instance, guid_comport_settings = guid_com_port)
                    add_com_port_link.save()
                
                else:
                    guid_ip_port_from_excel = connection.cursor()
                    sQuery="""SELECT tcpip_settings.guid
                                                    FROM 
                                                      public.tcpip_settings
                                                    WHERE 
                                                      tcpip_settings.ip_address = '%s' AND 
                                                      tcpip_settings.ip_port = '%s';"""%(ip_adr, ip_port)
                    #print sQuery
                    guid_ip_port_from_excel.execute(sQuery)
                    guid_ip_port_from_excel = guid_ip_port_from_excel.fetchall()
            
                    #print guid_ip_port_from_excel
                    if (len(guid_ip_port_from_excel)>0):
                        guid_ip_port = TcpipSettings.objects.get(guid=guid_ip_port_from_excel[0][0])
                        add_ip_port_link = LinkMetersTcpipSettings(guid_meters = instance, guid_tcpip_settings = guid_ip_port)            
                        add_ip_port_link.save()
                    else: writeToLog( u'Вы забыли загрузить порты')
            else:
                pass
            
signals.post_save.connect(add_link_meter, sender=Meters)


def add_link_abonents_taken_params(sender, instance, created, **kwargs):
    def get_taken_param_by_abonent_from_excel_cfg(input_taken_param):
        """Функция, которая читает excel файл. Составляет имя считываемого параметра типа "Пульсар 16M 33555 Пульсар 16M Канал 11". В случае совпадения должна привязать этот параметр к абоненту. Абоненты должны быть предварительно созданы."""    
        dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    
        def shrink_taken_param_name(taken_param_name):
            if taken_param_name.find(u'Текущий') != -1: # Ищем слово "Текущий"
                nn = taken_param_name.find(u'Текущий')  # Если нашли. то Записываем позицию где.        
            elif taken_param_name.find(u'Суточный') != -1:
                nn = taken_param_name.find(u'Суточный')
            else:
                pass
            return taken_param_name[:nn -1]

        for i in range(2,len(dtAll)):
            #taken_param = u'Пульсар' + u' ' + unicode(dtAll[i][3])[17:20] + u' ' + unicode(dtAll[i][3])[2:8] + u' ' + u'Пульсар' + u' ' + unicode(dtAll[i][3])[17:20] + u' ' + u'Канал' + u' ' + unicode(dtAll[i][4])
            taken_param = unicode(dtAll[i][6]) + u' ' + unicode(dtAll[i][5]) + u' '+ unicode(dtAll[i][6]) + u' ' + u'Канал' + u' ' + unicode(dtAll[i][4])
#            print taken_param
#            print shrink_taken_param_name(input_taken_param)
            if taken_param == shrink_taken_param_name(input_taken_param):
                try:
                    return unicode(dtAll[i][2])
                except:
                    return None
            else:
                pass
    
    writeToLog(u'--------')
    writeToLog(instance.name)
    writeToLog(u'==>', get_taken_param_by_abonent_from_excel_cfg(instance.name))
    if get_taken_param_by_abonent_from_excel_cfg(instance.name) is not None:
        writeToLog(u'Совпадение')
        try:
            add_link_abonents_taken_param = LinkAbonentsTakenParams (name = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(instance.name)).name + u" " + instance.guid_params.guid_names_params.name + u" " + instance.guid_params.guid_types_params.name ,coefficient=1, coefficient_2 = 1, guid_abonents = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(unicode(instance.name))) , guid_taken_params = instance )
            add_link_abonents_taken_param.save()
        except:
            pass
    else:
        pass
    
            
def add_link_abonents_taken_params2(sender, instance, created, **kwargs):
    writeToLog(instance.name)
    isExistTakenParam=SimpleCheckIfExist('taken_params','name',instance.name,"","","")
    if not isExistTakenParam:
        writeToLog(u'Параметра не существует!!! Связать невозможно')
        return None
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    for i in range(2,len(dtAll)):
        abon=unicode(dtAll[i][2])
        type_pulsar=unicode(dtAll[i][6])
        channel=unicode(dtAll[i][4])
        num_pulsar=unicode(dtAll[i][5])
        taken_param = type_pulsar+u' '+num_pulsar+u' '+type_pulsar+u' Канал '+channel+u' Суточный -- adress: '+channel+u'  channel: 0'
        #print taken_param
        if (taken_param==instance.name):
            isExistAbonent=SimpleCheckIfExist('abonents','name',abon,'','','')
            if isExistAbonent:
                writeToLog(u'Совпадение')
                #"ХВС, №47622 Канал 4 Суточный"
                guidAbon=GetSimpleTable('abonents','name',abon)[0][0]
                
                linkName=abon+u' Канал '+channel+' Суточный'
                writeToLog(linkName)
                try:
                    add_link_abonents_taken_param = LinkAbonentsTakenParams (name = linkName,coefficient=1, coefficient_2 = 1, guid_abonents = Abonents.objects.get(guid=guidAbon) , guid_taken_params = instance )
                    add_link_abonents_taken_param.save()
                    writeToLog(u'Связь добавлена: '+abon+u' -- '+taken_param)
                except:
                    writeToLog(u'ошибка')
                else:
                    pass
    
#    
#    
#    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
#    for i in range(2,len(dtAll)):
#            #taken_param = u'Пульсар' + u' ' + unicode(dtAll[i][3])[17:20] + u' ' + unicode(dtAll[i][3])[2:8] + u' ' + u'Пульсар' + u' ' + unicode(dtAll[i][3])[17:20] + u' ' + u'Канал' + u' ' + unicode(dtAll[i][4])
#            # "Пульсар 2M 062726 Пульсар 2M Канал 1 Суточный -- adress: 1  channel: 0"
#            # "Пульсар 10M 203677 Пульсар 10M Канал 7 Суточный -- adress: 7  channel: 0"
#        type_pulsar=unicode(dtAll[i][6])
#        channel=unicode(dtAll[i][4])
#        num_pulsar=unicode(dtAll[i][5])
#        taken_param = type_pulsar+u' '+num_pulsar+u' '+type_pulsar+u' Канал '+channel+u' Суточный -- adress: '+channel+u'  channel: 0'
#        print taken_param
#    
#    print u'--------'
#    print instance.name
#    print u'==>', get_taken_param_by_abonent_from_excel_cfg(instance.name)
#    if get_taken_param_by_abonent_from_excel_cfg(instance.name) is not None:
#        print u'Совпадение'
#        try:
#            add_link_abonents_taken_param = LinkAbonentsTakenParams (name = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(instance.name)).name + u" " + instance.guid_params.guid_names_params.name + u" " + instance.guid_params.guid_types_params.name ,coefficient=1, coefficient_2 = 1, guid_abonents = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(unicode(instance.name))) , guid_taken_params = instance )
#            add_link_abonents_taken_param.save()
#        except:
#            pass
#    else:
#        pass

def add_link_taken_params(sender, instance, created, **kwargs):
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    if (dtAll[1][1] == u'Объект'): #вода
        add_link_abonents_taken_params2(sender, instance, created, **kwargs)
    else:# электрика
        add_link_abonent_taken_params_from_excel_cfg_electric(sender, instance, created, **kwargs)


def add_link_abonent_taken_params_from_excel_cfg_electric(sender, instance, created, **kwargs):
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    #print dtAll[0][0]
    for i in range(1,len(dtAll)):
        meter=dtAll[i][6]
        abon=unicode(dtAll[i][3])
        obj=unicode(dtAll[i][2])
        if meter is not None:
            cursor = connection.cursor()
            sQuery="""SELECT abonents.guid FROM public.objects, public.abonents
                      WHERE objects.guid = abonents.guid_objects 
                      AND abonents.name = '%s' 
                      AND objects.name = '%s';"""%(abon,obj )
            #print sQuery
            cursor.execute(sQuery)
            guid_abonent_by_excel = cursor.fetchall()
            #print guid_abonent_by_excel

            if unicode(meter) == instance.guid_meters.factory_number_manual:
                writeToLog(u'Абонент найден' + u' ' + unicode(instance.name))
                #print guid_abonent_by_excel 
                add_link_abonents_taken_param = LinkAbonentsTakenParams (name = unicode(dtAll[i][3]) + u' - ' +  unicode(instance.guid_meters.name)  ,coefficient=unicode(dtAll[i][9]), coefficient_2 = 1, guid_abonents = Abonents.objects.get(guid =guid_abonent_by_excel[0][0]), guid_taken_params = instance)
                add_link_abonents_taken_param.save()
            else:
                pass
    
signals.post_save.connect(add_link_taken_params, sender=TakenParams)

def load_water_objects(request):
    args={}
    fileName=""
    sheet    = ""
    tcp_ip_status    = ""
    object_status    = ""
    counter_status    = ""
    result="Не загружено"
    if request.is_ajax():
        if request.method == 'GET':
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
            request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
            request.session["object_status"]    = object_status    = request.GET['object_status']
            request.session["counter_status"]    = counter_status    = request.GET['counter_status']
            
            directory=os.path.join(BASE_DIR,'static/cfg/')
            sPath=directory+fileName
            result=LoadObjectsAndAbons_water(sPath, sheet)
    
    object_status=result

    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["port_status"]=tcp_ip_status
    args["object_status"]=object_status
    args["counter_status"]=counter_status
    return render_to_response("service/service_water.html", args)
    
def CheckIfExistInObjects(name_parent, name_child):
    dt=[]
    cursor = connection.cursor()
    sQuery="""
    With obj as 
(Select guid as guid_child, objects.name as name_child, objects.level as level_child, guid_parent
 from objects)
Select guid as grand_parent, objects.name as name_parent, objects.level, objects.guid_parent, 
obj.guid_child,obj.name_child, obj.level_child, obj.guid_parent
FROM 
  public.objects, obj
where obj.guid_parent=objects.guid
and objects.name='%s' 
and obj.name_child='%s'
order by name_parent    """%(name_parent, name_child)
    cursor.execute(sQuery)
    dt = cursor.fetchall()

    if not dt:  
        return None
    else: 
        return dt[0][4]# возвращаем guid квариры
    
    
def LoadObjectsAndAbons_water(sPath, sheet):
    result=""
    dtAll=GetTableFromExcel(sPath,sheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    kv=0
    for i in range(2,len(dtAll)):
        obj_l0=u'Вода' # всегда будет Вода как объект-родитель
        obj_l1=dtAll[i][0] #корпус
        obj_l2=dtAll[i][1] #квартира
        if not dtAll[i][1] or dtAll[i][1]==None:
            j=i
            while not obj_l2 or obj_l2==None:
                j-=1
                obj_l2=dtAll[j][1]
        abon=dtAll[i][2] #абонент он же счётчик по воде
#        chanel=dtAll[i][4] # канал пульсара
#        numPulsar=dtAll[i][5] #номер пульсара
#        typePulsar=dtAll[i][5] #тип пульсара
        isNewObj_l0=SimpleCheckIfExist('objects','name',obj_l0,"","","")#вода
        isNewObj_l1=SimpleCheckIfExist('objects','name',obj_l1,"","","")#корпус
        
        guid_obj2=CheckIfExistInObjects(obj_l1, obj_l2)#возвращает guid квартиры или None
        
        isNewAbon=SimpleCheckIfExist('objects','name', obj_l2,'abonents', 'name', abon)
        
        #print 'isNewObj_l0 ', not isNewObj_l0,'isNewObj_l1 ', not isNewObj_l1, 'guid_obj2 ', str(guid_obj2), ' IsNewAbon', not isNewAbon 
        #print i, obj_l1, obj_l2, abon
        if not (isNewObj_l0):
            writeToLog('Level 0 create object '+obj_l0)
            add_parent_object = Objects(name=obj_l0, level=0) 
            add_parent_object.save()
            writeToLog( " Ok")
            writeToLog('create object '+obj_l1)
            #print add_parent_object
            add_object1=Objects(name=obj_l1, level=1, guid_parent = add_parent_object)
            add_object1.save()
            writeToLog('create object '+obj_l2)
            add_object2=Objects(name=obj_l2, level=2, guid_parent = add_object1)
            add_object2.save()            
            writeToLog('create abonent '+abon)
            add_abonent = Abonents(name = abon, guid_objects =add_object2, guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
            add_abonent.save()
            kv+=1
            result=u"Объекты: "+obj_l0+", "+obj_l1+u", "+obj_l2+u","+abon+u" созданы"
            continue
        if not (isNewObj_l1):#новый корпус
            writeToLog('Level 1 create object '+obj_l1)
            dtParent=GetSimpleTable('objects','name',obj_l0)
            if dtParent: #родительский объект есть - корпус
                guid_parent=dtParent[0][0]
                add_object1=Objects(name=obj_l1, level=1, guid_parent = Objects.objects.get(guid=guid_parent))
                add_object1.save()                
                writeToLog('create object '+obj_l2)
                add_object2=Objects(name=obj_l2, level=2, guid_parent = add_object1)
                add_object2.save()
                writeToLog('create abonent '+abon)
                add_abonent = Abonents(name = abon, guid_objects =add_object2, guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                add_abonent.save()
                kv+=1
                result+=u"Объекты: "+obj_l1+u", "+obj_l2+u","+abon+u" созданы"
                continue
            else: 
                writeToLog(u'Не удалось создать объект '+obj_l1)
                continue
            
        if bool(not guid_obj2): #новая квартира
            #переделать добавление на добавление по гуиду
            writeToLog('Level 2 create object '+obj_l2)
            dtParent=GetSimpleTable('objects','name',obj_l1)
            if dtParent: #родительский объект есть
                guid_parent=dtParent[0][0]
                add_object = Objects(name=obj_l2, level=2, guid_parent = Objects.objects.get(guid=guid_parent))
                add_object.save()
                result+=u"Объект: "+obj_l2+u" создан"
                add_abonent = Abonents(name = abon, guid_objects = add_object, guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                add_abonent.save()
                kv+=1
        if not (isNewAbon):
            writeToLog('Just create abonent '+ abon)
            if bool(guid_obj2): #родительский объект есть
                add_abonent = Abonents(name = abon, guid_objects = Objects.objects.get(guid=guid_obj2), guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                add_abonent.save()
                kv+=1            
#            else: 
#                print u'Не удалось создать объект '+abon
                continue

    result+=u" Прогружено "+str(kv)+u" водо-счётчиков"
    return result
    
def load_water_pulsar(request):
    args={}
    result=""
    if request.is_ajax():
        if request.method == 'GET':
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
            request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
            request.session["object_status"]    = object_status    = request.GET['object_status']
            request.session["counter_status"]    = counter_status    = request.GET['counter_status']
            directory=os.path.join(BASE_DIR,'static/cfg/')
            sPath=directory+fileName
            result=LoadWaterPulsar(sPath, sheet)
    counter_status=result#"Загрузка счётчиков условно прошла"
        
    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["tcp_ip_status"]=tcp_ip_status
    args["object_status"]=object_status
    args["counter_status"]=counter_status
    return render_to_response("service/service_water.html", args)
    
def LoadWaterPulsar(sPath, sSheet):
    global cfg_excel_name
    cfg_excel_name=sPath
    global cfg_sheet_name
    cfg_sheet_name=sSheet
    result=u""
    dtAll=GetTableFromExcel(sPath,sSheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    met=0
    con=0
    for i in range(2,len(dtAll)):
        obj_l0=u'Вода' # всегда будет Вода как объект-родитель
        obj_l1=dtAll[i][0] #корпус
        obj_l2=dtAll[i][1] #квартира
        if not dtAll[i][1] or dtAll[i][1]==None:
            j=i
            while not obj_l2 or obj_l2==None:
                j-=1
                obj_l2=dtAll[j][1]
        abon=unicode(dtAll[i][2]) #абонент он же счётчик по воде
        numPulsar=unicode(dtAll[i][5]) #номер пульсара
        typePulsar=unicode(dtAll[i][6]) #тип пульсара
        
        isNewAbon=SimpleCheckIfExist('objects','name', obj_l2,'abonents', 'name', abon)
        isNewPulsar=SimpleCheckIfExist('meters','address', numPulsar,'','','')
        #writeToLog(u'пульсар существует '+ unicode(isNewPulsar)+ typePulsar+ numPulsar)
        if not (isNewAbon):
            return u"Сначала создайте стурктуру объектов и счётчиков"
        if not (isNewPulsar):
            writeToLog(u'Обрабатываем строку '+unicode(obj_l2) +' '+ unicode(numPulsar))
            if unicode(typePulsar) == u'Пульсар 10M':
                    add_meter = Meters(name = unicode(typePulsar) + u' ' + unicode(numPulsar), address = unicode(numPulsar), factory_number_manual = unicode(numPulsar), guid_types_meters = TypesMeters.objects.get(guid = u"cae994a2-6ab9-4ffa-aac3-f21491a2de0b") )
                    add_meter.save()
                    writeToLog( u'OK Прибор добавлен в базу')
                    met+=1
            elif unicode(typePulsar) == u'Пульсар 16M':
                   add_meter = Meters(name = unicode(unicode(typePulsar) + u' ' + unicode(numPulsar)), address = unicode(numPulsar),  factory_number_manual = unicode(numPulsar), guid_types_meters = TypesMeters.objects.get(guid = u"7cd88751-d232-410c-a0ef-6354a79112f1") )
                   add_meter.save()
                   writeToLog(u'OK  Прибор добавлен в базу')
                   met+=1
            elif unicode(typePulsar) == u'Пульсар 2M':
                   add_meter = Meters(name = unicode(unicode(typePulsar) + u' ' + unicode(numPulsar)), address = unicode(numPulsar),  factory_number_manual = unicode(numPulsar), guid_types_meters = TypesMeters.objects.get(guid = u"6599be9a-1f4d-4a6e-a3d9-fb054b8d44e8") )
                   add_meter.save()
                   writeToLog(u'OK Прибор добавлен в базу')
                   met+=1
            else:
                writeToLog(u'Такой Пульсар уже есть')
        else:
            # надо проверить каналы и подсоединить их 
            #Пульсар 16M 029571 Пульсар 16M Канал 16 Суточный -- adress: 16  channel: 0
            chanel=unicode(dtAll[i][4])
            pulsarName=unicode(dtAll[i][6])
            abonent_name=unicode(dtAll[i][2])
            taken_param = pulsarName + u' ' + unicode(dtAll[i][5]) + u' '+ pulsarName + u' ' + u'Канал ' + chanel+ u' Суточный -- adress: ' +chanel+u'  channel: 0'
            writeToLog(taken_param)
            #Sravnenie(taken_param)
            dtTakenParam=GetSimpleTable('taken_params','name',taken_param)
            #writeToLog(bool(dtTakenParam))
            if dtTakenParam:                
                writeToLog(u'taken param найден')
                guid_taken_param=dtTakenParam[0][1]
                dtLink=GetSimpleTable('link_abonents_taken_params','guid_taken_params',guid_taken_param)
                if (dtLink):
                    result+=u"\n Привязка канала "+chanel+u" Пульсара "+pulsarName+u" уже существует. Перезапись НЕ произведена для счётчика "+abonent_name
                    continue
                dtAbon=GetSimpleTable('abonents','name', abonent_name)
                guidAbon=dtAbon[0][0]
                #print guidAbon
                #print guid_taken_param
                #print TakenParams.objects.get(guid=guid_taken_param) 
                #"миномес ГВС, №68208 Канал 5 Суточный"
                add_link_abonents_taken_param = LinkAbonentsTakenParams (name = abonent_name+u' Канал '+chanel+u' Суточный',coefficient=1, coefficient_2 = 1, guid_abonents = Abonents.objects.get(guid =guidAbon), guid_taken_params = TakenParams.objects.get(guid=guid_taken_param) )
                add_link_abonents_taken_param.save()
                writeToLog(u'Abonent connected with taken param')
                con+=1
    result=u'Прогружено новых пульсаров '+unicode(met)
    if con>0:
        result+=u'Созданы новые связи'
    return result

#def Sravnenie(takenParam):
#    str_bd='Пульсар 2М 062726 Пульсар 2M Канал 1 Суточный -- adress: 1 channel: 0'
#    i=0
#    print str_bd
#    while i!=len(takenParam):
#        if ord(takenParam[i])!=ord(str_bd[i]):
#            print i, takenParam[i]
#        i+=1

def load_water_port(request):
    args={}

    fileName=""
    sheet    = ""
    tcp_ip_status    = ""
    result=""
    if request.is_ajax():
        if request.method == 'GET':
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
            request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
            
            #print fileName
            directory=os.path.join(BASE_DIR,'static/cfg/')
            sPath=directory+fileName
            #print sPath, sheet
            result=load_tcp_ip_water_ports_from_excel(sPath, sheet)
    #print result
    if result:
        tcp_ip_status=u"Порт/ы был успешно добавлен"
    else:
        tcp_ip_status=u"Порт не был загружен, он уже существует в БД"
    
    
    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["tcp_ip_status"]=tcp_ip_status

    return render_to_response("service/service_water.html", args)
  
def change_electric_meters(request):
    args={}

    old_meter=u''
    new_meter=u''
    change_meter_status=u"Функция в разработке"
    if request.is_ajax():
        if request.method == 'GET':
            
            request.session["old_meter"]    = old_meter    = request.GET.get('old_meter')
            request.session["new_meter"]    = new_meter   = request.GET.get('new_meter')
            if (not old_meter or old_meter==None or new_meter==None or not new_meter):
                change_meter_status=u"Заполните обе ячейки"
            else:
                change_meter_status=ChangeMeters(old_meter, new_meter)

                
    #change_meter_status=unicode(old_meter)+unicode(new_meter)
    args["change_meter_status"]=change_meter_status

    return render_to_response("service/service_change_electric.html", args)
  
def ChangeMeters(old_meter, new_meter):
    result=u""
    isExistOldMeter=SimpleCheckIfExist('meters','factory_number_manual',old_meter,"","","")
    isExistNewMeter=SimpleCheckIfExist('meters','factory_number_manual',new_meter,"","","")
    if not isExistOldMeter:
        return u"Номера старого счётчика нет в базе"
    if isExistNewMeter:
        return u"Новый счётчик уже существует а базе"
    
    dtOldMeter=GetSimpleTable('meters','factory_number_manual', old_meter)
    guidOldMeter=unicode(dtOldMeter[0][0])
    
    dtTakenParams=GetSimpleTable('taken_params','guid_meters', guidOldMeter)
    
    oldName=unicode(dtOldMeter[0][1])
    newName=oldName.replace(old_meter,new_meter) #поменять на срез+ добавление или формировать полность по новой
    old_factory_number_manual=unicode(dtOldMeter[0][5])
    new_factory_number_manual=old_factory_number_manual.replace(old_meter,new_meter)  #поменять на срез+ добавление или формировать полность по новой
    old_address=unicode(dtOldMeter[0][2])
    new_address=old_address.replace(old_meter,new_meter)
    
    if UpdateTable('meters','guid', guidOldMeter, 'name', newName, 'factory_number_manual', new_factory_number_manual,'address', new_address):
        result=u"Счётчик "+unicode(old_meter)+ " успешно заменён на "+unicode(new_meter)
    #print result
    con=0
    for i in range(len(dtTakenParams)):
        dtTakenParams[i]=list(dtTakenParams[i])
        guidTaken=unicode(dtTakenParams[i][1])
        dtLinkAbonentsTakenParams=GetSimpleTable('link_abonents_taken_params','guid_taken_params', guidTaken)
        oldTakenParamName=unicode(dtTakenParams[i][4])
        #newTakenParamName=oldTakenParamName.replace(old_meter,new_meter)
        OldLinkAbonentTakenParamName=unicode(dtLinkAbonentsTakenParams[0][1])
        #newLinkAbonentTakenParamName= OldLinkAbonentTakenParamName.replace(old_meter,new_meter)
        #get_taken_param_by_abonent_from_excel_cfg(instance.name)).name + u" " + instance.guid_params.guid_names_params.name + u" " + instance.guid_params.guid_types_params.name
        #"Квартира 0103 - М-230 21949676"
        if (OldLinkAbonentTakenParamName.find('М-230')):
            typeMeter=u'М-230'
        if (OldLinkAbonentTakenParamName.find('Саяны Комбик')):
            typeMeter=u'Саяны Комбик'  
        newLinkAbonentTakenParamName=OldLinkAbonentTakenParamName.split('-')[0]+ u' - '+ typeMeter +u' ' + unicode(new_meter)
        
        # "М-230 22633939 Меркурий 230 T0 A+ Суточный -- adress: 0  channel: 0"
        #"Саяны Комбик 4443 Саяны Комбик Q Система1 Суточный -- adress: 0  channel: 1"
        n=oldTakenParamName.find(old_meter)
        s=oldTakenParamName[n+len(old_meter):]
        newTakenParamName= typeMeter + u' ' + unicode(new_meter) + s
        #print newTakenParamName
#        print newLinkAbonentTakenParamName
        if UpdateTable('link_abonents_taken_params','guid_taken_params', guidTaken, 'name', newLinkAbonentTakenParamName,"","","","") and UpdateTable('taken_params','guid', guidTaken, 'name',newTakenParamName,"","","",""):
            con+=1
    result+=u"; Изменено связей:"+unicode(con)
    
    return result
    
def UpdateTable(table,whereFieled, whereValue,field1,value1,field2,value2,field3,value3):
    isOk=False
    dt=[]
    cursor = connection.cursor()
    if (field2==""):
        sQuery="""           
     UPDATE %s
     SET  %s='%s'       
     WHERE %s='%s'
     RETURNING * 
   """%(table, field1, value1, whereFieled, whereValue)
    elif (field3==""):
        sQuery="""           
     UPDATE %s
     SET  %s='%s', %s='%s'      
     WHERE %s='%s'
     RETURNING * 
   """%(table, field1, value1,field2,value2,whereFieled, whereValue)
    else:
       sQuery="""           
     UPDATE %s
     SET  %s='%s', %s='%s', %s='%s'       
     WHERE %s='%s'
     RETURNING * 
   """%(table, field1, value1,field2,value2,field3,value3,whereFieled, whereValue)
    print sQuery
    cursor.execute(sQuery)
    dt = cursor.fetchall()
    if len(dt):
        isOk=True   
    return isOk
    
def load_tcp_ip_water_ports_from_excel(sPath, sheet):
    #Добавление tcp_ip портов

    wb = load_workbook(filename = sPath)
    sheet_ranges = wb[sheet]
    row = 3
    IsAdded=False
    result=""
    writeToLog('Load port')
    writeToLog(u'Загрузка портов')
    while (bool(sheet_ranges[u'H%s'%(row)].value)):
        if sheet_ranges[u'H%s'%(row)].value is not None:
            ip_adr=unicode(sheet_ranges[u'H%s'%(row)].value)
            ip_port=unicode(sheet_ranges[u'I%s'%(row)].value)
            
            writeToLog(u'Обрабатываем адрес ' +ip_adr + ip_port)
            
            # проверка есть ли уже такой порт, запрос в БД с адресом и портом, если ответ пустой-добавляем, в противном случае continue
            if not ip_adr or not ip_port or ip_adr==None or ip_port==None: 
                result+=u"Отсутствует значение/я для tcp/ip-порта в строке"+unicode(row)+". Заполните все ячейки excel таблицы."
                break
            else:
                if (checkPortIsExist(ip_adr,ip_port)):
                    add_port=TcpipSettings(ip_address = ip_adr, ip_port =int(ip_port), write_timeout =300 , read_timeout =700 , attempts =3 , delay_between_sending =400)
                    add_port.save()
                    result =u'Новый tcp/ip порт добавлен'
                    IsAdded=True
                else: result+= u'Порт '+unicode(ip_adr)+": "+unicode(ip_port)+u" уже существует"
        writeToLog( result)
        row+=1
    return IsAdded
    
def replace_electric_meters(request):
    args={}

    meter1=u''
    meter2=u''
    change_meter_status=u""
    replace_meter_status=u'НЕ удалось поменять счётчики местами'
    if request.is_ajax():
        if request.method == 'GET':                        
            request.session["meter1"]    = meter1    = request.GET.get('meter1')
            request.session["meter2"]    = meter2   = request.GET.get('meter2')
            
            if (not meter1 or meter1==None or meter2==None or not meter2):
                replace_meter_status=u"Заполните обе ячейки"
            else:                
                replace_meter_status=ReplaceMeters(meter1, meter2)

    args["change_meter_status"]=change_meter_status
    args["replace_meter_status"]=replace_meter_status
    return render_to_response("service/service_change_electric.html", args)
    
def ReplaceMeters(meter1, meter2):
    result=u''
    
    isExistOldMeter=SimpleCheckIfExist('meters','factory_number_manual',meter1,"","","")
    isExistNewMeter=SimpleCheckIfExist('meters','factory_number_manual',meter2,"","","")
    if not isExistOldMeter:
        return u"Номера первого счётчика нет в базе"
    if not isExistNewMeter:
        return u"Номера второго счётчика нет в базе"
        
#  objects.guid as obj_guid,      0 
#  objects.name as obj_name,      1
#  abonents.guid as ab_guid,      2 
#  abonents.name as ab_name,      3
#  link_abonents_taken_params.guid as link_ab_taken_guid,       4
#  link_abonents_taken_params.name as link_ab_taken_name,       5
#  taken_params.guid as taken_guid,       6
#  taken_params.name as taken_name,       7
#  meters.guid as meter_guid,             8
#  meters.name as meter_name,             9
#  meters.address as meter_adr,           10
#  meters.factory_number_manual           11
        
    dtAllTakenMeter1=GetSimpleTable('all_taken_params','factory_number_manual', meter1)
    guidAbonent1=unicode(dtAllTakenMeter1[0][2])
    abName1=unicode(dtAllTakenMeter1[0][3])
    
    dtAllTakenMeter2=GetSimpleTable('all_taken_params','factory_number_manual', meter2)    
    guidAbonent2=unicode(dtAllTakenMeter2[0][2])
    abName2=unicode(dtAllTakenMeter2[0][3])
    
    nameParam1=unicode(dtAllTakenMeter1[0][7])
    nameParam2=unicode(dtAllTakenMeter2[0][7])
        
    typeMeter1=getTypeMeter(nameParam1)
    typeMeter2=getTypeMeter(nameParam2)
    
    if len(typeMeter1)<1 or len(typeMeter2)<1:
        return u'Для этого типа счётчика ещё нет функции обработки'
    if typeMeter1 !=typeMeter2:
        return u'Типы счётчиков не совпадают'
    
    result+=changeConnectionMeterAbonent(dtAllTakenMeter1, typeMeter1, meter1, meter2, guidAbonent2, abName2)
    result+=changeConnectionMeterAbonent(dtAllTakenMeter2, typeMeter1, meter2, meter1, guidAbonent1, abName1) 
           
        
    return result

def getTypeMeter(nameParam1):
    typeMeter1=u''
    if (nameParam1.find('М-230') or nameParam1.find('Меркурий 230')):
        typeMeter1=u'М-230'
    elif (nameParam1.find('Саяны Комбик')):
        typeMeter1=u'Саяны Комбик' 
    elif (nameParam1.find('М-200') or nameParam1.find('Меркурий 200')):
        typeMeter1=u"Меркурий 200"
    return  typeMeter1

def changeConnectionMeterAbonent(dtAllTakenMeter1, typeMeter, meter1, meter2, guidAbonent2, abName2):
    result=u''
    #guidMeter1=unicode(dtAllTakenMeter1[0][8])
    #guidAbonent1=unicode(dtAllTakenMeter1[0][1])
    #meterName1=unicode(dtAllTakenMeter1[0][9])
    #abName1=unicode(dtAllTakenMeter1[0][2])
    con1=0
    for i in range(len(dtAllTakenMeter1)):
        dtAllTakenMeter1[i]=list(dtAllTakenMeter1[i])
        guidParam1=unicode(dtAllTakenMeter1[i][6])   
        nameParam1=unicode(dtAllTakenMeter1[i][7])
        guidLinkAbonentParam1=unicode(dtAllTakenMeter1[i][4])
        #nameLinkAbonentParam1=unicode(dtAllTakenMeter1[i][5])
        
        newTakenParamName1= makeNewTakenParamName(nameParam1, meter1, meter2, typeMeter)
        newLinkAbonentTakenParamName1=makeLinkabonentTakenParamName(abName2,typeMeter,meter2)
        print newTakenParamName1        
        print newLinkAbonentTakenParamName1
        
        isUpdateTakenParam=UpdateTable('taken_params','guid', guidParam1, 'name',newTakenParamName1,"","","","")
        isUpdateLinkAbonTakenParam=UpdateTable('link_abonents_taken_params','guid', guidLinkAbonentParam1, 'guid_abonents', guidAbonent2,'name', newLinkAbonentTakenParamName1,"","")   
        if isUpdateTakenParam and isUpdateLinkAbonTakenParam:
            con1+=1
            print con1
      
    if (con1>0):
        result+=u' Счётчик '+meter1+u' привязан к абоненту '+abName2+'. Изменено привязок: '+unicode(con1)
    else: result+=u' Что-то пошло не так, ни одной привязки не изменено! '+meter1
    return result

def makeLinkabonentTakenParamName(abName,typeMeter,new_meter):
    #"Квартира 0103 - М-230 21949676"   
#LinkAbonentsTakenParams (name = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(instance.name)).name + u" " + instance.guid_params.guid_names_params.name + u" " + instance.guid_params.guid_types_params.name 
    newLinkAbonentTakenParamName=abName+ u' - '+ typeMeter +u' ' + unicode(new_meter)
    return newLinkAbonentTakenParamName

def makeNewTakenParamName(nameParam1, old_meter, new_meter, typeMeter):
    newName=u''

        # "М-230 22633939 Меркурий 230 T0 A+ Суточный -- adress: 0  channel: 0"
        #"Саяны Комбик 4443 Саяны Комбик Q Система1 Суточный -- adress: 0  channel: 1"
    n=nameParam1.find(old_meter)
    s=nameParam1[n+len(old_meter):]
    newName= typeMeter + u' ' + unicode(new_meter) + s
    return newName
    
def get_electric_progruz(request):
    pass

def get_water_progruz(request):
    pass

def get_heat_progruz(request):
    pass

def get_info(request):
    args={}
    

    return render_to_response("service/service_get_info.html", args)
