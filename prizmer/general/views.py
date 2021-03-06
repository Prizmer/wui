# coding -*- coding: utf-8 -*-
from __future__ import unicode_literals
from django.contrib.auth.decorators import login_required
from django.contrib import auth
from django.shortcuts import render_to_response, HttpResponse
from django.core.context_processors import csrf
import simplejson as json
from django.db.models import Max
from django.db import connection
import re
from excel_response import ExcelResponse
import datetime
#---------
import calendar
import common_sql
from django.shortcuts import redirect
#---------

from general.models import Objects, Abonents, BalanceGroups, Meters, LinkBalanceGroupsMeters

def dictfetchall(cursor):
#"Returns all rows from a cursor as a dict"
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]
    
#!!!!!!!!!!!!
def simple_query(): # Пример запроса в БД на чистом SQL
    simpleq = connection.cursor()
    simpleq.execute("""SELECT 
                                 monthly_values.id, 
                                 monthly_values.date
                               FROM 
                                 public.monthly_values;""")
    simpleq = simpleq.fetchall()
    return simpleq
#!!!!!!!!!!!!    


#------------------------------------------------------------------------------------------------------------------------



    

# Отчет по СПГ на начало суток
def get_data_table_by_date_spg(obj_title, obj_parent_title, electric_data):
    data_table = []
    
    my_parametr = "Время работы узла учёта"    
    data_table_time = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, u"СПГ762__")
    
    my_parametr = "Время работы при ненулевом расходе"    
    data_table_time_rashod = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, u"СПГ762__")
    
    my_parametr = "Атмосферное давление"    
    data_table_p_atm = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, u"СПГ762__")

    my_parametr = "Температура наружного воздуха"    
    data_table_temp_air = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, u"СПГ762__")
    
    my_parametr = "Значение времени интегрирования"    
    data_table_time_integr = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, u"СПГ762__")
    
    my_parametr = "Среднее значение расхода газа"    
    data_table_sr_rashod = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, u"СПГ762__")
    
    my_parametr = "Среднее значение температуры газа"    
    data_table_sr_temp_air = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, u"СПГ762__")
    
    my_parametr = "Среднее значение абсолютного давления"    
    data_table_sr_abs_p = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, u"СПГ762__")
    
    my_parametr = "Среднее значение с доп. датчика 1"    
    data_table_sr_dop_1 = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, u"СПГ762__")
    
    my_parametr = "Среднее значение с доп. датчика 2"    
    data_table_sr_dop_2 = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, u"СПГ762__")
    
    my_parametr = "Масса газа при стандартных условиях"    
    data_table_m_gas_standart = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, u"СПГ762__")
    
    my_parametr = "Объем газа при стандартных условиях"    
    data_table_v_gas_standart = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, u"СПГ762__")
    
    my_parametr = "Объем газа при рабочих условиях"    
    data_table_v_work = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, u"СПГ762__")
    
    my_parametr = "Обобщённое сообщение о нештатных ситуациях"    
    data_table_err_common = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, u"СПГ762__")
    
              
    for x in range(len(data_table_time)):
        data_table_temp = []
        try:
            data_table_temp.append(data_table_time[x][0]) # дата
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_time[x][2]) # имя абонента
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_time[x][3]) # заводской номер
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_time[x][4]) # Время работы узла учёта
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_time_rashod[x][4]) # Время работы при ненулевом расходе
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_p_atm[x][4]) # Атмосферное давление
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")            
        try:
            data_table_temp.append(data_table_temp_air[x][4]) # Температура наружного воздуха
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")            
        try:
            data_table_temp.append(data_table_time_integr[x][4]) # Значение времени интегрирования
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")            
        try:
            data_table_temp.append(data_table_sr_rashod[x][4]) # Среднее значение расхода газа
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_sr_temp_air[x][4]) # Среднее значение температуры газа
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_sr_abs_p[x][4]) # Среднее значение абсолютного давления
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")            
        try:
            data_table_temp.append(data_table_sr_dop_1[x][4]) # Среднее значение с доп. датчика 1
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")            
        try:
            data_table_temp.append(data_table_sr_dop_2[x][4]) # Среднее значение с доп. датчика 2
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_m_gas_standart[x][4]) # Масса газа при стандартных условиях
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_v_gas_standart[x][4]) # Объем газа при стандартных условиях
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")            
        try:
            data_table_temp.append(data_table_v_work[x][4]) # Объем газа при рабочих условиях
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")            
        try:
            data_table_temp.append(data_table_err_common[x][4]) # Обобщённое сообщение о нештатных ситуациях
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
            

        data_table.append(data_table_temp)
    return data_table


    

#------------    


#-------------------------------------------------------------------------------------------------------------------------
    
def get_data_table_by_date_monthly(obj_title, obj_parent_title, electric_data):
    data_table = []
    my_parametr = "T0 A+"
    cursor_t0_aplus = connection.cursor()
    cursor_t0_aplus.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t0_aplus = cursor_t0_aplus.fetchall()
    
    my_parametr = "T1 A+"            
    cursor_t1_aplus = connection.cursor()
    cursor_t1_aplus.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t1_aplus = cursor_t1_aplus.fetchall()

    my_parametr = "T2 A+"                
    cursor_t2_aplus = connection.cursor()
    cursor_t2_aplus.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t2_aplus = cursor_t2_aplus.fetchall()
                
    my_parametr = "T3 A+"
    cursor_t3_aplus = connection.cursor()
    cursor_t3_aplus.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t3_aplus = cursor_t3_aplus.fetchall()

    my_parametr = "T4 A+"                
    cursor_t4_aplus = connection.cursor()
    cursor_t4_aplus.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t4_aplus = cursor_t4_aplus.fetchall()
    
    my_parametr = "T0 R+"
    cursor_t0_rplus = connection.cursor()
    cursor_t0_rplus.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t0_rplus = cursor_t0_rplus.fetchall()
                
    for x in range(len(data_table_t0_aplus)):
        data_table_temp = []
        try:
            data_table_temp.append(data_table_t0_aplus[x][0])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_t0_aplus[x][2])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_t0_aplus[x][6])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_t0_aplus[x][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_t1_aplus[x][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_t2_aplus[x][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_t3_aplus[x][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_t4_aplus[x][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_t0_rplus[x][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        data_table.append(data_table_temp) 
    
    return data_table
#!!!!!!!!!!!!
    
def get_data_table_by_date_daily(obj_title, obj_parent_title, electric_data):
    data_table = []
    my_parametr = "T0 A+"
    cursor_t0_aplus = connection.cursor()
    cursor_t0_aplus.execute("""SELECT 
                                daily_values.date, 
                                daily_values.value, 
                                abonents.name, 
                                daily_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.daily_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = daily_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                daily_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t0_aplus = cursor_t0_aplus.fetchall()
    
    my_parametr = "T1 A+"            
    cursor_t1_aplus = connection.cursor()
    cursor_t1_aplus.execute("""SELECT 
                                daily_values.date, 
                                daily_values.value, 
                                abonents.name, 
                                daily_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.daily_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = daily_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                daily_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t1_aplus = cursor_t1_aplus.fetchall()

    my_parametr = "T2 A+"                
    cursor_t2_aplus = connection.cursor()
    cursor_t2_aplus.execute("""SELECT 
                                daily_values.date, 
                                daily_values.value, 
                                abonents.name, 
                                daily_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.daily_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = daily_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                daily_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t2_aplus = cursor_t2_aplus.fetchall()
                
    my_parametr = "T3 A+"
    cursor_t3_aplus = connection.cursor()
    cursor_t3_aplus.execute("""SELECT 
                                daily_values.date, 
                                daily_values.value, 
                                abonents.name, 
                                daily_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.daily_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = daily_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                daily_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t3_aplus = cursor_t3_aplus.fetchall()

    my_parametr = "T4 A+"                
    cursor_t4_aplus = connection.cursor()
    cursor_t4_aplus.execute("""SELECT 
                                daily_values.date, 
                                daily_values.value, 
                                abonents.name, 
                                daily_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.daily_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = daily_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                daily_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t4_aplus = cursor_t4_aplus.fetchall()
    
    my_parametr = "T0 R+"
    cursor_t0_rplus = connection.cursor()
    cursor_t0_rplus.execute("""SELECT 
                                daily_values.date, 
                                daily_values.value, 
                                abonents.name, 
                                daily_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.daily_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = daily_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                daily_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t0_rplus = cursor_t0_rplus.fetchall()
                
    for x in range(len(data_table_t0_aplus)):
        data_table_temp = []
        try:
            data_table_temp.append(data_table_t0_aplus[x][0])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_t0_aplus[x][2])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_t0_aplus[x][6])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_t0_aplus[x][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_t1_aplus[x][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_t2_aplus[x][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_t3_aplus[x][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(data_table_t4_aplus[x][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        
        try:
            data_table_temp.append(data_table_t0_rplus[x][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        data_table.append(data_table_temp)
    
    return data_table

# Create your views here.
@login_required(login_url='/auth/login/') 
def default(request):
    args={}
    args.update(csrf(request))
    
    
    #-------------- get data new tree
    max_level = Objects.objects.aggregate(Max('level'))['level__max'] # Max number of levels
    if max_level < 3:
        all_level_0 = Objects.objects.filter(level=0)
        tree_data = []
        for l0 in range(len(all_level_0)):
            filter_level_1 = Objects.objects.filter(level=1).filter(guid_parent = all_level_0[l0].guid)
            children_data_l1 = []
            for l1 in range(len(filter_level_1)):
                children_data_l2 = []
                filter_level_2 = Objects.objects.filter(level=2).filter(guid_parent = filter_level_1[l1].guid)
                for l2 in range(len(filter_level_2)):
                    abonents_data = []
                    list_of_level_2 = {"key":u"level2-"+str(l2), "title": filter_level_2[l2].name, "children":abonents_data}
                    filter_level_abonents = Abonents.objects.filter(guid_objects = filter_level_2[l2].guid)
                    for l3 in range(len(filter_level_abonents)):
                        meters_data = []
                        cursor = connection.cursor()
                        cursor.execute("""SELECT 
                                      abonents.name, 
                                      meters.name, 
                                      meters.factory_number_manual
                                    FROM 
                                      public.abonents, 
                                      public.meters, 
                                      public.taken_params, 
                                      public.link_abonents_taken_params
                                    WHERE 
                                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                                      meters.guid = taken_params.guid_meters AND
                                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                      abonents.name = %s
                                    GROUP BY
                                      abonents.name,
                                      meters.name, 
                                      meters.factory_number_manual;""", [filter_level_abonents[l3].name])
                        filter_level_meters = dictfetchall(cursor)
                        for meter in range(len(filter_level_meters)):
                           list_of_level_meters = {"key":u"meter-"+str(meter), "title": filter_level_meters[meter]['factory_number_manual']}
                           meters_data.append(list_of_level_meters)
                        list_of_level_abonents = {"key":u"abonent-"+str(l2), "title": filter_level_abonents[l3].name, "children":meters_data}
                        abonents_data.append(list_of_level_abonents)
                    list_of_level_2 = {"key":u"level2-"+str(l2), "title": filter_level_2[l2].name, "children":abonents_data}                     
                    children_data_l2.append(list_of_level_2)             
                list_of_level_1 = {"key":u"level1-"+str(l1), "title": filter_level_1[l1].name, "children":children_data_l2, "folder":bool(children_data_l2)}
                children_data_l1.append(list_of_level_1)
            list_of_level_0 = {"key":u"level0-"+str(l0), "title": all_level_0[l0].name, "children":children_data_l1, "folder":bool(children_data_l1)}
            tree_data.append(list_of_level_0)
        tree_data_json = json.dumps(tree_data, )
        args['tree_data_json'] = tree_data_json
    else:
        pass
    #-------------- get data new tree end
    args['ico_url_electric'] = "/static/images/electric-ico36.png"
    args['ico_url_water'] = "/static/images/water-ico36.png"
    args['ico_url_heat'] = "/static/images/heat-ico36.png"    
    args['ico_url_gas'] = "/static/images/gas-ico36.png"
    args['ico_url_economic'] = "/static/images/economic-ico36.png"
    return render_to_response('base.html', args)

def go_out(request):
    auth.logout(request)

    return redirect(default)
    
@login_required(login_url='/auth/login/') 
def tree_data_json(request):
    args={}
    args.update(csrf(request))
    
    
    #-------------- get data new tree
    max_level = Objects.objects.aggregate(Max('level'))['level__max'] # Max number of levels
    if max_level < 3:
        all_level_0 = Objects.objects.filter(level=0)
        tree_data = []
        for l0 in range(len(all_level_0)):
            filter_level_1 = Objects.objects.filter(level=1).filter(guid_parent = all_level_0[l0].guid)
            children_data_l1 = []
            for l1 in range(len(filter_level_1)):
                children_data_l2 = []
                filter_level_2 = Objects.objects.filter(level=2).filter(guid_parent = filter_level_1[l1].guid)
                for l2 in range(len(filter_level_2)):
                    abonents_data = []
                    list_of_level_2 = {"key":u"level2-"+str(l2), "title": filter_level_2[l2].name, "children":abonents_data}
                    filter_level_abonents = Abonents.objects.filter(guid_objects = filter_level_2[l2].guid).order_by('name')
                    for l3 in range(len(filter_level_abonents)):
                        meters_data = []
                        cursor = connection.cursor()
                        cursor.execute("""SELECT 
                                          abonents.name, 
                                          meters.name, 
                                          meters.factory_number_manual
                                        FROM 
                                          public.abonents, 
                                          public.meters, 
                                          public.taken_params, 
                                          public.link_abonents_taken_params, 
                                          public.objects
                                        WHERE 
                                          abonents.guid = link_abonents_taken_params.guid_abonents AND
                                          meters.guid = taken_params.guid_meters AND
                                          taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                          objects.guid = abonents.guid_objects AND
                                          abonents.name = %s AND 
                                          objects.name = %s
                                        GROUP BY
                                          abonents.name,
                                          meters.name, 
                                          meters.factory_number_manual;""", [filter_level_abonents[l3].name, filter_level_2[l2].name ])
                        filter_level_meters = dictfetchall(cursor)
                        for meter in range(len(filter_level_meters)):
                           list_of_level_meters = {"key":u"meter-"+str(meter), "title": filter_level_meters[meter]['factory_number_manual']}
                           meters_data.append(list_of_level_meters)
                        list_of_level_abonents = {"key":u"abonent-"+str(l2), "title": filter_level_abonents[l3].name, "children":meters_data}
                        abonents_data.append(list_of_level_abonents)
                    list_of_level_2 = {"key":u"level2-"+str(l2), "title": filter_level_2[l2].name, "children":abonents_data}                     
                    children_data_l2.append(list_of_level_2)             
                list_of_level_1 = {"key":u"level1-"+str(l1), "title": filter_level_1[l1].name, "children":children_data_l2, "folder":bool(children_data_l2)}
                children_data_l1.append(list_of_level_1)
            list_of_level_0 = {"key":u"level0-"+str(l0), "title": all_level_0[l0].name, "children":children_data_l1, "folder":bool(children_data_l1)}
            tree_data.append(list_of_level_0)
            
        # Получаем информацию по балансным группам
            balance_groups_list = []
            simpleq = connection.cursor()
            simpleq.execute(""" SELECT 
                                  balance_groups.name
                                FROM 
                                  public.balance_groups;""")
            simpleq = simpleq.fetchall()
            for x in range (len(simpleq)):
                balance_groups_list.append({"key": u"group-"+str(x), "title": simpleq[x][0]})
             
        # Получаем информацию по группам 80020
            groups_80020_list = []
            simpleq = connection.cursor()
            simpleq.execute(""" SELECT 
                                  groups_80020.name
                                FROM 
                                  public.groups_80020;
                                """)
            simpleq = simpleq.fetchall()
            for x in range (len(simpleq)):
                groups_80020_list.append({"key": u"group80020-"+str(x), "title": simpleq[x][0]})
        
        tree_data.append({"key": u"group" + str(1000), "title": u"Группы 80020", "children":groups_80020_list , "folder":bool(True)})
        tree_data.append({"key": u"group" + str(1000), "title": u"Группы", "children":balance_groups_list , "folder":bool(True)})
        
        # Создаем json данные для дерева объектов
        tree_data_json = json.dumps(tree_data, )
        args['tree_data_json'] = tree_data_json
    else:
        pass
    return HttpResponse(tree_data_json)
    #-------------- get data new tree end
    
def get_object_title(request):
    if request.is_ajax():
        if request.method == 'GET':
            object_title = request.GET['object_title']
        elif request.method == 'POST':
            object_title = u'Не выбран'
    else:
        object_title = u'Не выбран'
    return HttpResponse(object_title)

    
def get_object_key(request):
    if request.is_ajax():
        if request.method == 'GET':
            object_key = request.GET['object_key']
        elif request.method == 'POST':
            object_key = u'Не выбран'
    else:
        object_key = u'Не выбран'
    return HttpResponse(object_key)

@login_required(login_url='/auth/login/')     
def get_data_table(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = get_data_table_by_date_monthly(obj_title, obj_parent_title, electric_data_end)
#                request.session["data_table_export"] = data_table
                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = get_data_table_by_date_daily(obj_title, obj_parent_title, electric_data_end)

#                request.session["data_table_export"] = data_table  ! Check. Not Working

            elif (is_electric_current == "1") & (bool(is_abonent_level.search(obj_key))):
                pass
            elif (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                cursor_t0_aplus_delta_start = connection.cursor()
                cursor_t0_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t0_aplus_delta_start = cursor_t0_aplus_delta_start.fetchall()
                
                cursor_t1_aplus_delta_start = connection.cursor()
                cursor_t1_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t1_aplus_delta_start = cursor_t1_aplus_delta_start.fetchall()
                
                cursor_t2_aplus_delta_start = connection.cursor()
                cursor_t2_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t2_aplus_delta_start = cursor_t2_aplus_delta_start.fetchall()
                
                cursor_t3_aplus_delta_start = connection.cursor()
                cursor_t3_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t3_aplus_delta_start = cursor_t3_aplus_delta_start.fetchall() 
                
                cursor_t4_aplus_delta_start = connection.cursor()
                cursor_t4_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t4_aplus_delta_start = cursor_t4_aplus_delta_start.fetchall()

                cursor_t0_aplus_delta_end = connection.cursor()
                cursor_t0_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t0_aplus_delta_end = cursor_t0_aplus_delta_end.fetchall()
                
                cursor_t1_aplus_delta_end = connection.cursor()
                cursor_t1_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t1_aplus_delta_end = cursor_t1_aplus_delta_end.fetchall()
                
                cursor_t2_aplus_delta_end = connection.cursor()
                cursor_t2_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t2_aplus_delta_end = cursor_t2_aplus_delta_end.fetchall()
                
                cursor_t3_aplus_delta_end = connection.cursor()
                cursor_t3_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t3_aplus_delta_end = cursor_t3_aplus_delta_end.fetchall() 
                
                cursor_t4_aplus_delta_end = connection.cursor()
                cursor_t4_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t4_aplus_delta_end = cursor_t4_aplus_delta_end.fetchall()
#                data_table = []
                for x in range(len(data_table_t0_aplus_delta_end)):
                    data_table_temp = []

                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][2])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][1] - data_table_t0_aplus_delta_start[x][1] )
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end[x][1] - data_table_t1_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end[x][1] - data_table_t2_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end[x][1] - data_table_t3_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end[x][1] - data_table_t4_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
                
            elif (is_electric_period == "1") & (is_electric_daily =="1") & (bool(is_abonent_level.search(obj_key))): # pokazaniya za period

                end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
                start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
                dates = [x for x in common_sql.daterange(start_date,
                              end_date,
                              step=datetime.timedelta(days=1),
                              inclusive=True)]
                '''for x in range(len(dates)):
                    data_table_temp = [dates[x], dates[x], datetime.datetime.strftime(dates[x], "%d.%m.%Y")]
                    data_table.append(data_table_temp)'''

                for x in range(len(dates)):
                    data_table_temp = get_data_table_by_date_daily(obj_title, obj_parent_title, datetime.datetime.strftime(dates[x], "%d.%m.%Y"))
                    data_table.extend(data_table_temp)
                #------------

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # monthly for abonents group
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    # query data for each abonent
                    cursor_t0_aplus_monthly_temp = connection.cursor()
                    cursor_t0_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T0 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_monthly_temp = cursor_t0_aplus_monthly_temp.fetchall()
                    
                    cursor_t1_aplus_monthly_temp = connection.cursor()
                    cursor_t1_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T1 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_monthly_temp = cursor_t1_aplus_monthly_temp.fetchall()
                    
                    cursor_t2_aplus_monthly_temp = connection.cursor()
                    cursor_t2_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date,
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T2 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY 
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_monthly_temp = cursor_t2_aplus_monthly_temp.fetchall()
                    
                    cursor_t3_aplus_monthly_temp = connection.cursor()
                    cursor_t3_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T3 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_monthly_temp = cursor_t3_aplus_monthly_temp.fetchall()
                
                    cursor_t4_aplus_monthly_temp = connection.cursor()
                    cursor_t4_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T4 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_monthly_temp = cursor_t4_aplus_monthly_temp.fetchall()
                    
                    cursor_t0_rplus_monthly_temp = connection.cursor()
                    cursor_t0_rplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T0 R+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_rplus_monthly_temp = cursor_t0_rplus_monthly_temp.fetchall()
                
                    data_table_temp = []
                    try:
                        data_table_temp.append(electric_data_end)
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    data_table_temp.append(abonents_list[x][0])
                    
                    try:
                        data_table_temp.append(data_table_t0_aplus_monthly_temp[0][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t0_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:    
                        data_table_temp.append(data_table_t1_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t2_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_rplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for abonents group
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()                              
#                data_table = []
                for x in range(len(abonents_list)):
                    cursor_t0_aplus_daily_temp = connection.cursor()
                    cursor_t0_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_daily_temp = cursor_t0_aplus_daily_temp.fetchall()
                
                    cursor_t1_aplus_daily_temp = connection.cursor()
                    cursor_t1_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_daily_temp = cursor_t1_aplus_daily_temp.fetchall()
                
                    cursor_t2_aplus_daily_temp = connection.cursor()
                    cursor_t2_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_daily_temp = cursor_t2_aplus_daily_temp.fetchall()
                
                    cursor_t3_aplus_daily_temp = connection.cursor()
                    cursor_t3_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_daily_temp = cursor_t3_aplus_daily_temp.fetchall() 
                
                    cursor_t4_aplus_daily_temp = connection.cursor()
                    cursor_t4_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_daily_temp = cursor_t4_aplus_daily_temp.fetchall()
                    
                    data_table_temp = []
                    try:
                        data_table_temp.append(electric_data_end)
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    data_table_temp.append(abonents_list[x][0])
                    try:
                        data_table_temp.append(data_table_t0_aplus_daily_temp[0][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
            elif (is_electric_current == '1') & (bool(is_object_level.search(obj_key))):
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                              ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    data_table_temp = []
                    data_table_temp.append(u'Дата')
                    data_table_temp.append(abonents_list[x][0])
                    data_table_temp.append(u'Какой-то заводской номер')
                    data_table_temp.append(0)
                    data_table_temp.append(100)
                    data_table_temp.append(200)
                    data_table_temp.append(300)
                    data_table_temp.append(400)
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
                   
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                              ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    # delta for groups abonents 'start date'
                    cursor_t0_aplus_delta_start_temp = connection.cursor()
                    cursor_t0_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t0_aplus_delta_start_temp = cursor_t0_aplus_delta_start_temp.fetchall()
                
                    cursor_t1_aplus_delta_start_temp = connection.cursor()
                    cursor_t1_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t1_aplus_delta_start_temp = cursor_t1_aplus_delta_start_temp.fetchall()
                
                    cursor_t2_aplus_delta_start_temp = connection.cursor()
                    cursor_t2_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t2_aplus_delta_start_temp = cursor_t2_aplus_delta_start_temp.fetchall()
                
                    cursor_t3_aplus_delta_start_temp = connection.cursor()
                    cursor_t3_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t3_aplus_delta_start_temp = cursor_t3_aplus_delta_start_temp.fetchall() 
                
                    cursor_t4_aplus_delta_start_temp = connection.cursor()
                    cursor_t4_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t4_aplus_delta_start_temp = cursor_t4_aplus_delta_start_temp.fetchall()
                    
                    # delta for groups abonents 'end date'
                    cursor_t0_aplus_delta_end_temp = connection.cursor()
                    cursor_t0_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_delta_end_temp = cursor_t0_aplus_delta_end_temp.fetchall()
                
                    cursor_t1_aplus_delta_end_temp = connection.cursor()
                    cursor_t1_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_delta_end_temp = cursor_t1_aplus_delta_end_temp.fetchall()
                
                    cursor_t2_aplus_delta_end_temp = connection.cursor()
                    cursor_t2_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_delta_end_temp = cursor_t2_aplus_delta_end_temp.fetchall()
                
                    cursor_t3_aplus_delta_end_temp = connection.cursor()
                    cursor_t3_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_delta_end_temp = cursor_t3_aplus_delta_end_temp.fetchall() 
                
                    cursor_t4_aplus_delta_end_temp = connection.cursor()
                    cursor_t4_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_delta_end_temp = cursor_t4_aplus_delta_end_temp.fetchall()
                    
                    data_table_temp = []
                    data_table_temp.append(abonents_list[x][0])
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1] - data_table_t0_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end_temp[0][1] - data_table_t1_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end_temp[0][1] - data_table_t2_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end_temp[0][1] - data_table_t3_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end_temp[0][1] - data_table_t4_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                           
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table                 
#*********************************************************************************************************************************************************************          
            else:
                pass

        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render_to_response("data_table.html", args)

def export_excel_electric(request):
    data_table = request.session["data_table_export"]
    return ExcelResponse(data_table, 'report')
    


@login_required(login_url='/auth/login/')  
def electric(request):
    args={}
    args['ico_url_electric'] = "/static/images/electric-ico42.png"
    args['ico_url_water'] = "/static/images/water-ico36.png"
    args['ico_url_heat'] = "/static/images/heat-ico36.png"    
    args['ico_url_gas'] = "/static/images/gas-ico36.png"
    args['ico_url_economic'] = "/static/images/economic-ico36.png"
    return render_to_response('control.html', args)

@login_required(login_url='/auth/login/')
def economic(request):
    args={}
    args['ico_url_electric'] = "/static/images/electric-ico36.png"
    args['ico_url_water'] = "/static/images/water-ico36.png"
    args['ico_url_heat'] = "/static/images/heat-ico36.png"    
    args['ico_url_gas'] = "/static/images/gas-ico36.png"
    args['ico_url_economic'] = "/static/images/economic-ico42.png"
    return render_to_response('economic.html', args)

@login_required(login_url='/auth/login/')    
def water(request):
    args={}
    args['ico_url_electric'] = "/static/images/electric-ico36.png"
    args['ico_url_water'] = "/static/images/water-ico42.png"
    args['ico_url_heat'] = "/static/images/heat-ico36.png"    
    args['ico_url_gas'] = "/static/images/gas-ico36.png"
    args['ico_url_economic'] = "/static/images/economic-ico36.png"
    return render_to_response('water.html', args)
    
@login_required(login_url='/auth/login/')    
def heat(request):
    args={}
    args['ico_url_electric'] = "/static/images/electric-ico36.png"
    args['ico_url_water'] = "/static/images/water-ico36.png"
    args['ico_url_heat'] = "/static/images/heat-ico42.png"    
    args['ico_url_gas'] = "/static/images/gas-ico36.png"
    args['ico_url_economic'] = "/static/images/economic-ico36.png"
    return render_to_response('heat.html', args)
    
@login_required(login_url='/auth/login/')    
def gas(request):
    args={}
    args['ico_url_electric'] = "/static/images/electric-ico36.png"
    args['ico_url_water'] = "/static/images/water-ico36.png"
    args['ico_url_heat'] = "/static/images/heat-ico36.png"    
    args['ico_url_gas'] = "/static/images/gas-ico42.png"
    args['ico_url_economic'] = "/static/images/economic-ico36.png"
    return render_to_response('gas.html', args)
      
   
    
# образец выгрузги экселя -------------------------------------------------------------------------------------------    
def test_xlsx(request):
    import StringIO
    response = StringIO.StringIO()
    from openpyxl import Workbook
    from openpyxl.compat import range
    from openpyxl.cell import get_column_letter
    
    wb = Workbook()
    

    ws = wb.active
    
    ws.title = "range names"
    
    for col_idx in range(1,40):
        col = get_column_letter(col_idx)
        for row in range(1,600):
            ws.cell('%s%s'%(col,row)).value = '%s%s' % (col, row)
            
    ws = wb.create_sheet()
    
    ws.title = 'Pi'
    
    ws['F5'] = 3.14
    
    wb.save(response)

    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    response['Content-Disposition'] = "attachment; filename=test.xlsx"

    return response
#--------------------------------------------------------------------------------------------------------------------
def choose_report(request):
    return render_to_response("choose_report.html")

def report_2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = get_data_table_by_date_monthly(obj_title, obj_parent_title, electric_data_end)

                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = get_data_table_by_date_daily(obj_title, obj_parent_title, electric_data_end)


            elif (is_electric_current == "1") & (bool(is_abonent_level.search(obj_key))):
                pass
                            
            elif (is_electric_period == "1") & (is_electric_daily =="1") & (bool(is_abonent_level.search(obj_key))): # pokazaniya za period

                end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
                start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
                dates = [x for x in common_sql.daterange(start_date,
                              end_date,
                              step=datetime.timedelta(days=1),
                              inclusive=True)]
                '''for x in range(len(dates)):
                    data_table_temp = [dates[x], dates[x], datetime.datetime.strftime(dates[x], "%d.%m.%Y")]
                    data_table.append(data_table_temp)'''

                for x in range(len(dates)):
                    data_table_temp = get_data_table_by_date_daily(obj_title, obj_parent_title, datetime.datetime.strftime(dates[x], "%d.%m.%Y"))
                    data_table.extend(data_table_temp)
                #------------

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # monthly for abonents group
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    # query data for each abonent
                    cursor_t0_aplus_monthly_temp = connection.cursor()
                    cursor_t0_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T0 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_monthly_temp = cursor_t0_aplus_monthly_temp.fetchall()
                    
                    cursor_t1_aplus_monthly_temp = connection.cursor()
                    cursor_t1_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T1 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_monthly_temp = cursor_t1_aplus_monthly_temp.fetchall()
                    
                    cursor_t2_aplus_monthly_temp = connection.cursor()
                    cursor_t2_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date,
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T2 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY 
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_monthly_temp = cursor_t2_aplus_monthly_temp.fetchall()
                    
                    cursor_t3_aplus_monthly_temp = connection.cursor()
                    cursor_t3_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T3 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_monthly_temp = cursor_t3_aplus_monthly_temp.fetchall()
                
                    cursor_t4_aplus_monthly_temp = connection.cursor()
                    cursor_t4_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T4 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_monthly_temp = cursor_t4_aplus_monthly_temp.fetchall()
                    
                    cursor_t0_rplus_monthly_temp = connection.cursor()
                    cursor_t0_rplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T0 R+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_rplus_monthly_temp = cursor_t0_rplus_monthly_temp.fetchall()
                
                    data_table_temp = []
                    try:
                        data_table_temp.append(electric_data_end)
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    data_table_temp.append(abonents_list[x][0])
                    
                    try:
                        data_table_temp.append(data_table_t0_aplus_monthly_temp[0][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t0_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:    
                        data_table_temp.append(data_table_t1_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t2_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_rplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1'):# & (bool(is_object_level.search(obj_key))): # daily for abonents group
               
                    
                if (bool(is_object_level.search(obj_key))):
                    cursor_abonents_list = connection.cursor()
                    cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title])
                    abonents_list = cursor_abonents_list.fetchall()

                    
                elif (bool(is_group_level.search(obj_key))):
                    cursor_abonents_list = connection.cursor()
                    cursor_abonents_list.execute("""
                                                SELECT 
                                                  meters.name,
                                                  link_balance_groups_meters.type
                                                FROM 
                                                  public.meters, 
                                                  public.link_balance_groups_meters, 
                                                  public.balance_groups
                                                WHERE 
                                                  link_balance_groups_meters.guid_balance_groups = balance_groups.guid AND
                                                  link_balance_groups_meters.guid_meters = meters.guid AND
                                                  balance_groups.name = %s
                                                ORDER BY
                                                  meters.name ASC;""",[obj_title])
                    abonents_list = cursor_abonents_list.fetchall()
                    obj_title=u'Завод'
                else:
                    abonents_list = [12345678]
                              

                for x in range(len(abonents_list)):
                    cursor_t0_aplus_daily_temp = connection.cursor()
                    cursor_t0_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_daily_temp = cursor_t0_aplus_daily_temp.fetchall()
                
                    cursor_t1_aplus_daily_temp = connection.cursor()
                    cursor_t1_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_daily_temp = cursor_t1_aplus_daily_temp.fetchall()
                
                    cursor_t2_aplus_daily_temp = connection.cursor()
                    cursor_t2_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_daily_temp = cursor_t2_aplus_daily_temp.fetchall()
                
                    cursor_t3_aplus_daily_temp = connection.cursor()
                    cursor_t3_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_daily_temp = cursor_t3_aplus_daily_temp.fetchall() 
                
                    cursor_t4_aplus_daily_temp = connection.cursor()
                    cursor_t4_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_daily_temp = cursor_t4_aplus_daily_temp.fetchall()
                    
                    cursor_t0_rplus_daily_temp = connection.cursor()
                    cursor_t0_rplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_rplus_daily_temp = cursor_t0_rplus_daily_temp.fetchall()
                    
                    data_table_temp = []
                    try:
                        data_table_temp.append(electric_data_end)
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    data_table_temp.append(abonents_list[x][0])
                    try:
                        data_table_temp.append(data_table_t0_aplus_daily_temp[0][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        if (bool(is_group_level.search(obj_key))):                           
                            if abonents_list[x][1]: # Если абонент входит в группу со знаком плюс, то показания как есть
                                data_table_temp.append(data_table_t0_aplus_daily_temp[0][1])
                            else:                   # Если абонент входит в группу со знаком минус, то показазния инвертируются
                                data_table_temp.append(-data_table_t0_aplus_daily_temp[0][1])
                        else:
                           data_table_temp.append(data_table_t0_aplus_daily_temp[0][1]) 
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        if (bool(is_group_level.search(obj_key))):                                           
                            if abonents_list[x][1]: # Если абонент входит в группу со знаком плюс, то показания как есть
                                data_table_temp.append(data_table_t0_rplus_daily_temp[0][1])
                            else:
                                data_table_temp.append(-data_table_t0_rplus_daily_temp[0][1])
                        else:
                            data_table_temp.append(data_table_t0_rplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    data_table.append(data_table_temp)
                if (bool(is_group_level.search(obj_key))):  # Если это группа добавляем еще одну строку с суммой показаний
                    sum_a_plus = 0
                    sum_r_plus = 0
                    for x in range(len(data_table)):
                        try:
                            sum_a_plus = sum_a_plus + data_table[x][3]
                            sum_r_plus = sum_r_plus + data_table[x][8]
                        except:
                            next
                    data_table.append([])
                    data_table.append([u' ',u' ',u'<strong>Сумма</strong>',sum_a_plus,u'-',u'-',u'-',u'-',sum_r_plus])

                request.session["data_table_export"] = data_table
            elif (is_electric_current == '1') & (bool(is_object_level.search(obj_key))): # текущие для объекта учёта
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                              ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    data_table_temp = []
                    data_table_temp.append(u'Дата')
                    data_table_temp.append(abonents_list[x][0])
                    data_table_temp.append(u'Какой-то заводской номер')
                    data_table_temp.append(0)
                    data_table_temp.append(100)
                    data_table_temp.append(200)
                    data_table_temp.append(300)
                    data_table_temp.append(400)
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render_to_response("data_table/2.html", args)
    
def data_table_3_tarifa_k(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u'1'
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                cursor_t0_aplus_delta_start = connection.cursor()
                cursor_t0_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t0_aplus_delta_start = cursor_t0_aplus_delta_start.fetchall()
                
                cursor_t1_aplus_delta_start = connection.cursor()
                cursor_t1_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t1_aplus_delta_start = cursor_t1_aplus_delta_start.fetchall()
                
                cursor_t2_aplus_delta_start = connection.cursor()
                cursor_t2_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t2_aplus_delta_start = cursor_t2_aplus_delta_start.fetchall()
                
                cursor_t3_aplus_delta_start = connection.cursor()
                cursor_t3_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t3_aplus_delta_start = cursor_t3_aplus_delta_start.fetchall() 
                
                cursor_t4_aplus_delta_start = connection.cursor()
                cursor_t4_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t4_aplus_delta_start = cursor_t4_aplus_delta_start.fetchall()

                cursor_t0_aplus_delta_end = connection.cursor()
                cursor_t0_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t0_aplus_delta_end = cursor_t0_aplus_delta_end.fetchall()
                
                cursor_t1_aplus_delta_end = connection.cursor()
                cursor_t1_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t1_aplus_delta_end = cursor_t1_aplus_delta_end.fetchall()
                
                cursor_t2_aplus_delta_end = connection.cursor()
                cursor_t2_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t2_aplus_delta_end = cursor_t2_aplus_delta_end.fetchall()
                
                cursor_t3_aplus_delta_end = connection.cursor()
                cursor_t3_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t3_aplus_delta_end = cursor_t3_aplus_delta_end.fetchall() 
                
                cursor_t4_aplus_delta_end = connection.cursor()
                cursor_t4_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t4_aplus_delta_end = cursor_t4_aplus_delta_end.fetchall()
                
                cursor_t0_rplus_delta_start = connection.cursor()
                cursor_t0_rplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t0_rplus_delta_start = cursor_t0_rplus_delta_start.fetchall()
                
                cursor_t0_rplus_delta_end = connection.cursor()
                cursor_t0_rplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t0_rplus_delta_end = cursor_t0_rplus_delta_end.fetchall()
                
                
#                data_table = []
                for x in range(len(data_table_t0_aplus_delta_end)):
                    data_table_temp = []

                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][2])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][1] - data_table_t0_aplus_delta_start[x][1] )
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end[x][1] - data_table_t1_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end[x][1] - data_table_t2_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end[x][1] - data_table_t3_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end[x][1] - data_table_t4_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")                    
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_start[x][1]) # Показания R+ начальные
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end[x][1]) # Показания R+ конечные
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end[x][1] - data_table_t0_rplus_delta_start[x][1]) # Показания R+ разница
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    
                    try:
                        data_table_temp.append(common_sql.get_k_t_t(obj_title)) # Коэффициент трансформации тока параметр 20
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(common_sql.get_k_t_n(obj_title)) # Коэффициент трансформации напряжения параметр 21
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[12]) # Энергия А+ параметр 22
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")

                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[19]) # Энергия R+ параметр 23
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                           
                    data_table.append(data_table_temp)
                    
                    
                    
                request.session["data_table_export"] = data_table
                
                               
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                              ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    # delta for groups abonents 'start date'
                    cursor_t0_aplus_delta_start_temp = connection.cursor()
                    cursor_t0_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t0_aplus_delta_start_temp = cursor_t0_aplus_delta_start_temp.fetchall()
                
                    cursor_t1_aplus_delta_start_temp = connection.cursor()
                    cursor_t1_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t1_aplus_delta_start_temp = cursor_t1_aplus_delta_start_temp.fetchall()
                
                    cursor_t2_aplus_delta_start_temp = connection.cursor()
                    cursor_t2_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t2_aplus_delta_start_temp = cursor_t2_aplus_delta_start_temp.fetchall()
                
                    cursor_t3_aplus_delta_start_temp = connection.cursor()
                    cursor_t3_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t3_aplus_delta_start_temp = cursor_t3_aplus_delta_start_temp.fetchall() 
                
                    cursor_t4_aplus_delta_start_temp = connection.cursor()
                    cursor_t4_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t4_aplus_delta_start_temp = cursor_t4_aplus_delta_start_temp.fetchall()
                    
                    # delta for groups abonents 'end date'
                    cursor_t0_aplus_delta_end_temp = connection.cursor()
                    cursor_t0_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_delta_end_temp = cursor_t0_aplus_delta_end_temp.fetchall()
                
                    cursor_t1_aplus_delta_end_temp = connection.cursor()
                    cursor_t1_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_delta_end_temp = cursor_t1_aplus_delta_end_temp.fetchall()
                
                    cursor_t2_aplus_delta_end_temp = connection.cursor()
                    cursor_t2_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_delta_end_temp = cursor_t2_aplus_delta_end_temp.fetchall()
                
                    cursor_t3_aplus_delta_end_temp = connection.cursor()
                    cursor_t3_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_delta_end_temp = cursor_t3_aplus_delta_end_temp.fetchall() 
                
                    cursor_t4_aplus_delta_end_temp = connection.cursor()
                    cursor_t4_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_delta_end_temp = cursor_t4_aplus_delta_end_temp.fetchall()
                    
                    cursor_t0_rplus_delta_start_temp = connection.cursor()
                    cursor_t0_rplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t0_rplus_delta_start_temp = cursor_t0_rplus_delta_start_temp.fetchall()
                    
                    cursor_t0_rplus_delta_end_temp = connection.cursor()
                    cursor_t0_rplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_rplus_delta_end_temp = cursor_t0_rplus_delta_end_temp.fetchall()
                    
                    data_table_temp = []
                    data_table_temp.append(abonents_list[x][0])
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1] - data_table_t0_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end_temp[0][1] - data_table_t1_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end_temp[0][1] - data_table_t2_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end_temp[0][1] - data_table_t3_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end_temp[0][1] - data_table_t4_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_start_temp[0][1]) # Показания R+ начальные
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1]) # Показания R+ конечные
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1] - data_table_t0_rplus_delta_start_temp[0][1]) # Показания R+ разница параметр 19
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(common_sql.get_k_t_t(abonents_list[x][0])) # Коэффициент трансформации тока параметр 20
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(common_sql.get_k_t_n(abonents_list[x][0])) # Коэффициент трансформации напряжения параметр 21
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[12]) # Энергия А+ параметр 22
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")

                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[19]) # Энергия R+ параметр 23
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                           
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table                 
#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render_to_response("data_table/1.html", args)
    
def data_table_period_3_tarifa(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    data_table = []
    data_table_graph_a_plus = []
    data_table_graph_r_plus = []

    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = u'1'
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period                       
            if (is_electric_period == "1") & (is_electric_daily =="1") & (bool(is_abonent_level.search(obj_key))): # pokazaniya za period
                 
                end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
                start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
                # dates формирует список дат от начальной до конечной даты                
                dates = [x for x in common_sql.daterange(start_date,
                              end_date,
                              step=datetime.timedelta(days=1),
                              inclusive=True)]
                # делаем выборку показаний по каждой дате в диапазоне указанных
                for x in range(len(dates)):
                    data_table_temp = get_data_table_by_date_daily(obj_title, obj_parent_title, datetime.datetime.strftime(dates[x], "%d.%m.%Y"))
                    if x >0:
                        try:
                            data_table_temp[0].append(data_table_temp[0][3] - data_table[x-1][3]) # Считаем разницу показаний по A+ за предыдущие сутки
                            data_table_temp[0].append(data_table_temp[0][8] - data_table[x-1][8]) # Считаем разницу показаний по R+ за предыдущие сутки                            
                        except:
                            next
                    # Блок проверки показаний за отдельную дату в диапазоне. Если показаний нет, то вставляем Н/Д в соответствующие поля        
                    if data_table_temp:
                        data_table.extend(data_table_temp)
                    else:
                        data_table.append([datetime.datetime.strftime(dates[x], "%d.%m.%Y"),obj_title,common_sql.get_serial_number_by_meter_name(obj_title), u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д'])
                #------------
                        
#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
               
    for x in range(len(data_table)):
        data_table_graph_a_plus_temp = []
        data_table_graph_r_plus_temp = []

        try:
            data_table_graph_a_plus_temp.append(data_table[x][0].strftime("%d.%m.%y"))
            data_table_graph_a_plus_temp.append(data_table[x][9])
            
            data_table_graph_r_plus_temp.append(data_table[x][0].strftime("%d.%m.%y"))
            data_table_graph_r_plus_temp.append(data_table[x][10])

            data_table_graph_a_plus.append(data_table_graph_a_plus_temp)
            data_table_graph_r_plus.append(data_table_graph_r_plus_temp)

        except:
            next
            
    # Сдвигаем дату на 1 число назад, потому что считаем энергию за прошедшие сутки            
    for x in range(len(data_table_graph_a_plus)):
        data_table_graph_a_plus[x][0] = (datetime.datetime.strptime(data_table_graph_a_plus[x][0],"%d.%m.%y")-datetime.timedelta(days=1)).strftime("%d.%m.%y")
        data_table_graph_a_plus[x][1] = data_table_graph_a_plus[x][1]*common_sql.get_k_t_n(obj_title)*common_sql.get_k_t_t(obj_title)
        
    # Сдвигаем дату на 1 число назад, потому что считаем энергию за прошедшие сутки            
    for x in range(len(data_table_graph_r_plus)):
        data_table_graph_r_plus[x][0] = (datetime.datetime.strptime(data_table_graph_r_plus[x][0],"%d.%m.%y")-datetime.timedelta(days=1)).strftime("%d.%m.%y")
        data_table_graph_r_plus[x][1] = data_table_graph_r_plus[x][1]*common_sql.get_k_t_n(obj_title)*common_sql.get_k_t_t(obj_title)
        

                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    args['data_table_graph_a_plus'] = data_table_graph_a_plus
    args['data_table_graph_r_plus'] = data_table_graph_r_plus
    

    

    return render_to_response("data_table/3.html", args)
    


    
    
    
    
def profil_30_aplus(request):
    args = {}
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = meters_name           = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']                     
            
            a_plus = connection.cursor()
            a_plus.execute("""SELECT 
                                  various_values.date, 
                                  various_values."time", 
                                  various_values.value, 
                                  meters.name, 
                                  meters.address, 
                                  names_params.name
                                FROM 
                                  public.various_values, 
                                  public.meters, 
                                  public.params, 
                                  public.taken_params, 
                                  public.names_params
                                WHERE 
                                  params.guid_names_params = names_params.guid AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  taken_params.id = various_values.id_taken_params AND
                                  various_values.date = %s AND 
                                  meters.name = %s AND 
                                  names_params.name = 'A+ Профиль';""",[electric_data_end, meters_name])
            a_plus = a_plus.fetchall()
            val_table_a_plus = []
           
            for x in range(len(a_plus)):
                my_val_table = [] 
                my_val_table.append(float(calendar.timegm(datetime.datetime.combine(a_plus[x][0], a_plus[x][1]).timetuple())*1000))
                my_val_table.append(a_plus[x][2]*2*common_sql.get_k_t_n(meters_name)*common_sql.get_k_t_t(meters_name))
                val_table_a_plus.append(my_val_table)
                
            r_plus = connection.cursor()
            r_plus.execute("""SELECT 
                                  various_values.date, 
                                  various_values."time", 
                                  various_values.value, 
                                  meters.name, 
                                  meters.address, 
                                  names_params.name
                                FROM 
                                  public.various_values, 
                                  public.meters, 
                                  public.params, 
                                  public.taken_params, 
                                  public.names_params
                                WHERE 
                                  params.guid_names_params = names_params.guid AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  taken_params.id = various_values.id_taken_params AND
                                  various_values.date = %s AND 
                                  meters.name = %s AND 
                                  names_params.name = 'R+ Профиль';""",[electric_data_end, meters_name])
            r_plus = r_plus.fetchall()
            val_table_r_plus = []
           
            for x in range(len(r_plus)):
                my_val_table = [] 
                my_val_table.append(float(calendar.timegm(datetime.datetime.combine(r_plus[x][0], r_plus[x][1]).timetuple())*1000))
                my_val_table.append(r_plus[x][2]*2*common_sql.get_k_t_n(meters_name)*common_sql.get_k_t_t(meters_name))
                val_table_r_plus.append(my_val_table)
                
            data_table = []
            for x in range(len(a_plus)):
                data_table_temp = []
                try:
                    data_table_temp.append(a_plus[x][0])
                except IndexError:
                    data_table_temp.append(u"Н/Д")
                except TypeError:
                    data_table_temp.append(u"Н/Д")
                try:
                    data_table_temp.append(a_plus[x][1])
                except IndexError:
                    data_table_temp.append(u"Н/Д")
                except TypeError:
                    data_table_temp.append(u"Н/Д")
                try:
                    data_table_temp.append(a_plus[x][3])
                except IndexError:
                    data_table_temp.append(u"Н/Д")
                except TypeError:
                    data_table_temp.append(u"Н/Д")
                try:
                    data_table_temp.append(a_plus[x][2]*2*common_sql.get_k_t_n(meters_name)*common_sql.get_k_t_t(meters_name))
                except IndexError:
                    data_table_temp.append(u"Н/Д")
                except TypeError:
                    data_table_temp.append(u"Н/Д")
                try:
                    data_table_temp.append(r_plus[x][2]*2*common_sql.get_k_t_n(meters_name)*common_sql.get_k_t_t(meters_name))
                except IndexError:
                    data_table_temp.append(u"Н/Д")
                except TypeError:
                    data_table_temp.append(u"Н/Д")
                data_table.append(data_table_temp)

        
     
            args['min30_a_plus'] = val_table_a_plus
            args['min30_r_plus'] = val_table_r_plus
                
            args['data_table'] = data_table
            args['k_t_n'] = common_sql.get_k_t_n(meters_name)
            args['k_t_t'] = common_sql.get_k_t_t(meters_name)
            args['meters_name'] = meters_name
            args['electric_data_end'] = electric_data_end
    
    

    
        
    return render_to_response("data_table/4.html", args)
    



def hour_increment(request):
    args = {}
#    meters_name= u'Не выбран'
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']  
    time_list = ['00:00', '00:30','01:00', '01:30', '02:00', '02:30', '03:00', '03:30', '04:00', '04:30', '05:00', '05:30', '06:00', '06:30', '07:00', '07:30', '08:00', '08:30', '09:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00', '12:30', '13:00', '13:30', '14:00', '14:30', '15:00', '15:30', '16:00', '16:30', '17:00', '17:30', '18:00', '18:30', '19:00', '19:30', '20:00', '20:30', '21:00', '21:30', '22:00', '22:30', '23:00', '23:30']
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = meters_name           = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
    
    serial_number = common_sql.get_serial_number_by_meter_name(meters_name)
        
    data_table = []
    if meters_name != u'Не выбран':
        # Добавляем первую строку в таблицу данных. Делаем запрос показаний на начало суток.
        data_table.append([electric_data_end,u'00:00', meters_name, serial_number, common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ),common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 R+' ), u'---', u'---'])
        
        if common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ) != u'Нет данных':  # Если есть показания на начало суток выполняем почасовое приращение  
            for x in range(24):
                data_table_temp = []
                data_table_temp.append(electric_data_end)
                data_table_temp.append(time_list[(2*x)])
                data_table_temp.append(meters_name)
                data_table_temp.append(serial_number)
                data_table_temp.append(data_table[len(data_table)-1][4] + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))
                data_table_temp.append(common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 R+' ) + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))
                data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))
                data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))    
                if x == 0: # Убираем первую строку. Так как показания на 00:00 берем отдельным запросом
                    next
                else:
                    data_table.append(data_table_temp)    
    
    args['data_table'] = data_table
    args['meters_name'] = meters_name           
    args['electric_data_end'] = electric_data_end
    return render_to_response("data_table/6.html", args)
    


    

def economic_electric(request):
    args = {}
    data_table = []
    
    electric_data_end   = request.session["electric_data_end"]
    electric_data_start   = request.session["electric_data_start"]



    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']    

    end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
    start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
    dates = [x for x in common_sql.daterange(start_date,
                  end_date,
                  step=datetime.timedelta(days=1),
                  inclusive=True)]

                  
    for x in range(len(dates)):
        try:
            data_table_temp = []
            delta_a_plus = 1
            delta_r_plus = 1

            try:
                delta_a_plus = common_sql.delta_sum_a_plus(dates[x+1])-common_sql.delta_sum_a_plus(dates[x])
                if delta_a_plus > 0:
				    delta_a_plus = delta_a_plus
                else:
				    delta_a_plus = u'Н/Д'
                delta_r_plus = common_sql.delta_sum_r_plus(dates[x+1])-common_sql.delta_sum_r_plus(dates[x])
                if delta_r_plus > 0:
				    delta_r_plus = delta_r_plus
                else:
                    delta_r_plus = u'Н/Д'

            except:
                delta_a_plus = u'Н/Д'
                delta_r_plus = u'Н/Д'

            data_table_temp.append(dates[x])
            data_table_temp.append(common_sql.product_sum(dates[x]))
            data_table_temp.append(delta_a_plus)
            data_table_temp.append(delta_a_plus/(common_sql.product_sum(dates[x])))
            data_table_temp.append(delta_r_plus)
            data_table_temp.append(delta_r_plus/(common_sql.product_sum(dates[x])))
        except:
            next
        data_table.append(data_table_temp)
    data_graph = []    
    for x in range(len(data_table)):
        data_graph_temp = []
        #data_graph_temp.append(x)
        try:
            data_graph_temp.append(data_table[x][0].strftime("%d.%m.%y"))
            data_graph_temp.append(data_table[x][3])
            data_graph.append(data_graph_temp)
        except:
            next
        
    
    args['economic_graph_data'] = json.dumps(data_graph)#[[u'Jan', 13], [datetime.datetime.now().strftime("%Y-%m-%d"), 17], [50, 9]])
    #args['economic_graph_data'] = data_graph
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    return render_to_response("data_table/7.html", args)
    
def rejim_day(request):
    args = {}
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = meters_name           = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
    return render_to_response("data_table/8.html", args)    

def load_balance_groups(request):
    #сделать потом воду, пока тут балансные группы!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    parent_name         = request.GET['obj_parent_title']
    ab_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
                # Добавляем привязку к балансной группе 
    cfg_excel_name = 'C:/work/mitino/prizmer/static/cfg/kB_balance_for_load.xlsx'
    cfg_sheet_name = u'ВРУ-1'
    is_electic_cfg = True
    is_water_cfg = False
    is_heat_cfg = False
    from django.db import connection
    from openpyxl import load_workbook
    wb = load_workbook(filename = cfg_excel_name)
    sheet_ranges = wb[cfg_sheet_name]
    row = 2
    dt=[]

    while (bool(sheet_ranges[u'A%s'%(row)].value)):
        guid_balance_groups_from_excel = connection.cursor()
        balance_group_name=[unicode(sheet_ranges[u'A%s'%(row)].value)]
        guid_balance_groups_from_excel.execute("""SELECT balance_groups.guid FROM public.balance_groups WHERE balance_groups.name = %s;""",balance_group_name)
        guid_balance_groups_from_excel = guid_balance_groups_from_excel.fetchall()
        if len(guid_balance_groups_from_excel)>0:
            guid_balance_groups = BalanceGroups.objects.get(guid=guid_balance_groups_from_excel[0][0])
        else: 
            #print u'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!'
            #print u'Надо создать балансную группу(вручную), прежде чем добавлять в неё что-то.'
            break
        guid_meters_from_excel = connection.cursor()
        meters_name=[unicode(sheet_ranges[u'E%s'%(row)].value)]
        znak=[bool(sheet_ranges[u'D%s'%(row)].value)]
        guid_meters_from_excel.execute("""SELECT meters.guid FROM public.meters WHERE meters.factory_number_manual = %s;""",meters_name)
        guid_meters_from_excel = guid_meters_from_excel.fetchall()
        if len(guid_meters_from_excel)>0:
            guid_meters = Meters.objects.get(guid=guid_meters_from_excel[0][0])
        else:
            #print u'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!'
            #print u'Такого счётчика не существует в БД, он не может быть добавлен в балансную группу.'
            continue
        add_link_meter_balance_group = LinkBalanceGroupsMeters(guid_balance_groups = guid_balance_groups, guid_meters = guid_meters, type=znak[0])
        add_link_meter_balance_group.save()
        
        dt.append([balance_group_name,meters_name,znak])
        #print unicode(sheet_ranges[u'A%s'%(row)].value), unicode(row), unicode([balance_group_name,meters_name,znak])
        row = row + 1
    
    args = {}

    args['data_table'] = dt
    args['electric_data_end'] = electric_data_end
    return render_to_response("data_table/water/24.html", args)

def pokazaniya_water(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_channel(meters_name, electric_data_end)
    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        #print list_of_abonents_2
        #print common_sql.return_parent_guid_by_abonent_name(parent_name)
        #print meters_name
        #print common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []        
        for x in range(len(list_of_abonents_2)):
            data_table_temp = common_sql.get_daily_water_channel(list_of_abonents_2[x], electric_data_end)
            data_table.extend(data_table_temp)
            
    elif(bool(is_object_level_1.search(obj_key))):
        list_of_objects_2 = common_sql.list_of_objects(common_sql.return_parent_guid_by_abonent_name(meters_name)) #Список квартир для объекта с пульсарами
        data_table = []
        for x in range(len(list_of_objects_2)):
            data_table_temp = [(list_of_objects_2[x][0],)]
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(meters_name), list_of_objects_2[x][0])
            for y in range(len(list_of_abonents_2)):
                data_table_temp2 = common_sql.get_daily_water_channel(list_of_abonents_2[y], electric_data_end)

                data_table_temp.extend(data_table_temp2)                                
                      
            data_table.extend(data_table_temp)
              
    else:
        data_table = []
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end

    return render_to_response("data_table/water/10.html", args)
    
def pokazaniya_water_identificators(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_channel(meters_name, electric_data_end)
    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []        
        for x in range(len(list_of_abonents_2)):
            data_table_temp = common_sql.get_daily_water_channel(list_of_abonents_2[x], electric_data_end)
            data_table.extend(data_table_temp)
    elif(bool(is_object_level_1.search(obj_key))):
        
        list_of_objects_2 = common_sql.list_of_objects(common_sql.return_parent_guid_by_abonent_name(meters_name)) #Список квартир для объекта с пульсарами
        data_table = []
        for x in range(len(list_of_objects_2)):
            data_table_temp = [(list_of_objects_2[x][0],)]
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(meters_name), list_of_objects_2[x][0])
            for y in range(len(list_of_abonents_2)):
                data_table_temp2 = common_sql.get_daily_water_channel(list_of_abonents_2[y], electric_data_end)

                data_table_temp.extend(data_table_temp2)                                
                      
            data_table.extend(data_table_temp)
              
    else:
        data_table = []
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name

    return render_to_response("data_table/water/12.html", args)     

def pokazaniya_water_gvs_hvs_daily(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.GET['obj_parent_title']
    obj_title         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end, 'daily', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table=common_sql.get_daily_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end, 'daily', False)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = obj_title

    return render_to_response("data_table/water/28.html", args)     
def pokazaniya_water_gvs_hvs_current(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.GET['obj_parent_title']
    obj_title         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_current_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end,  True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table_temp=common_sql.get_current_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end,  False)
        for row in data_table_temp:
            if row[4]==u'Н/Д' and row[5]==u'Н/Д':
                row2=common_sql.get_current_water_gvs_hvs(unicode(row[2]), unicode(row[6]) , electric_data_end, True)
                #print row2
                #print unicode(row[2]), unicode(row[6]), electric_data_end, True
                if len(row2)==0:
                    r=[unicode(electric_data_end), u'Н/Д', unicode(row[2]),unicode(row[3]), u'Н/Д', u'Н/Д']
                    data_table.append(r)
                else:
                    data_table.append(row2[0])
            else:
                data_table.append(row)

    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = obj_title

    return render_to_response("data_table/water/26.html", args)     

def water_elf_hvs_by_date(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.GET['obj_parent_title']
    obj_title         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '1','attr1', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '1', 'attr1',False)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['res'] = u'ХВС'
    args['obj_title'] = obj_title
    return render_to_response("data_table/water/52.html", args)     

def water_elf_gvs_by_date(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.GET['obj_parent_title']
    obj_title         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '2','attr2', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '2', 'attr2',False)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['res'] = u'ГВС'
    args['obj_title'] = obj_title

    return render_to_response("data_table/water/52.html", args)

def water_elf_hvs_potreblenie(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.GET['obj_parent_title']
    obj_title         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    electric_data_start = request.GET['electric_data_start']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end, electric_data_start,'1','attr1', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end, electric_data_start,'1','attr1', False)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['res'] = u'ХВС'
    args['obj_title'] = obj_title

    return render_to_response("data_table/water/53.html", args)
    
def water_elf_gvs_potreblenie(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    electric_data_start = request.GET['electric_data_start']
    obj_parent_title         = request.GET['obj_parent_title']
    obj_title         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end,electric_data_start, '2','attr2', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end,electric_data_start, '2','attr2', False)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['res'] = u'ГВС'
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = obj_title

    return render_to_response("data_table/water/53.html", args)

def potreblenie_water(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']                        
    obj_key             = request.GET['obj_key']

    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']                        
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table_start = common_sql.get_daily_water_channel(meters_name, electric_data_start) # Таблица с начальными значениями
        data_table_end = common_sql.get_daily_water_channel(meters_name, electric_data_end)     # Таблица с конечными значениями
        
        data_table = [[data_table_start[0][0],data_table_start[0][1],data_table_start[0][2],data_table_start[0][3],data_table_end[0][2],data_table_end[0][2]-data_table_start[0][2]]]

        
    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []        
        for x in range(len(list_of_abonents_2)):
            data_table_temp_start = common_sql.get_daily_water_channel(list_of_abonents_2[x], electric_data_start)
            data_table_temp_end = common_sql.get_daily_water_channel(list_of_abonents_2[x], electric_data_end)
            data_table_temp = [[data_table_temp_start[0][0],data_table_temp_start[0][1],data_table_temp_start[0][2],data_table_temp_start[0][3],data_table_temp_end[0][2],data_table_temp_end[0][2]-data_table_temp_start[0][2]]]

            data_table.extend(data_table_temp)

    elif(bool(is_object_level_1.search(obj_key))):
        
        list_of_objects_2 = common_sql.list_of_objects(common_sql.return_parent_guid_by_abonent_name(meters_name)) #Список квартир для объекта с пульсарами
        data_table = []
        for x in range(len(list_of_objects_2)):
            data_table_temp = [(list_of_objects_2[x][0],)]
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(meters_name), list_of_objects_2[x][0])
            for y in range(len(list_of_abonents_2)):
                data_table_temp2_end = common_sql.get_daily_water_channel(list_of_abonents_2[y], electric_data_end)
                data_table_temp2_start = common_sql.get_daily_water_channel(list_of_abonents_2[y], electric_data_start)
                #print data_table_temp2_end
                if bool(data_table_temp2_end) and bool(data_table_temp2_start):
                
                    data_table_temp2 = [[data_table_temp2_start[0][0],data_table_temp2_start[0][1],data_table_temp2_start[0][2],data_table_temp2_start[0][3],data_table_temp2_end[0][2],data_table_temp2_end[0][2]-data_table_temp2_start[0][2]]]
                else:
                    data_table_temp2 = [[list_of_abonents_2[y][0], u'Н/Д', u'-', u'-', u'-']]                

                data_table_temp.extend(data_table_temp2)
            data_table.extend(data_table_temp)

    else:
        data_table = []
                                                     
    
    args['data_table'] = data_table
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name


    return render_to_response("data_table/water/11.html", args)

def num_from_name(name):
    start = name.find(u'№')
    num = name[start+1:]
    return num
    
    
def add_numbers(request):
    g =  Abonents.objects.values_list("guid")
    for x in range(len(g)):
        t = Abonents.objects.get(guid = g[x][0])
        if num_from_name(t.name):
            t.account_2 = num_from_name(t.name)  # change field
            t.save() # this will update only
        else:
            next
    html = u'Готово'
    return HttpResponse(html)

def electric_simple_2_zones_old(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = get_data_table_by_date_monthly(obj_title, obj_parent_title, electric_data_end)

                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = get_data_table_by_date_daily(obj_title, obj_parent_title, electric_data_end)


            elif (is_electric_current == "1") & (bool(is_abonent_level.search(obj_key))):
                pass
                            
            elif (is_electric_period == "1") & (is_electric_daily =="1") & (bool(is_abonent_level.search(obj_key))): # pokazaniya za period

                end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
                start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
                dates = [x for x in common_sql.daterange(start_date,
                              end_date,
                              step=datetime.timedelta(days=1),
                              inclusive=True)]
                '''for x in range(len(dates)):
                    data_table_temp = [dates[x], dates[x], datetime.datetime.strftime(dates[x], "%d.%m.%Y")]
                    data_table.append(data_table_temp)'''

                for x in range(len(dates)):
                    data_table_temp = get_data_table_by_date_daily(obj_title, obj_parent_title, datetime.datetime.strftime(dates[x], "%d.%m.%Y"))
                    data_table.extend(data_table_temp)
                #------------

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # monthly for abonents group
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    # query data for each abonent
                    cursor_t0_aplus_monthly_temp = connection.cursor()
                    cursor_t0_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T0 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_monthly_temp = cursor_t0_aplus_monthly_temp.fetchall()
                    
                    cursor_t1_aplus_monthly_temp = connection.cursor()
                    cursor_t1_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T1 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_monthly_temp = cursor_t1_aplus_monthly_temp.fetchall()
                    
                    cursor_t2_aplus_monthly_temp = connection.cursor()
                    cursor_t2_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date,
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T2 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY 
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_monthly_temp = cursor_t2_aplus_monthly_temp.fetchall()
                    
                    cursor_t3_aplus_monthly_temp = connection.cursor()
                    cursor_t3_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T3 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_monthly_temp = cursor_t3_aplus_monthly_temp.fetchall()
                
                    cursor_t4_aplus_monthly_temp = connection.cursor()
                    cursor_t4_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T4 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_monthly_temp = cursor_t4_aplus_monthly_temp.fetchall()
                    
                    cursor_t0_rplus_monthly_temp = connection.cursor()
                    cursor_t0_rplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T0 R+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_rplus_monthly_temp = cursor_t0_rplus_monthly_temp.fetchall()
                
                    data_table_temp = []
                    try:
                        data_table_temp.append(electric_data_end)
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    data_table_temp.append(abonents_list[x][0])
                    
                    try:
                        data_table_temp.append(data_table_t0_aplus_monthly_temp[0][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t0_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:    
                        data_table_temp.append(data_table_t1_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t2_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_rplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1'):# & (bool(is_object_level.search(obj_key))): # daily for abonents group
               
                    
                if (bool(is_object_level.search(obj_key))):
                    cursor_abonents_list = connection.cursor()
                    cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title])
                    abonents_list = cursor_abonents_list.fetchall()

                    
                elif (bool(is_group_level.search(obj_key))):
                    cursor_abonents_list = connection.cursor()
                    cursor_abonents_list.execute("""
                                                SELECT 
                                                  meters.name,
                                                  link_balance_groups_meters.type
                                                FROM 
                                                  public.meters, 
                                                  public.link_balance_groups_meters, 
                                                  public.balance_groups
                                                WHERE 
                                                  link_balance_groups_meters.guid_balance_groups = balance_groups.guid AND
                                                  link_balance_groups_meters.guid_meters = meters.guid AND
                                                  balance_groups.name = %s
                                                ORDER BY
                                                  meters.name ASC;""",[obj_title])
                    abonents_list = cursor_abonents_list.fetchall()
                    obj_title=u'Завод'
                else:
                    abonents_list = [12345678]
                              

                for x in range(len(abonents_list)):
                    cursor_t0_aplus_daily_temp = connection.cursor()
                    cursor_t0_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_daily_temp = cursor_t0_aplus_daily_temp.fetchall()
                
                    cursor_t1_aplus_daily_temp = connection.cursor()
                    cursor_t1_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_daily_temp = cursor_t1_aplus_daily_temp.fetchall()
                
                    cursor_t2_aplus_daily_temp = connection.cursor()
                    cursor_t2_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_daily_temp = cursor_t2_aplus_daily_temp.fetchall()
                
                    cursor_t3_aplus_daily_temp = connection.cursor()
                    cursor_t3_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_daily_temp = cursor_t3_aplus_daily_temp.fetchall() 
                
                    cursor_t4_aplus_daily_temp = connection.cursor()
                    cursor_t4_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_daily_temp = cursor_t4_aplus_daily_temp.fetchall()
                    
                    cursor_t0_rplus_daily_temp = connection.cursor()
                    cursor_t0_rplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_rplus_daily_temp = cursor_t0_rplus_daily_temp.fetchall()
                    
                    data_table_temp = []
                    try:
                        data_table_temp.append(electric_data_end)
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    data_table_temp.append(abonents_list[x][0])
                    try:
                        data_table_temp.append(data_table_t0_aplus_daily_temp[0][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        if (bool(is_group_level.search(obj_key))):                           
                            if abonents_list[x][1]: # Если абонент входит в группу со знаком плюс, то показания как есть
                                data_table_temp.append(data_table_t0_aplus_daily_temp[0][1])
                            else:                   # Если абонент входит в группу со знаком минус, то показазния инвертируются
                                data_table_temp.append(-data_table_t0_aplus_daily_temp[0][1])
                        else:
                           data_table_temp.append(data_table_t0_aplus_daily_temp[0][1]) 
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        if (bool(is_group_level.search(obj_key))):                                           
                            if abonents_list[x][1]: # Если абонент входит в группу со знаком плюс, то показания как есть
                                data_table_temp.append(data_table_t0_rplus_daily_temp[0][1])
                            else:
                                data_table_temp.append(-data_table_t0_rplus_daily_temp[0][1])
                        else:
                            data_table_temp.append(data_table_t0_rplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    data_table.append(data_table_temp)
                if (bool(is_group_level.search(obj_key))):  # Если это группа добавляем еще одну строку с суммой показаний
                    sum_a_plus = 0
                    sum_r_plus = 0
                    for x in range(len(data_table)):
                        try:
                            sum_a_plus = sum_a_plus + data_table[x][3]
                            sum_r_plus = sum_r_plus + data_table[x][8]
                        except:
                            next
                    data_table.append([])
                    data_table.append([u' ',u' ',u'<strong>Сумма</strong>',sum_a_plus,u'-',u'-',u'-',u'-',sum_r_plus])

                request.session["data_table_export"] = data_table
            elif (is_electric_current == '1') & (bool(is_object_level.search(obj_key))): # текущие для объекта учёта
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                              ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    data_table_temp = []
                    data_table_temp.append(u'Дата')
                    data_table_temp.append(abonents_list[x][0])
                    data_table_temp.append(u'Какой-то заводской номер')
                    data_table_temp.append(0)
                    data_table_temp.append(100)
                    data_table_temp.append(200)
                    data_table_temp.append(300)
                    data_table_temp.append(400)
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render_to_response("data_table/electric/14.html", args)

def electric_simple_3_zones_old(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = get_data_table_by_date_monthly(obj_title, obj_parent_title, electric_data_end)

                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = get_data_table_by_date_daily(obj_title, obj_parent_title, electric_data_end)


            elif (is_electric_current == "1") & (bool(is_abonent_level.search(obj_key))):
                pass
                            
            elif (is_electric_period == "1") & (is_electric_daily =="1") & (bool(is_abonent_level.search(obj_key))): # pokazaniya za period

                end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
                start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
                dates = [x for x in common_sql.daterange(start_date,
                              end_date,
                              step=datetime.timedelta(days=1),
                              inclusive=True)]
                '''for x in range(len(dates)):
                    data_table_temp = [dates[x], dates[x], datetime.datetime.strftime(dates[x], "%d.%m.%Y")]
                    data_table.append(data_table_temp)'''

                for x in range(len(dates)):
                    data_table_temp = get_data_table_by_date_daily(obj_title, obj_parent_title, datetime.datetime.strftime(dates[x], "%d.%m.%Y"))
                    data_table.extend(data_table_temp)
                #------------

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # monthly for abonents group
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    # query data for each abonent
                    cursor_t0_aplus_monthly_temp = connection.cursor()
                    cursor_t0_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T0 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_monthly_temp = cursor_t0_aplus_monthly_temp.fetchall()
                    
                    cursor_t1_aplus_monthly_temp = connection.cursor()
                    cursor_t1_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T1 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_monthly_temp = cursor_t1_aplus_monthly_temp.fetchall()
                    
                    cursor_t2_aplus_monthly_temp = connection.cursor()
                    cursor_t2_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date,
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T2 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY 
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_monthly_temp = cursor_t2_aplus_monthly_temp.fetchall()
                    
                    cursor_t3_aplus_monthly_temp = connection.cursor()
                    cursor_t3_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T3 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_monthly_temp = cursor_t3_aplus_monthly_temp.fetchall()
                
                    cursor_t4_aplus_monthly_temp = connection.cursor()
                    cursor_t4_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T4 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_monthly_temp = cursor_t4_aplus_monthly_temp.fetchall()
                    
                    cursor_t0_rplus_monthly_temp = connection.cursor()
                    cursor_t0_rplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T0 R+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_rplus_monthly_temp = cursor_t0_rplus_monthly_temp.fetchall()
                
                    data_table_temp = []
                    try:
                        data_table_temp.append(electric_data_end)
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    data_table_temp.append(abonents_list[x][0])
                    
                    try:
                        data_table_temp.append(data_table_t0_aplus_monthly_temp[0][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t0_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:    
                        data_table_temp.append(data_table_t1_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t2_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_rplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1'):# & (bool(is_object_level.search(obj_key))): # daily for abonents group
               
                    
                if (bool(is_object_level.search(obj_key))):
                    cursor_abonents_list = connection.cursor()
                    cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title])
                    abonents_list = cursor_abonents_list.fetchall()

                    
                elif (bool(is_group_level.search(obj_key))):
                    cursor_abonents_list = connection.cursor()
                    cursor_abonents_list.execute("""
                                                SELECT 
                                                  meters.name,
                                                  link_balance_groups_meters.type
                                                FROM 
                                                  public.meters, 
                                                  public.link_balance_groups_meters, 
                                                  public.balance_groups
                                                WHERE 
                                                  link_balance_groups_meters.guid_balance_groups = balance_groups.guid AND
                                                  link_balance_groups_meters.guid_meters = meters.guid AND
                                                  balance_groups.name = %s
                                                ORDER BY
                                                  meters.name ASC;""",[obj_title])
                    abonents_list = cursor_abonents_list.fetchall()
                    obj_title=u'Завод'
                else:
                    abonents_list = [12345678]
                              

                for x in range(len(abonents_list)):
                    cursor_t0_aplus_daily_temp = connection.cursor()
                    cursor_t0_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_daily_temp = cursor_t0_aplus_daily_temp.fetchall()
                
                    cursor_t1_aplus_daily_temp = connection.cursor()
                    cursor_t1_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_daily_temp = cursor_t1_aplus_daily_temp.fetchall()
                
                    cursor_t2_aplus_daily_temp = connection.cursor()
                    cursor_t2_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_daily_temp = cursor_t2_aplus_daily_temp.fetchall()
                
                    cursor_t3_aplus_daily_temp = connection.cursor()
                    cursor_t3_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_daily_temp = cursor_t3_aplus_daily_temp.fetchall() 
                
                    cursor_t4_aplus_daily_temp = connection.cursor()
                    cursor_t4_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_daily_temp = cursor_t4_aplus_daily_temp.fetchall()
                    
                    cursor_t0_rplus_daily_temp = connection.cursor()
                    cursor_t0_rplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_rplus_daily_temp = cursor_t0_rplus_daily_temp.fetchall()
                    
                    data_table_temp = []
                    try:
                        data_table_temp.append(electric_data_end)
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    data_table_temp.append(abonents_list[x][0])
                    try:
                        data_table_temp.append(data_table_t0_aplus_daily_temp[0][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        if (bool(is_group_level.search(obj_key))):                           
                            if abonents_list[x][1]: # Если абонент входит в группу со знаком плюс, то показания как есть
                                data_table_temp.append(data_table_t0_aplus_daily_temp[0][1])
                            else:                   # Если абонент входит в группу со знаком минус, то показазния инвертируются
                                data_table_temp.append(-data_table_t0_aplus_daily_temp[0][1])
                        else:
                           data_table_temp.append(data_table_t0_aplus_daily_temp[0][1]) 
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        if (bool(is_group_level.search(obj_key))):                                           
                            if abonents_list[x][1]: # Если абонент входит в группу со знаком плюс, то показания как есть
                                data_table_temp.append(data_table_t0_rplus_daily_temp[0][1])
                            else:
                                data_table_temp.append(-data_table_t0_rplus_daily_temp[0][1])
                        else:
                            data_table_temp.append(data_table_t0_rplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    data_table.append(data_table_temp)
                if (bool(is_group_level.search(obj_key))):  # Если это группа добавляем еще одну строку с суммой показаний
                    sum_a_plus = 0
                    sum_r_plus = 0
                    for x in range(len(data_table)):
                        try:
                            sum_a_plus = sum_a_plus + data_table[x][3]
                            sum_r_plus = sum_r_plus + data_table[x][8]
                        except:
                            next
                    data_table.append([])
                    data_table.append([u' ',u' ',u'<strong>Сумма</strong>',sum_a_plus,u'-',u'-',u'-',u'-',sum_r_plus])

                request.session["data_table_export"] = data_table
            elif (is_electric_current == '1') & (bool(is_object_level.search(obj_key))): # текущие для объекта учёта
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                              ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    data_table_temp = []
                    data_table_temp.append(u'Дата')
                    data_table_temp.append(abonents_list[x][0])
                    data_table_temp.append(u'Какой-то заводской номер')
                    data_table_temp.append(0)
                    data_table_temp.append(100)
                    data_table_temp.append(200)
                    data_table_temp.append(300)
                    data_table_temp.append(400)
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    
    return render_to_response("data_table/electric/16.html", args)
    
def electric_potreblenie_2_zones(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u'1'
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                cursor_t0_aplus_delta_start = connection.cursor()
                cursor_t0_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t0_aplus_delta_start = cursor_t0_aplus_delta_start.fetchall()
                
                cursor_t1_aplus_delta_start = connection.cursor()
                cursor_t1_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t1_aplus_delta_start = cursor_t1_aplus_delta_start.fetchall()
                
                cursor_t2_aplus_delta_start = connection.cursor()
                cursor_t2_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t2_aplus_delta_start = cursor_t2_aplus_delta_start.fetchall()
                
                cursor_t3_aplus_delta_start = connection.cursor()
                cursor_t3_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t3_aplus_delta_start = cursor_t3_aplus_delta_start.fetchall() 
                
                cursor_t4_aplus_delta_start = connection.cursor()
                cursor_t4_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t4_aplus_delta_start = cursor_t4_aplus_delta_start.fetchall()

                cursor_t0_aplus_delta_end = connection.cursor()
                cursor_t0_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t0_aplus_delta_end = cursor_t0_aplus_delta_end.fetchall()
                
                cursor_t1_aplus_delta_end = connection.cursor()
                cursor_t1_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t1_aplus_delta_end = cursor_t1_aplus_delta_end.fetchall()
                
                cursor_t2_aplus_delta_end = connection.cursor()
                cursor_t2_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t2_aplus_delta_end = cursor_t2_aplus_delta_end.fetchall()
                
                cursor_t3_aplus_delta_end = connection.cursor()
                cursor_t3_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t3_aplus_delta_end = cursor_t3_aplus_delta_end.fetchall() 
                
                cursor_t4_aplus_delta_end = connection.cursor()
                cursor_t4_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t4_aplus_delta_end = cursor_t4_aplus_delta_end.fetchall()
                
                cursor_t0_rplus_delta_start = connection.cursor()
                cursor_t0_rplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t0_rplus_delta_start = cursor_t0_rplus_delta_start.fetchall()
                
                cursor_t0_rplus_delta_end = connection.cursor()
                cursor_t0_rplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t0_rplus_delta_end = cursor_t0_rplus_delta_end.fetchall()
                
                
#                data_table = []
                for x in range(len(data_table_t0_aplus_delta_end)):
                    data_table_temp = []

                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][2])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][1] - data_table_t0_aplus_delta_start[x][1] )
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end[x][1] - data_table_t1_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end[x][1] - data_table_t2_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end[x][1] - data_table_t3_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end[x][1] - data_table_t4_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")                    
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_start[x][1]) # Показания R+ начальные
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end[x][1]) # Показания R+ конечные
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end[x][1] - data_table_t0_rplus_delta_start[x][1]) # Показания R+ разница
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    
                    try:
                        data_table_temp.append(common_sql.get_k_t_t(obj_title)) # Коэффициент трансформации тока параметр 20
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(common_sql.get_k_t_n(obj_title)) # Коэффициент трансформации напряжения параметр 21
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[12]) # Энергия А+ параметр 22
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")

                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[19]) # Энергия R+ параметр 23
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                           
                    data_table.append(data_table_temp)
                    
                    
                    
                request.session["data_table_export"] = data_table
                
                               
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                              ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    # delta for groups abonents 'start date'
                    cursor_t0_aplus_delta_start_temp = connection.cursor()
                    cursor_t0_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t0_aplus_delta_start_temp = cursor_t0_aplus_delta_start_temp.fetchall()
                
                    cursor_t1_aplus_delta_start_temp = connection.cursor()
                    cursor_t1_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t1_aplus_delta_start_temp = cursor_t1_aplus_delta_start_temp.fetchall()
                
                    cursor_t2_aplus_delta_start_temp = connection.cursor()
                    cursor_t2_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t2_aplus_delta_start_temp = cursor_t2_aplus_delta_start_temp.fetchall()
                
                    cursor_t3_aplus_delta_start_temp = connection.cursor()
                    cursor_t3_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t3_aplus_delta_start_temp = cursor_t3_aplus_delta_start_temp.fetchall() 
                
                    cursor_t4_aplus_delta_start_temp = connection.cursor()
                    cursor_t4_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t4_aplus_delta_start_temp = cursor_t4_aplus_delta_start_temp.fetchall()
                    
                    # delta for groups abonents 'end date'
                    cursor_t0_aplus_delta_end_temp = connection.cursor()
                    cursor_t0_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_delta_end_temp = cursor_t0_aplus_delta_end_temp.fetchall()
                
                    cursor_t1_aplus_delta_end_temp = connection.cursor()
                    cursor_t1_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_delta_end_temp = cursor_t1_aplus_delta_end_temp.fetchall()
                
                    cursor_t2_aplus_delta_end_temp = connection.cursor()
                    cursor_t2_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_delta_end_temp = cursor_t2_aplus_delta_end_temp.fetchall()
                
                    cursor_t3_aplus_delta_end_temp = connection.cursor()
                    cursor_t3_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_delta_end_temp = cursor_t3_aplus_delta_end_temp.fetchall() 
                
                    cursor_t4_aplus_delta_end_temp = connection.cursor()
                    cursor_t4_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_delta_end_temp = cursor_t4_aplus_delta_end_temp.fetchall()
                    
                    cursor_t0_rplus_delta_start_temp = connection.cursor()
                    cursor_t0_rplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t0_rplus_delta_start_temp = cursor_t0_rplus_delta_start_temp.fetchall()
                    
                    cursor_t0_rplus_delta_end_temp = connection.cursor()
                    cursor_t0_rplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_rplus_delta_end_temp = cursor_t0_rplus_delta_end_temp.fetchall()
                    
                    data_table_temp = []
                    data_table_temp.append(abonents_list[x][0])
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1] - data_table_t0_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end_temp[0][1] - data_table_t1_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end_temp[0][1] - data_table_t2_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end_temp[0][1] - data_table_t3_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end_temp[0][1] - data_table_t4_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_start_temp[0][1]) # Показания R+ начальные
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1]) # Показания R+ конечные
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1] - data_table_t0_rplus_delta_start_temp[0][1]) # Показания R+ разница параметр 19
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(common_sql.get_k_t_t(abonents_list[x][0])) # Коэффициент трансформации тока параметр 20
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(common_sql.get_k_t_n(abonents_list[x][0])) # Коэффициент трансформации напряжения параметр 21
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[12]) # Энергия А+ параметр 22
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")

                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[19]) # Энергия R+ параметр 23
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                           
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table                 
#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    return render_to_response("data_table/electric/15.html", args)
    
def electric_potreblenie_3_zones(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u'1'
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                cursor_t0_aplus_delta_start = connection.cursor()
                cursor_t0_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t0_aplus_delta_start = cursor_t0_aplus_delta_start.fetchall()
                
                cursor_t1_aplus_delta_start = connection.cursor()
                cursor_t1_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t1_aplus_delta_start = cursor_t1_aplus_delta_start.fetchall()
                
                cursor_t2_aplus_delta_start = connection.cursor()
                cursor_t2_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t2_aplus_delta_start = cursor_t2_aplus_delta_start.fetchall()
                
                cursor_t3_aplus_delta_start = connection.cursor()
                cursor_t3_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t3_aplus_delta_start = cursor_t3_aplus_delta_start.fetchall() 
                
                cursor_t4_aplus_delta_start = connection.cursor()
                cursor_t4_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t4_aplus_delta_start = cursor_t4_aplus_delta_start.fetchall()

                cursor_t0_aplus_delta_end = connection.cursor()
                cursor_t0_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t0_aplus_delta_end = cursor_t0_aplus_delta_end.fetchall()
                
                cursor_t1_aplus_delta_end = connection.cursor()
                cursor_t1_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t1_aplus_delta_end = cursor_t1_aplus_delta_end.fetchall()
                
                cursor_t2_aplus_delta_end = connection.cursor()
                cursor_t2_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t2_aplus_delta_end = cursor_t2_aplus_delta_end.fetchall()
                
                cursor_t3_aplus_delta_end = connection.cursor()
                cursor_t3_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t3_aplus_delta_end = cursor_t3_aplus_delta_end.fetchall() 
                
                cursor_t4_aplus_delta_end = connection.cursor()
                cursor_t4_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t4_aplus_delta_end = cursor_t4_aplus_delta_end.fetchall()
                
                cursor_t0_rplus_delta_start = connection.cursor()
                cursor_t0_rplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t0_rplus_delta_start = cursor_t0_rplus_delta_start.fetchall()
                
                cursor_t0_rplus_delta_end = connection.cursor()
                cursor_t0_rplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t0_rplus_delta_end = cursor_t0_rplus_delta_end.fetchall()
                
                
#                data_table = []
                for x in range(len(data_table_t0_aplus_delta_end)):
                    data_table_temp = []

                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][2])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][1] - data_table_t0_aplus_delta_start[x][1] )
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end[x][1] - data_table_t1_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end[x][1] - data_table_t2_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end[x][1] - data_table_t3_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end[x][1] - data_table_t4_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")                    
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_start[x][1]) # Показания R+ начальные
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end[x][1]) # Показания R+ конечные
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end[x][1] - data_table_t0_rplus_delta_start[x][1]) # Показания R+ разница
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    
                    try:
                        data_table_temp.append(common_sql.get_k_t_t(obj_title)) # Коэффициент трансформации тока параметр 20
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(common_sql.get_k_t_n(obj_title)) # Коэффициент трансформации напряжения параметр 21
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[12]) # Энергия А+ параметр 22
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")

                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[19]) # Энергия R+ параметр 23
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                           
                    data_table.append(data_table_temp)
                    
                    
                    
                request.session["data_table_export"] = data_table
                
                               
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                              ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    # delta for groups abonents 'start date'
                    cursor_t0_aplus_delta_start_temp = connection.cursor()
                    cursor_t0_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t0_aplus_delta_start_temp = cursor_t0_aplus_delta_start_temp.fetchall()
                
                    cursor_t1_aplus_delta_start_temp = connection.cursor()
                    cursor_t1_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t1_aplus_delta_start_temp = cursor_t1_aplus_delta_start_temp.fetchall()
                
                    cursor_t2_aplus_delta_start_temp = connection.cursor()
                    cursor_t2_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t2_aplus_delta_start_temp = cursor_t2_aplus_delta_start_temp.fetchall()
                
                    cursor_t3_aplus_delta_start_temp = connection.cursor()
                    cursor_t3_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t3_aplus_delta_start_temp = cursor_t3_aplus_delta_start_temp.fetchall() 
                
                    cursor_t4_aplus_delta_start_temp = connection.cursor()
                    cursor_t4_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t4_aplus_delta_start_temp = cursor_t4_aplus_delta_start_temp.fetchall()
                    
                    # delta for groups abonents 'end date'
                    cursor_t0_aplus_delta_end_temp = connection.cursor()
                    cursor_t0_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_delta_end_temp = cursor_t0_aplus_delta_end_temp.fetchall()
                
                    cursor_t1_aplus_delta_end_temp = connection.cursor()
                    cursor_t1_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_delta_end_temp = cursor_t1_aplus_delta_end_temp.fetchall()
                
                    cursor_t2_aplus_delta_end_temp = connection.cursor()
                    cursor_t2_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_delta_end_temp = cursor_t2_aplus_delta_end_temp.fetchall()
                
                    cursor_t3_aplus_delta_end_temp = connection.cursor()
                    cursor_t3_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_delta_end_temp = cursor_t3_aplus_delta_end_temp.fetchall() 
                
                    cursor_t4_aplus_delta_end_temp = connection.cursor()
                    cursor_t4_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_delta_end_temp = cursor_t4_aplus_delta_end_temp.fetchall()
                    
                    cursor_t0_rplus_delta_start_temp = connection.cursor()
                    cursor_t0_rplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t0_rplus_delta_start_temp = cursor_t0_rplus_delta_start_temp.fetchall()
                    
                    cursor_t0_rplus_delta_end_temp = connection.cursor()
                    cursor_t0_rplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_rplus_delta_end_temp = cursor_t0_rplus_delta_end_temp.fetchall()
                    
                    data_table_temp = []
                    data_table_temp.append(abonents_list[x][0])
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][6])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1] - data_table_t0_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end_temp[0][1] - data_table_t1_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end_temp[0][1] - data_table_t2_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end_temp[0][1] - data_table_t3_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end_temp[0][1] - data_table_t4_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_start_temp[0][1]) # Показания R+ начальные
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1]) # Показания R+ конечные
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1] - data_table_t0_rplus_delta_start_temp[0][1]) # Показания R+ разница параметр 19
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(common_sql.get_k_t_t(abonents_list[x][0])) # Коэффициент трансформации тока параметр 20
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(common_sql.get_k_t_n(abonents_list[x][0])) # Коэффициент трансформации напряжения параметр 21
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[12]) # Энергия А+ параметр 22
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")

                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[19]) # Энергия R+ параметр 23
                    except IndexError:
                        data_table_temp.append(u"Н/Д")
                    except TypeError:
                        data_table_temp.append(u"Н/Д")
                           
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table                 
#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    return render_to_response("data_table/electric/17.html", args)
       

def electric_potreblenie_3_zones_v2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u'1'
    electric_data_start = u''
    electric_data_end = u''
    
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            res='Электричество'

            if (is_electric_monthly=="1"):
                dm='monthly'
            else:
                dm='daily'
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                    isAbon=True                    
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
                
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                    isAbon=False
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
            #*********************************************************************************************************************************************************************
            elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
                    data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
                    request.session["data_table_export"] = data_table
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    return render_to_response("data_table/electric/17.html", args)
    

def electric_potreblenie_2_zones_v2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u'1'
    electric_data_start = u''
    electric_data_end = u''
    
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            res='Электричество'
            if (is_electric_monthly=="1"):
                dm='monthly'
            else:
                dm='daily'
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                    isAbon=True                    
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
                
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                    isAbon=False
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
            #*********************************************************************************************************************************************************************
            elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
                    data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
                    request.session["data_table_export"] = data_table
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    return render_to_response("data_table/electric/31.html", args)

def electric_simple_2_zones(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_2_zones(obj_title, obj_parent_title, electric_data_end)

                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_by_date_daily_2_zones(obj_title, obj_parent_title, electric_data_end)


            elif (is_electric_current == "1") & (bool(is_abonent_level.search(obj_key))):
                pass
                            
            elif (is_electric_period == "1") & (is_electric_daily =="1") & (bool(is_abonent_level.search(obj_key))): # pokazaniya za period

                end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
                start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
                dates = [x for x in common_sql.daterange(start_date,
                              end_date,
                              step=datetime.timedelta(days=1),
                              inclusive=True)]
                '''for x in range(len(dates)):
                    data_table_temp = [dates[x], dates[x], datetime.datetime.strftime(dates[x], "%d.%m.%Y")]
                    data_table.append(data_table_temp)'''

                for x in range(len(dates)):
                    data_table_temp = common_sql.get_data_table_by_date_daily_2_zones(obj_title, obj_parent_title, datetime.datetime.strftime(dates[x], "%d.%m.%Y"))
                    data_table.extend(data_table_temp)
                #------------

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # monthly for abonents group
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    # query data for each abonent
                    data_table_temp = common_sql.get_data_table_by_date_monthly_2_zones(abonents_list[x][0], obj_title, electric_data_end)
                    if not data_table_temp:
                        data_table_temp = [[electric_data_end, abonents_list[x][0], u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
                    data_table.extend(data_table_temp)
#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1'): # daily for abonents group
               
                    
                if (bool(is_object_level.search(obj_key))):
                    cursor_abonents_list = connection.cursor()
                    cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title])
                    abonents_list = cursor_abonents_list.fetchall()
                            
                for x in range(len(abonents_list)):
                    data_table_temp = common_sql.get_data_table_by_date_daily_2_zones(abonents_list[x][0], obj_title, electric_data_end)
                    if not data_table_temp:
                        data_table_temp = [[electric_data_end, abonents_list[x][0], u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
                    data_table.extend(data_table_temp)

            elif (is_electric_current == '1') & (bool(is_object_level.search(obj_key))): # текущие для объекта учёта
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                              ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    data_table_temp = []
                    data_table_temp.append(u'Дата')
                    data_table_temp.append(abonents_list[x][0])
                    data_table_temp.append(u'Какой-то заводской номер')
                    data_table_temp.append(0)
                    data_table_temp.append(100)
                    data_table_temp.append(200)
                    data_table_temp.append(300)
                    data_table_temp.append(400)
                    data_table.append(data_table_temp)
#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render_to_response("data_table/electric/14.html", args)

def electric_between_3_zones(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    data_table=[]
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end)
            else:
                pass
            
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    
    return render_to_response("data_table/electric/29.html", args)

def electric_between_2_zones(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    data_table=[]
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end)
            else:
                pass
            
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    
    return render_to_response("data_table/electric/27.html", args)

def electric_between(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    data_table=[]
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end)
            else:
                pass
            
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render_to_response("data_table/electric/25.html", args)

def electric_simple_2_zones_v2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'monthly')

                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'daily')


            elif (is_electric_current == "1") & (bool(is_abonent_level.search(obj_key))):
                pass
                            
            elif (is_electric_period == "1") & (is_electric_daily =="1") & (bool(is_abonent_level.search(obj_key))): # pokazaniya za period
                pass
                #------------

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # показания на начало месяца для объекта
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        

#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for abonents group
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]

            elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
              
            elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]

            elif (is_electric_current == '1') & (bool(is_object_level.search(obj_key))): # текущие для объекта учёта
                    pass
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    
    return render_to_response("data_table/electric/14.html", args)


def electric_simple_3_zones(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones(obj_title, obj_parent_title, electric_data_end)

                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_by_date_daily_3_zones(obj_title, obj_parent_title, electric_data_end)


            elif (is_electric_current == "1") & (bool(is_abonent_level.search(obj_key))):
                pass
                            
            elif (is_electric_period == "1") & (is_electric_daily =="1") & (bool(is_abonent_level.search(obj_key))): # pokazaniya za period

                end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
                start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
                dates = [x for x in common_sql.daterange(start_date,
                              end_date,
                              step=datetime.timedelta(days=1),
                              inclusive=True)]
                '''for x in range(len(dates)):
                    data_table_temp = [dates[x], dates[x], datetime.datetime.strftime(dates[x], "%d.%m.%Y")]
                    data_table.append(data_table_temp)'''

                for x in range(len(dates)):
                    data_table_temp = common_sql.get_data_table_by_date_daily_3_zones(obj_title, obj_parent_title, datetime.datetime.strftime(dates[x], "%d.%m.%Y"))
                    data_table.extend(data_table_temp)
                #------------

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # показания на начало месяца для объекта
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s 
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title]) 
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    # query data for each abonent
                    data_table_temp = common_sql.get_data_table_by_date_monthly_3_zones(abonents_list[x][0], obj_title, electric_data_end)
                    if not data_table_temp:
                        data_table_temp = [[electric_data_end, abonents_list[x][0], u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        
                    data_table.extend(data_table_temp)
#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1'): # daily for abonents group
                    
                if (bool(is_object_level.search(obj_key))):
                    cursor_abonents_list = connection.cursor()
                    cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title])
                    abonents_list = cursor_abonents_list.fetchall()

                elif(bool(is_group_level.search(obj_key))):
                    cursor_abonents_list = connection.cursor()
                    cursor_abonents_list.execute("""
                                              SELECT 
                                                  abonents.name
                                                FROM 
                                                  public.balance_groups, 
                                                  public.abonents, 
                                                  public.link_balance_groups_meters, 
                                                  public.meters, 
                                                  public.link_abonents_taken_params, 
                                                  public.taken_params
                                                WHERE 
                                                  balance_groups.guid = link_balance_groups_meters.guid_balance_groups AND
                                                  meters.guid = link_balance_groups_meters.guid_meters AND
                                                  meters.guid = taken_params.guid_meters AND
                                                  link_abonents_taken_params.guid_abonents = abonents.guid AND
                                                  link_abonents_taken_params.guid_taken_params = taken_params.guid AND
                                                  balance_groups.name = %s
                                                GROUP BY
                                                 abonents.name
                                                ORDER BY
                                                  abonents.name ASC;""",[obj_title])
                    abonents_list = cursor_abonents_list.fetchall()
                    obj_title_group = obj_title
                            
                for x in range(len(abonents_list)):
                    #print abonents_list[x][0]
                    if bool(is_group_level.search(obj_key)):
                        cursor_obj_title = connection.cursor()
                        cursor_obj_title.execute("""
                                              SELECT 
                                              objects.name
                                            FROM 
                                              public.abonents, 
                                              public.balance_groups, 
                                              public.objects, 
                                              public.link_balance_groups_meters, 
                                              public.meters, 
                                              public.taken_params, 
                                              public.link_abonents_taken_params
                                            WHERE 
                                              abonents.guid_objects = objects.guid AND
                                              link_balance_groups_meters.guid_balance_groups = balance_groups.guid AND
                                              link_balance_groups_meters.guid_meters = meters.guid AND
                                              taken_params.guid_meters = meters.guid AND
                                              link_abonents_taken_params.guid_abonents = abonents.guid AND
                                              link_abonents_taken_params.guid_taken_params = taken_params.guid AND
                                              balance_groups.name = %s AND 
                                              abonents.name = %s
                                             GROUP BY
                                             objects.name;""",[obj_title_group, abonents_list[x][0]])
                        obj_title = cursor_obj_title.fetchall()
                        obj_title = obj_title[0][0]
                        #print obj_title
                    data_table_temp = common_sql.get_data_table_by_date_daily_3_zones(abonents_list[x][0], obj_title, electric_data_end)

                    if not data_table_temp:
                        data_table_temp = [[electric_data_end, abonents_list[x][0], u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]                    
                    data_table.extend(data_table_temp)
              


            elif (is_electric_current == '1') & (bool(is_object_level.search(obj_key))): # текущие для объекта учёта
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                              ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)): # Заполняем табличку тестовыми значениями
                    data_table_temp = []
                    data_table_temp.append(u'Дата')
                    data_table_temp.append(abonents_list[x][0])
                    data_table_temp.append(u'Какой-то заводской номер')
                    data_table_temp.append(0)
                    data_table_temp.append(100)
                    data_table_temp.append(200)
                    data_table_temp.append(300)
                    data_table_temp.append(400)
                    data_table.append(data_table_temp)

#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render_to_response("data_table/electric/16.html", args)
    
def electric_simple_3_zones_v2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'monthly')

                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'daily')


            elif (is_electric_current == "1") & (bool(is_abonent_level.search(obj_key))):
                pass
                            
            elif (is_electric_period == "1") & (is_electric_daily =="1") & (bool(is_abonent_level.search(obj_key))): # pokazaniya za period
                pass
                #------------

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # показания на начало месяца для объекта
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        

#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for abonents group
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]

            elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
              
            elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]

            elif (is_electric_current == '1') & (bool(is_object_level.search(obj_key))): # текущие для объекта учёта
                    pass

#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render_to_response("data_table/electric/16.html", args)
    
    
#________________-
def pokazaniya_heat(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    
    list_except = [u'ВРУ Счётчик01',u'ВРУ Счётчик02',u'ВРУ Счётчик03',u'ВРУ Счётчик04',u'ВРУ Счётчик05',u'ВРУ Счётчик06',u'ВРУ Счётчик07',u'ВРУ Счётчик08',u'ВРУ Счётчик09',u'ВРУ Счётчик10',u'ВРУ Счётчик11',u'ВРУ Счётчик12',u'ВРУ Счётчик13',u'ВРУ Счётчик14',u'ВРУ Счётчик15',u'ВРУ Счётчик16',u'ВРУ Счётчик17',u'ВРУ Счётчик18',u'ВРУ Счётчик19',u'ВРУ Счётчик20',u'ВРУ Счётчик21',u'ВРУ Счётчик22',u'ВРУ Счётчик23',u'Гараж Счётчик 1',u'Гараж Счётчик 2']
                     
    if (bool(is_abonent_level.search(obj_key))):     
        data_table = common_sql.get_data_table_by_date_heat(meters_name, parent_name, electric_data_end)

    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []
        for x in range(len(list_of_abonents_2)):
            data_table_temp = common_sql.get_data_table_by_date_heat(list_of_abonents_2[x], meters_name, electric_data_end)

            if list_of_abonents_2[x][0] in list_except:
                next
            elif data_table_temp:            
                data_table.extend(data_table_temp)
            else:
                data_table.extend([[0,list_of_abonents_2[x][0],u'Н/Д',u'Н/Д',u'Н/Д']])
                
              
    else:
        data_table = []
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 

    return render_to_response("data_table/heat/18.html", args)

def pokazaniya_heat_v2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    data_table = []
    list_except = []
    if (bool(is_abonent_level.search(obj_key))):     
        data_table = common_sql.get_data_table_by_date_heat_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_v2(meters_name, parent_name, electric_data_end, False)
        for row in data_table:
            for x in list_except:
                if x==row[2]:
                    data_table.remove(x)


    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
    return render_to_response("data_table/heat/18.html", args)

def potreblenie_heat(request): 
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']
    electric_data_start   = request.GET['electric_data_start']                        
    obj_key             = request.GET['obj_key']
    list_except = [u'ВРУ Счётчик01',u'ВРУ Счётчик02',u'ВРУ Счётчик03',u'ВРУ Счётчик04',u'ВРУ Счётчик05',u'ВРУ Счётчик06',u'ВРУ Счётчик07',u'ВРУ Счётчик08',u'ВРУ Счётчик09',u'ВРУ Счётчик10',u'ВРУ Счётчик11',u'ВРУ Счётчик12',u'ВРУ Счётчик13',u'ВРУ Счётчик14',u'ВРУ Счётчик15',u'ВРУ Счётчик16',u'ВРУ Счётчик17',u'ВРУ Счётчик18',u'ВРУ Счётчик19',u'ВРУ Счётчик20',u'ВРУ Счётчик21',u'ВРУ Счётчик22',u'ВРУ Счётчик23',u'Гараж Счётчик 1',u'Гараж Счётчик 2']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table_end = common_sql.get_data_table_by_date_heat(meters_name, parent_name, electric_data_end)
        data_table_start = common_sql.get_data_table_by_date_heat(meters_name, parent_name, electric_data_start)
        data_table = []
        for x in range(len(data_table_end)):
            try:
                data_table_temp=[data_table_end[x][0], data_table_end[x][1], data_table_end[x][2], data_table_start[x][3], data_table_end[x][3], data_table_end[x][3]-data_table_start[x][3], data_table_end[x][5] - data_table_start[x][5]]
                data_table.append(data_table_temp)
            except:
                data_table = []
            

    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []
        for x in range(len(list_of_abonents_2)):
            data_table_end_temp = common_sql.get_data_table_by_date_heat(list_of_abonents_2[x][0], meters_name, electric_data_end)
            data_table_start_temp = common_sql.get_data_table_by_date_heat(list_of_abonents_2[x][0], meters_name, electric_data_start)
            data_table_temp = []
            for x in range(len(data_table_end_temp)):

                data_table_temp_2 = []
                try:
                    data_table_temp_2.append(data_table_end_temp[x][0])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_end_temp[x][1])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_end_temp[x][2])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_start_temp[x][3])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_end_temp[x][3])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_end_temp[x][3]-data_table_start_temp[x][3])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_end_temp[x][5]-data_table_start_temp[x][5])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")

                data_table_temp.append(data_table_temp_2)
            data_table_end_temp = []
            data_table_start_temp = []
            

            if list_of_abonents_2[x][0] in list_except:
                next
            elif data_table_temp:            
                data_table.extend(data_table_temp)
            else:
                data_table.extend([[0,list_of_abonents_2[x][0],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
                
              
    else:
        data_table = []
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = meters_name 


    return render_to_response("data_table/heat/19.html", args)
#--------------------------------------------
def potreblenie_heat_v2(request): 
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']
    electric_data_start = request.GET['electric_data_start']
    obj_key             = request.GET['obj_key']
    list_except = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_for_period_for_abon_heat_v2(meters_name, parent_name, electric_data_start, electric_data_end)

    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_for_period_heat_v2(meters_name, parent_name, electric_data_start, electric_data_end)

    else:
        data_table = []
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = meters_name 


    return render_to_response("data_table/heat/19.html", args)
    
def pokazaniya_heat_current(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key'] 
    list_except = [u'ВРУ Счётчик01',u'ВРУ Счётчик02',u'ВРУ Счётчик03',u'ВРУ Счётчик04',u'ВРУ Счётчик05',u'ВРУ Счётчик06',u'ВРУ Счётчик07',u'ВРУ Счётчик08',u'ВРУ Счётчик09',u'ВРУ Счётчик10',u'ВРУ Счётчик11',u'ВРУ Счётчик12',u'ВРУ Счётчик13',u'ВРУ Счётчик14',u'ВРУ Счётчик15',u'ВРУ Счётчик16',u'ВРУ Счётчик17',u'ВРУ Счётчик18',u'ВРУ Счётчик19',u'ВРУ Счётчик20',u'ВРУ Счётчик21',u'ВРУ Счётчик22',u'ВРУ Счётчик23',u'Гараж Счётчик 1',u'Гараж Счётчик 2']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_current_heat(meters_name, parent_name)
    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []
        for x in range(len(list_of_abonents_2)):
            data_table_temp = common_sql.get_data_table_current_heat(list_of_abonents_2[x], meters_name)
            
            if list_of_abonents_2[x][0] in list_except:
                next
            elif data_table_temp:            
                data_table.extend(data_table_temp)
            else:
                data_table.extend([[u'Н/Д',u'Н/Д',list_of_abonents_2[x][0],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
                
              
    else:
        data_table = []
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 

    return render_to_response("data_table/heat/20.html", args)


def pokazaniya_heat_current_v2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key'] 
    list_except = []
    data_table=[]
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']

    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_current_heat_v2(meters_name, parent_name, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_current_heat_v2(meters_name, parent_name, False)
        for row in data_table:
            for x in list_except:
                if x==row[2]:
                    data_table.remove(x)

    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
    return render_to_response("data_table/heat/20.html", args)

# Test SPG
def pokazaniya_spg(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = get_data_table_by_date_spg(meters_name, parent_name, electric_data_end)
    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents_heat(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []
        for x in range(len(list_of_abonents_2)):
            data_table_temp = get_data_table_by_date_spg(list_of_abonents_2[x], parent_name, electric_data_end)
            if data_table_temp:            
                data_table.extend(data_table_temp)
            else:
                data_table.extend([[electric_data_end,list_of_abonents_2[x][0],u'Н/Д',u'Н/Д',u'Н/Д']])
                
              
    else:
        data_table = [] 
        
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 

    return render_to_response("data_table/gas/22.html", args)
    
def pokazaniya_sayany(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_by_date_heat_sayany(meters_name, parent_name, electric_data_end)
    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents_heat(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []
        for x in range(len(list_of_abonents_2)):

            data_table_temp = common_sql.get_data_table_by_date_heat_sayany(list_of_abonents_2[x], meters_name, electric_data_end)
            if data_table_temp:            
                data_table.extend(data_table_temp)
            else:
                data_table.extend([[electric_data_end,list_of_abonents_2[x][0],u'Н/Д',u'Н/Д',u'Н/Д']])
                
              
    else:
        data_table = [] 
        
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
      
    return render_to_response("data_table/heat/30.html", args)
    
def pokazaniya_sayany_v2(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, False)

#    for i in range(len(data_table)):
#        data_table[i]=list(data_table[i])
#        if (data_table[i][3] is None):
#            print data_table[i][1], meters_name
#            data_table[i][0]=electric_data_end
#            dt=common_sql.get_data_table_by_date_heat_sayany_v2(data_table[i][1], meters_name, None, True)
#            if (len(dt)>0):
#                data_table[i]=dt[0]
#        data_table[i]=tuple(data_table[i])
    
    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
      
    return render_to_response("data_table/heat/30.html", args)
    
def pokazaniya_sayany_last(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, False)


    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        if (data_table[i][3] is None):
            #print data_table[i][1], meters_name
            data_table[i][0]=electric_data_end
            dt=common_sql.get_data_table_by_date_heat_sayany_v2(data_table[i][1], meters_name, None, True)
            if (len(dt)>0):
                data_table[i]=dt[0]
        data_table[i]=tuple(data_table[i])
    
    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
      
    return render_to_response("data_table/heat/32.html", args)
    
def heat_potreblenie_sayany(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']

    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_period_heat_sayany(meters_name, parent_name,electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_period_heat_sayany(meters_name, parent_name,electric_data_start, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = meters_name 
      
    return render_to_response("data_table/heat/33.html", args)

def electric_check_factory_numbers(request):
    args= {}
    
    data_table = []
    data_table = common_sql.get_data_table_diferent_numbers()
    
    args['data_table'] = data_table
    
    return render_to_response("data_table/electric/40.html", args)

def water_by_date(request):
    args= {}
    is_abonent_level = re.compile(r'level2')
    is_object_level_2 = re.compile(r'level1')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']

    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_by_date(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_by_date(meters_name, parent_name, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
      
    return render_to_response("data_table/water/38.html", args)
    
def water_potreblenie_pulsar(request):
    args= {}
    is_abonent_level = re.compile(r'level2')
    is_object_level_2 = re.compile(r'level1')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    electric_data_start   = request.GET['electric_data_start']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_end"]   = electric_data_start   = request.GET['electric_data_start']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']

    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_period_pulsar(meters_name, parent_name,electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_period_pulsar(meters_name, parent_name,electric_data_start, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        num=data_table[i][3]
        if ('ХВС, №' in num) or ('ГВС, №' in num):
            num=num.replace(u'ХВС, №', ' ')
            num=num.replace(u'ГВС, №', ' ')
            data_table[i][3]=num
            #print num
        data_table[i]=tuple(data_table[i])
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = meters_name 
      
    return render_to_response("data_table/water/39.html", args)


def pokazaniya_water_hvs_tekon(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']            
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, u'Канал 1',  u'Tekon_hvs',True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, u'Канал 1',  u'Tekon_hvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
      
    return render_to_response("data_table/water/34.html", args)

def water_potreblenie_hvs_tekon(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name, electric_data_start, electric_data_end, u'Канал 1',  u'Tekon_hvs', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name,electric_data_start, electric_data_end, u'Канал 1',  u'Tekon_hvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = meters_name 
      
    return render_to_response("data_table/water/35.html", args)

def pokazaniya_water_gvs_tekon(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']            
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, u'Канал 2',  u'Tekon_gvs', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, u'Канал 2',  u'Tekon_gvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
      
    return render_to_response("data_table/water/36.html", args)

def water_potreblenie_gvs_tekon(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name, electric_data_start, electric_data_end, u'Канал 2',  u'Tekon_gvs', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name,electric_data_start, electric_data_end, u'Канал 2',  u'Tekon_gvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = meters_name 
      
    return render_to_response("data_table/water/37.html", args)

def tekon_heat_by_date(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']    
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']            
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_heat_daily(meters_name, parent_name, electric_data_end, u'Канал 3',  u'Tekon_heat', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_heat_daily(meters_name, parent_name, electric_data_end, u'Канал 3',  u'Tekon_heat', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
      
    return render_to_response("data_table/heat/50.html", args)

def tekon_heat_potreblenie(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name, electric_data_start, electric_data_end, u'Канал 3',  u'Tekon_heat', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name,electric_data_start, electric_data_end, u'Канал 3',  u'Tekon_heat', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = meters_name 
      
    return render_to_response("data_table/heat/51.html", args)
    
def resources_all(request):
    args= {}
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']            
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']
            data_table = common_sql.get_data_table_report_all_res_period3(electric_data_start, electric_data_end)
            #data_table = common_sql.get_data_table_report_all_res_period(u'10.02.2017', u'20.02.2017')

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
#        
#    #удаляем из номеров счётчиков лишнее
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        num=data_table[i][3]
        if ('ХВС, №' in num) or ('ГВС, №' in num):
            num=num.replace(u'ХВС, №', ' ')
            num=num.replace(u'ГВС, №', ' ')
            data_table[i][3]=num
            #print num
        data_table[i]=tuple(data_table[i])

    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
      
    return render_to_response("data_table/9.html", args)
    
def resources_all_by_date(request):
    args= {}
    electric_data_end   = request.GET['electric_data_end']            
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            data_table = common_sql.get_data_table_report_all_res_by_date(electric_data_end)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
      
    return render_to_response("data_table/42.html", args)
    
def resources_electric_by_date(request):
    args= {}
    electric_data_end   = request.GET['electric_data_end']            
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            data_table = common_sql.get_data_table_report_electric_res_by_date(electric_data_end)

    #zamenyem None na N/D vezde
#    if len(data_table)>0: 
#        data_table=common_sql.ChangeNull(data_table, None)

    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
      
    return render_to_response("data_table/44.html", args)
    
def resources_water_by_date(request):
    args= {}
    electric_data_end   = request.GET['electric_data_end']            
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            data_table = common_sql.get_data_table_report_water_res_by_date(electric_data_end)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
      
    return render_to_response("data_table/46.html", args)
    
def resources_heat_by_date(request):
    args= {}
    electric_data_end   = request.GET['electric_data_end']            
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            data_table = common_sql.get_data_table_report_heat_res_by_date(electric_data_end)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    
      
    return render_to_response("data_table/48.html", args)

def test_test(request):
    args={}
    args['test_test'] = 10
    countAll=300
    
    
    return render_to_response("data_table/test/23.html", args)

def forma_80020(request):
    args= {}
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']            
    
    data_table = []
    data_table_check_data_header = []
    data_table_check_data = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]    = electric_data_end    = request.GET['electric_data_end']
            request.session["electric_data_start"]  = electric_data_start  = request.GET['electric_data_start']
            request.session["obj_title"]            = group_name           = request.GET['obj_title']
            
            # Формируем список дат на основе начальной и конечной даты полученной от web-календарей
            end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
            start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
            list_of_dates = [x for x in common_sql.daterange(start_date,
                          end_date,
                          step=datetime.timedelta(days=1),
                          inclusive=True)]

            # Делаем информационную табличку со счётчиками входящими в данную группу
            data_table = common_sql.get_info_group_80020_meters(group_name)
            #print group_name
            #print u'Это наш data_TABLE --> ', data_table
        
        # Делаем проверку на сумму профилей мощности и размности показаний счётчика
            # Узнаем начальные и конечные показания по счётчику
            # T0 A+ Начальная дата
                
            for y in range(len(data_table)):
                data_table[y] = list(data_table[y])
                my_parametr = u'T0 A+'
                result = common_sql.get_data_table_electric_parametr_daily_by_meters_number(data_table[y][3], electric_data_start, my_parametr)

                if result:
                    data_table[y].append(result[0][0])
                else:
                    data_table[y].append(u'Н/Д')
            # T0 A+ Конечная  дата
            for z in range(len(data_table)):
                data_table[z] = list(data_table[z])
                my_parametr = u'T0 A+'
                result = common_sql.get_data_table_electric_parametr_daily_by_meters_number(data_table[z][3], electric_data_end, my_parametr)
               
                if result:
                    data_table[z].append(result[0][0])
                else:
                    data_table[z].append(u'Н/Д')
                    
            for w in range(len(data_table)):
                try:
                    data_table[w].append(data_table[w][8]-data_table[w][7])
                except:
                    data_table[w].append(u'Н/Д')
                    
            #считаем сумму получасовок А+
            for m in range(len(data_table)):
                sum_of_30_min_a = 0
                for date in range(len(list_of_dates)-1):
                    try:
                        sum_of_30_min_a = sum_of_30_min_a + common_sql.get_sum_of_30_profil_by_meter_number(list_of_dates[date], data_table[m][3], u'A+ Профиль')
                    except:
                        pass
                try:
                    data_table[m].append(sum_of_30_min_a)
                except:
                    data_table[m].append(u'Н/Д')
                
                # Вычисляем процент сбора получасовок
                sum_of_count_30_min = 0
                for j in range (len(list_of_dates)-1):
                    sum_of_count_30_min = sum_of_count_30_min + common_sql.get_count_of_30_profil_by_meter_number(list_of_dates[j], data_table[m][3], u'A+ Профиль')
                data_table[m].append((sum_of_count_30_min*100.0)/((len(list_of_dates)-1)*48))
                    
                                         
                    
            #Заполняем list со значениями нужных параметров
            list_of_taken_params = []
            for x in range(len(data_table)):
                #Получаем считываемые параметры по заводскому номеру прибора.
                 #A+
                name_of_type_meters = common_sql.get_name_of_type_meter_by_serial_number(data_table[x][3])
                guid_params = u''
                if name_of_type_meters[0][0] == u'Меркурий 230-УМ':
                    guid_params = u'922ad57c-8f5e-4f00-a78d-e3ba89ef859f'
                elif name_of_type_meters[0][0] == u'Меркурий 230':
                    guid_params = u'6af9ddce-437a-4e07-bd70-6cf9dcc10b31'
                else:
                    pass
                result = common_sql.get_taken_param_by_meters_number_and_guid_params(data_table[x][3], guid_params)
                list_of_taken_params.append(unicode(result[0][0]) + u' ' + unicode(result[0][1]))
                 #R+
                if name_of_type_meters[0][0] == u'Меркурий 230-УМ':
                     guid_params = u'61101fa3-a96a-4934-9482-e32036c12829'
                elif name_of_type_meters[0][0] == u'Меркурий 230':
                     guid_params = u'66e997c0-8128-40a7-ae65-7e8993fbea61'
                else:
                    pass
                result = common_sql.get_taken_param_by_meters_number_and_guid_params(data_table[x][3], guid_params)
                list_of_taken_params.append(unicode(result[0][0]) + u' ' + unicode(result[0][1]))
                                
            # Добавляем дату в лист с параметрами и делаем таблицу для шапки таблицы 
            list_of_taken_params.insert(0, u'Дата')
            data_table_check_data_header = list_of_taken_params
            
                     
        
        #Проверяем сколько получасовок имеем за каждые сутки промежутка по каждому считываемому параметру
            for x in range(len(list_of_dates)):
                list_of_one_date_check = []
                list_of_one_date_check.append(list_of_dates[x])
                for y in range(1, len(list_of_taken_params)):
                    my_split_params = list_of_taken_params[y].split()
                    my_names_param = my_split_params[1] + u' ' + my_split_params[2]
                    
                    list_of_one_date_check.append(common_sql.get_count_of_30_profil_by_meter_number(list_of_dates[x], my_split_params[0], my_names_param))
                data_table_check_data.append(list_of_one_date_check)
            
            data_table_check_data.pop()
                                                
            
    args['data_table'] = data_table
    args['data_table_check_data_header'] = data_table_check_data_header
    args['data_table_check_data'] = data_table_check_data
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    
    return render_to_response("data_table/electric/41.html", args)
    
def pulsar_heat_period(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            
#*********************************************************************************************************************************************************************
            if (bool(is_abonent_level.search(obj_key))):
                data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title,electric_data_end, electric_data_start, True)
            elif (bool(is_object_level_2.search(obj_key))):
                data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title, electric_data_end,electric_data_start, False)
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render_to_response("data_table/heat/59.html", args)

def pulsar_heat_period_2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            
#*********************************************************************************************************************************************************************
            if (bool(is_abonent_level.search(obj_key))):
                data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title,electric_data_end, electric_data_start, True)
            elif (bool(is_object_level_2.search(obj_key))):
                data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title, electric_data_end,electric_data_start, False)
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render_to_response("data_table/heat/61.html", args)

def pulsar_heat_daily(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']


#*********************************************************************************************************************************************************************
            if (bool(is_abonent_level.search(obj_key))):
                data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, True)
            elif (bool(is_object_level_2.search(obj_key))):
                data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, False)
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render_to_response("data_table/heat/56.html", args)
    
def pulsar_heat_daily_2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']


#*********************************************************************************************************************************************************************
            if (bool(is_abonent_level.search(obj_key))):
                data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, True)
            elif (bool(is_object_level_2.search(obj_key))):
                data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, False)
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render_to_response("data_table/heat/62.html", args)

def pulsar_water_daily(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["obj_parent_title"]    = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']            
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily(obj_parent_title, obj_title, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
   

    return render_to_response("data_table/water/58.html", args)
    
def pulsar_water_period(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["obj_parent_title"]    = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']              
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period(obj_parent_title, obj_title, electric_data_start,electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
   

    return render_to_response("data_table/water/57.html", args)
    
def pulsar_water_daily_row(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["obj_parent_title"]    = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']            
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily_row(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily_row(obj_parent_title, obj_title, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull_for_pulsar(data_table)
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
   

    return render_to_response("data_table/water/60.html", args)
    
def heat_elf_daily(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["obj_parent_title"]    = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']            
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_daily(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_daily(obj_parent_title, obj_title, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
   

    return render_to_response("data_table/heat/64.html", args)
    
def heat_elf_period(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["obj_parent_title"]    = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']   
            request.session["electric_data_end"]   = electric_data_start   = request.GET['electric_data_start'] 
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_elf_period(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_elf_period(obj_parent_title, obj_title, electric_data_start,electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
   

    return render_to_response("data_table/heat/63.html", args)
    
def heat_water_elf_daily(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u''
    electric_data_start = u''
    electric_data_end = u''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["obj_parent_title"]    = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']            
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_water_daily(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_water_daily(obj_parent_title, obj_title, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
   

    return render_to_response("data_table/66.html", args)

#Разработка формы 80040 ___________________---------------------------------------_______________________________
    
def forma_80040(request):
    args= {}
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']            
    
    data_table = []
    data_table_check_data_header = []
    data_table_check_data = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]    = electric_data_end    = request.GET['electric_data_end']
            request.session["electric_data_start"]  = electric_data_start  = request.GET['electric_data_start']
            request.session["obj_title"]            = group_name           = request.GET['obj_title']
            
            # Формируем список дат на основе начальной и конечной даты полученной от web-календарей
            end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
            start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
            list_of_dates = [x for x in common_sql.daterange(start_date,
                          end_date,
                          step=datetime.timedelta(days=1),
                          inclusive=True)]

            # Делаем информационную табличку со счётчиками входящими в данную группу
            data_table = common_sql.get_info_group_80020_meters(group_name)
            #print group_name
            #print u'Это наш data_TABLE --> ', data_table
        
        # Делаем проверку на сумму профилей мощности и размности показаний счётчика
            # Узнаем начальные и конечные показания по счётчику
            # T0 A+ Начальная дата
                
            for y in range(len(data_table)):
                data_table[y] = list(data_table[y])
                my_parametr = u'T0 A+'
                result = common_sql.get_data_table_electric_parametr_daily_by_meters_number(data_table[y][3], electric_data_start, my_parametr)

                if result:
                    data_table[y].append(result[0][0])
                else:
                    data_table[y].append(u'Н/Д')
            # T0 A+ Конечная  дата
            for z in range(len(data_table)):
                data_table[z] = list(data_table[z])
                my_parametr = u'T0 A+'
                result = common_sql.get_data_table_electric_parametr_daily_by_meters_number(data_table[z][3], electric_data_end, my_parametr)
               
                if result:
                    data_table[z].append(result[0][0])
                else:
                    data_table[z].append(u'Н/Д')
                    
            for w in range(len(data_table)):
                try:
                    data_table[w].append(data_table[w][8]-data_table[w][7])
                except:
                    data_table[w].append(u'Н/Д')
                    
            #считаем сумму получасовок А+
            for m in range(len(data_table)):
                sum_of_30_min_a = 0
                for date in range(len(list_of_dates)-1):
                    try:
                        sum_of_30_min_a = sum_of_30_min_a + common_sql.get_sum_of_30_profil_by_meter_number(list_of_dates[date], data_table[m][3], u'A+ Профиль')
                    except:
                        pass
                try:
                    data_table[m].append(sum_of_30_min_a)
                except:
                    data_table[m].append(u'Н/Д')
                
                # Вычисляем процент сбора получасовок
                sum_of_count_30_min = 0
                for j in range (len(list_of_dates)-1):
                    sum_of_count_30_min = sum_of_count_30_min + common_sql.get_count_of_30_profil_by_meter_number(list_of_dates[j], data_table[m][3], u'A+ Профиль')
                try: # Если начальная и конечная даты совпадают, то получается деление на нуль
                    data_table[m].append((sum_of_count_30_min*100.0)/((len(list_of_dates)-1)*48))
                except: # В случае деления на нуль ничего не делаем... Ждем ввода двух разных дат
                    pass
                    
                                         
                    
            #Заполняем list со значениями нужных параметров
            list_of_taken_params = []
            for x in range(len(data_table)):
                #Получаем считываемые параметры по заводскому номеру прибора.
                 #A+
                name_of_type_meters = common_sql.get_name_of_type_meter_by_serial_number(data_table[x][3])
                guid_params = u''
                if name_of_type_meters[0][0] == u'Меркурий 230-УМ':
                    guid_params = u'922ad57c-8f5e-4f00-a78d-e3ba89ef859f'
                elif name_of_type_meters[0][0] == u'Меркурий 230':
                    guid_params = u'6af9ddce-437a-4e07-bd70-6cf9dcc10b31'
                else:
                    pass
                result = common_sql.get_taken_param_by_meters_number_and_guid_params(data_table[x][3], guid_params)
                list_of_taken_params.append(unicode(result[0][0]) + u' ' + unicode(result[0][1]))
                 #R+
                if name_of_type_meters[0][0] == u'Меркурий 230-УМ':
                     guid_params = u'61101fa3-a96a-4934-9482-e32036c12829'
                elif name_of_type_meters[0][0] == u'Меркурий 230':
                     guid_params = u'66e997c0-8128-40a7-ae65-7e8993fbea61'
                else:
                    pass
                result = common_sql.get_taken_param_by_meters_number_and_guid_params(data_table[x][3], guid_params)
                list_of_taken_params.append(unicode(result[0][0]) + u' ' + unicode(result[0][1]))
                                
            # Добавляем дату в лист с параметрами и делаем таблицу для шапки таблицы 
            list_of_taken_params.insert(0, u'Дата')
            data_table_check_data_header = list_of_taken_params
            
                     
        
        #Проверяем сколько получасовок имеем за каждые сутки промежутка по каждому считываемому параметру
            for x in range(len(list_of_dates)):
                list_of_one_date_check = []
                list_of_one_date_check.append(list_of_dates[x])
                for y in range(1, len(list_of_taken_params)):
                    my_split_params = list_of_taken_params[y].split()
                    my_names_param = my_split_params[1] + u' ' + my_split_params[2]
                    
                    list_of_one_date_check.append(common_sql.get_count_of_30_profil_by_meter_number(list_of_dates[x], my_split_params[0], my_names_param))
                data_table_check_data.append(list_of_one_date_check)
            
            data_table_check_data.pop()
                                                
            
    args['data_table'] = data_table
    args['data_table_check_data_header'] = data_table_check_data_header
    args['data_table_check_data'] = data_table_check_data
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    
    return render_to_response("data_table/electric/71.html", args)
    