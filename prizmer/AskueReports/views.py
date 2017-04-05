# coding -*- coding: utf-8 -*-

#from django.shortcuts import render
from __future__ import unicode_literals
from django.shortcuts import render_to_response, HttpResponse
from django.db import connection
import StringIO
from openpyxl import Workbook
from openpyxl.compat import range
import datetime
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Font
from openpyxl.cell import get_column_letter
import common_sql
import re
from excel_response import ExcelResponse
from django.shortcuts import redirect
#from datetime import datetime, date, time

# для работы с xml
from lxml import etree


def zagotovka(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    ws['B5'] = 'Заготовка'
    ws['B5'].style = ali_grey
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'zagotovka' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

# Стили
ali_grey   = Style(fill=PatternFill(fill_type='solid', start_color='DCDCDC'), border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))
ali_white  = Style(border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))
ali_yellow = Style(fill=PatternFill(fill_type='solid', start_color='EEEE00'), border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))
ali_white_size_18  = Style(font=Font(size=18))
# Конец описания стилей

def translate(name):
 
    #Заменяем пробелы и преобразуем строку к нижнему регистру
    name = name.replace(' ','-').lower()
 
    #
    transtable = (
        ## Большие буквы
        (u"Щ", u"Sch"),
        (u"Щ", u"SCH"),
        # two-symbol
        (u"Ё", u"Yo"),
        (u"Ё", u"YO"),
        (u"Ж", u"Zh"),
        (u"Ж", u"ZH"),
        (u"Ц", u"Ts"),
        (u"Ц", u"TS"),
        (u"Ч", u"Ch"),
        (u"Ч", u"CH"),
        (u"Ш", u"Sh"),
        (u"Ш", u"SH"),
        (u"Ы", u"Yi"),
        (u"Ы", u"YI"),
        (u"Ю", u"Yu"),
        (u"Ю", u"YU"),
        (u"Я", u"Ya"),
        (u"Я", u"YA"),
        # one-symbol
        (u"А", u"A"),
        (u"Б", u"B"),
        (u"В", u"V"),
        (u"Г", u"G"),
        (u"Д", u"D"),
        (u"Е", u"E"),
        (u"З", u"Z"),
        (u"И", u"I"),
        (u"Й", u"J"),
        (u"К", u"K"),
        (u"Л", u"L"),
        (u"М", u"M"),
        (u"Н", u"N"),
        (u"О", u"O"),
        (u"П", u"P"),
        (u"Р", u"R"),
        (u"С", u"S"),
        (u"Т", u"T"),
        (u"У", u"U"),
        (u"Ф", u"F"),
        (u"Х", u"H"),
        (u"Э", u"E"),
        (u"Ъ", u"`"),
        (u"Ь", u"'"),
        ## Маленькие буквы
        # three-symbols
        (u"щ", u"sch"),
        # two-symbols
        (u"ё", u"yo"),
        (u"ж", u"zh"),
        (u"ц", u"ts"),
        (u"ч", u"ch"),
        (u"ш", u"sh"),
        (u"ы", u"yi"),
        (u"ю", u"yu"),
        (u"я", u"ya"),
        # one-symbol
        (u"а", u"a"),
        (u"б", u"b"),
        (u"в", u"v"),
        (u"г", u"g"),
        (u"д", u"d"),
        (u"е", u"e"),
        (u"з", u"z"),
        (u"и", u"i"),
        (u"й", u"j"),
        (u"к", u"k"),
        (u"л", u"l"),
        (u"м", u"m"),
        (u"н", u"n"),
        (u"о", u"o"),
        (u"п", u"p"),
        (u"р", u"r"),
        (u"с", u"s"),
        (u"т", u"t"),
        (u"у", u"u"),
        (u"ф", u"f"),
        (u"х", u"h"),
        (u"э", u"e"),
    )
    #перебираем символы в таблице и заменяем
    for symb_in, symb_out in transtable:
        name = name.replace(symb_in, symb_out)
    #возвращаем переменную
    return name

def get_k_t_n_by_serial_number(serial_number):
    """Получаем Ктн по серийному номеру счтчика"""
    simpleq = connection.cursor()
    simpleq.execute("""SELECT 
                          link_abonents_taken_params.coefficient_2
                        FROM 
                          public.meters, 
                          public.link_abonents_taken_params, 
                          public.taken_params
                        WHERE 
                          link_abonents_taken_params.guid_taken_params = taken_params.guid AND
                          taken_params.guid_meters = meters.guid AND
                          meters.factory_number_manual = %s
                        ORDER BY
                          link_abonents_taken_params.coefficient_2 DESC
                        LIMIT 1;""", [serial_number])
    simpleq = simpleq.fetchall()
    return simpleq[0][0]
    

    
def get_k_t_t_by_serial_number(serial_number):
    """Получаем Ктт по серийному номеру счтчика"""
    simpleq = connection.cursor()
    simpleq.execute("""SELECT 
                          link_abonents_taken_params.coefficient
                        FROM 
                          public.meters, 
                          public.link_abonents_taken_params, 
                          public.taken_params
                        WHERE 
                          link_abonents_taken_params.guid_taken_params = taken_params.guid AND
                          taken_params.guid_meters = meters.guid AND
                          meters.factory_number_manual = %s
                        ORDER BY
                          link_abonents_taken_params.coefficient DESC
                        LIMIT 1;""", [serial_number])
    simpleq = simpleq.fetchall()
    return simpleq[0][0]
    
def get_k_a_by_serial_number(serial_number):
    """Получаем Коэффициент А по серийному номеру счтчика"""
    simpleq = connection.cursor()
    simpleq.execute("""SELECT 
                          link_abonents_taken_params.coefficient_3
                        FROM 
                          public.meters, 
                          public.link_abonents_taken_params, 
                          public.taken_params
                        WHERE 
                          link_abonents_taken_params.guid_taken_params = taken_params.guid AND
                          taken_params.guid_meters = meters.guid AND
                          meters.factory_number_manual = %s
                        ORDER BY
                          link_abonents_taken_params.coefficient DESC
                        LIMIT 1;""", [serial_number])
    simpleq = simpleq.fetchall()
    return simpleq[0][0]


def report_3_tarifa_k(request): # Отчет по А+ и R+ с коэффициентами

    response = StringIO.StringIO()    
    wb = Workbook()    
    ws = wb.active                   

# Шапка отчета   
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Срез показаний счетчика с коэффициентами за период' + ' ' + str(request.session["electric_data_start"]) + " - " + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктт'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктн'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма A+
    ws.merge_cells('F4:I4')
    ws['F4'] = 'Суммарные показания/энергия A+'
    ws['F4'].style = ali_grey
    ws['G4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['F5'] = 'Показания на ' + str(request.session["electric_data_start"])
    ws['F5'].style = ali_grey
    ws['G5'] = 'Показания на ' + str(request.session["electric_data_end"])
    ws['G5'].style = ali_grey
    ws['H5'] = 'Разность показаний'
    ws['H5'].style = ali_grey
    ws['I5'] = 'Энергия кВт*ч'
    ws['I5'].style = ali_grey
    
    # Сумма R+
    ws.merge_cells('J4:M4')
    ws['J4'] = 'Суммарные показания/энергия R+'
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['L4'].style = ali_grey
    ws['M4'].style = ali_grey
    ws['J5'] = 'Показания на ' + str(request.session["electric_data_start"])
    ws['J5'].style = ali_grey
    ws['K5'] = 'Показания на ' + str(request.session["electric_data_end"])
    ws['K5'].style = ali_grey
    ws['L5'] = 'Разность показаний'
    ws['L5'].style = ali_grey
    ws['M5'] = 'Энергия кВт*ч'
    ws['M5'].style = ali_grey
    
   
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17
    
# Заполняем таблицу данными ----------------------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------------------------------------------------
    is_abonent_level    = re.compile(r'abonent')
    is_object_level     = re.compile(r'level')    
    obj_title           = request.session["obj_title"]
    electric_data_start = request.session["electric_data_start"]
    electric_data_end   = request.session["electric_data_end"]
    obj_key             = request.session["obj_key"]
    data_table = []
    
#--------------------------------------------------------------------------------------------------------------------------------------------------------------
    if bool(is_object_level.search(obj_key)): # Если это объект, то формируем список абонентов
        cursor_abonents_list = connection.cursor()
        cursor_abonents_list.execute("""
                                  SELECT 
                                   abonents.name
                                  FROM 
                                   public.objects, 
                                   public.abonents
                                  WHERE 
                                   objects.guid = abonents.guid_objects AND
                                   objects.name = %s
                                  ORDER BY
                                   abonents.name ASC;""",[obj_title])
        abonents_list = cursor_abonents_list.fetchall()

    elif bool(is_abonent_level.search(obj_key)):   # Если это отдельный абонент, то делаем выборку для одного абонента, а за имя родительсого объекта берем завод.
        abonents_list = [(obj_title,)]
        obj_title = u"Завод"
     
    for x in range(len(abonents_list)):
        # delta for groups abonents 'start date' A+
        cursor_t0_aplus_delta_start_temp = connection.cursor()
        cursor_t0_aplus_delta_start_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 A+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
        data_table_t0_aplus_delta_start_temp = cursor_t0_aplus_delta_start_temp.fetchall()
    
	# delta for groups abonents 'end date' A+
        cursor_t0_aplus_delta_end_temp = connection.cursor()
        cursor_t0_aplus_delta_end_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 A+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
        data_table_t0_aplus_delta_end_temp = cursor_t0_aplus_delta_end_temp.fetchall()
        
    # delta for groups abonents 'start date' R+
        cursor_t0_rplus_delta_start_temp = connection.cursor()
        cursor_t0_rplus_delta_start_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 R+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
        data_table_t0_rplus_delta_start_temp = cursor_t0_rplus_delta_start_temp.fetchall()
    
	# delta for groups abonents 'end date' R+
        cursor_t0_rplus_delta_end_temp = connection.cursor()
        cursor_t0_rplus_delta_end_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 R+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
        data_table_t0_rplus_delta_end_temp = cursor_t0_rplus_delta_end_temp.fetchall()
    
       
        data_table_temp = []
        data_table_temp.append(abonents_list[x][0]) # наименование канала
        try:# заводской номер
            data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][6])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:# T0 A+ нач
            data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")        
        try:# T0 A+ кон
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:# расход T0 A+
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1] - data_table_t0_aplus_delta_start_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try: #k_1
            data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][8])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try: #k_2
            data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][9])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:#k_3
            data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][10])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")

        try:# T0 R+ нач
            data_table_temp.append(data_table_t0_rplus_delta_start_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")        
        try:# T0 R+ кон
            data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:# расход T0 R+
            data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1] - data_table_t0_rplus_delta_start_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
                      
        data_table.append(data_table_temp)
        
    for row in range(6, len(abonents_list)+6):
        ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0]) # наименование канала
        ws.cell('A%s'%(row)).style = ali_grey
        
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
        ws.cell('B%s'%(row)).style = ali_white
        
        ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][2])  # T0 A+ нач
        ws.cell('F%s'%(row)).style = ali_white
               
        ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][3])  # T0 A+ кон
        ws.cell('G%s'%(row)).style = ali_white
        
        ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][4]) # расход T0 A+
        ws.cell('H%s'%(row)).style = ali_white
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][5]) # коэффициент 1  Ктт
        ws.cell('C%s'%(row)).style = ali_white
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][6]) # коэффициент 2  Ктн
        ws.cell('D%s'%(row)).style = ali_white
                
        ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][7]) # коэффициент 3  Постоянная счётчика
        ws.cell('E%s'%(row)).style = ali_white
        
        
        ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][8])  # T0 R+ нач
        ws.cell('J%s'%(row)).style = ali_white
               
        ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][9])  # T0 R+ кон
        ws.cell('K%s'%(row)).style = ali_white
        
        ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][10]) # расход T0 R+
        ws.cell('L%s'%(row)).style = ali_white
        
        try:
            ws.cell('I%s'%(row)).value = '%s' % (float(data_table[row-6][4])*float(data_table[row-6][5])*float(data_table[row-6][6])) # T0 R+ энергия с учёток коэффициентов
            ws.cell('I%s'%(row)).style = ali_yellow
        except UnicodeEncodeError:
            ws.cell('I%s'%(row)).value = '%s' % '-'
            ws.cell('I%s'%(row)).style = ali_yellow
        except TypeError:
            ws.cell('I%s'%(row)).value = '%s' % '-'
            ws.cell('I%s'%(row)).style = ali_yellow
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (float(data_table[row-6][10])*float(data_table[row-6][5])*float(data_table[row-6][6])) # T0 R+ энергия с учёток коэффициентов
            ws.cell('M%s'%(row)).style = ali_yellow
        except UnicodeEncodeError:
            ws.cell('M%s'%(row)).value = '%s' % '-'
            ws.cell('M%s'%(row)).style = ali_yellow
        except TypeError:
            ws.cell('M%s'%(row)).value = '%s' % '-'
            ws.cell('M%s'%(row)).style = ali_yellow
                    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
#    response['Content-Disposition'] = "attachment; filename=electric.xlsx"
    output_name = u'otchet za period ' + electric_data_start + '-' + electric_data_end+ translate(obj_title)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   

    return response
    
    
def pokazania(request): # Показания по общему тарифу по А+ и R+
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

# Шапка отчета   
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Срез показаний с коэффициентами на дату' + ' ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктт'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктн'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F4:I4')
    ws['F4'] = 'Суммарные показания/энергия'
    ws['F4'].style = ali_grey
    ws['G4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['F5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['G5'].style = ali_yellow
    
    ws['H5'] = 'Показания R+ на ' + str(request.session["electric_data_end"])
    ws['H5'].style = ali_grey
    
    ws['I5'] = 'Энергия R+ на ' + str(request.session["electric_data_end"])
    ws['I5'].style = ali_yellow
    
   
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
# Заполняем таблицу данными ----------------------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------------------------------------------------
    is_abonent_level    = re.compile(r'abonent')
    is_object_level     = re.compile(r'level')    
    obj_title           = request.session["obj_title"]
    electric_data_end   = request.session["electric_data_end"]
    obj_key             = request.session["obj_key"]
    data_table = []
    
#--------------------------------------------------------------------------------------------------------------------------------------------------------------
    if bool(is_object_level.search(obj_key)): # Если это объект, то формируем список абонентов
        cursor_abonents_list = connection.cursor()
        cursor_abonents_list.execute("""
                                  SELECT 
                                   abonents.name
                                  FROM 
                                   public.objects, 
                                   public.abonents
                                  WHERE 
                                   objects.guid = abonents.guid_objects AND
                                   objects.name = %s
                                  ORDER BY
                                   abonents.name ASC;""",[obj_title])
        abonents_list = cursor_abonents_list.fetchall()

    elif bool(is_abonent_level.search(obj_key)):   # Если это отдельный абонент, то делаем выборку для одного абонента, а за имя родительсого объекта берем завод.
        abonents_list = [(obj_title,)]
        obj_title = u"Завод"

    for x in range(len(abonents_list)):    
	# delta A+ for groups abonents 'end date'
        cursor_t0_aplus_delta_end_temp = connection.cursor()
        cursor_t0_aplus_delta_end_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 A+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
        data_table_t0_aplus_delta_end_temp = cursor_t0_aplus_delta_end_temp.fetchall()
        
        	# delta R+ for groups abonents 'end date'
        cursor_t0_rplus_delta_end_temp = connection.cursor()
        cursor_t0_rplus_delta_end_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 R+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
        data_table_t0_rplus_delta_end_temp = cursor_t0_rplus_delta_end_temp.fetchall()
           
        data_table_temp = []
        data_table_temp.append(abonents_list[x][0]) # наименование канала
        try:# заводской номер
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][6])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:# T0 A+ кон
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")        
        try: #k_1
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][8])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try: #k_2
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][9])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:#k_3
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][10])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:# T0 R+ кон
            data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
                       
        data_table.append(data_table_temp)
        
    for row in range(6, len(abonents_list)+6):
        ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0]) # наименование канала
        ws.cell('A%s'%(row)).style = ali_grey
        
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
        ws.cell('B%s'%(row)).style = ali_white
        
        ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][2])  # T0 A+ кон
        ws.cell('F%s'%(row)).style = ali_white
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3]) # коэффициент 1  Ктт
        ws.cell('C%s'%(row)).style = ali_white
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4]) # коэффициент 2  Ктн
        ws.cell('D%s'%(row)).style = ali_white
                
        ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5]) # коэффициент 3  Постоянная счётчика
        ws.cell('E%s'%(row)).style = ali_white

        ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][6]) # T0 R+ кон
        ws.cell('H%s'%(row)).style = ali_white
                
        try:
            ws.cell('G%s'%(row)).value = '%s' % (float(data_table[row-6][2])*float(data_table[row-6][3])*float(data_table[row-6][4])) # T0 A+ энергия с учёток коэффициентов
            ws.cell('G%s'%(row)).style = ali_yellow
        except UnicodeEncodeError:
            ws.cell('G%s'%(row)).value = '%s' % '-'
            ws.cell('G%s'%(row)).style = ali_yellow
        except TypeError:
            ws.cell('G%s'%(row)).value = '%s' % '-'
            ws.cell('G%s'%(row)).style = ali_yellow
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (float(data_table[row-6][6])*float(data_table[row-6][3])*float(data_table[row-6][4])) # T0 R+ энергия с учёток коэффициентов
            ws.cell('I%s'%(row)).style = ali_yellow
        except UnicodeEncodeError:
            ws.cell('I%s'%(row)).value = '%s' % '-'
            ws.cell('I%s'%(row)).style = ali_yellow
        except TypeError:
            ws.cell('I%s'%(row)).value = '%s' % '-'
            ws.cell('I%s'%(row)).style = ali_yellow

    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel") 
#    response['Content-Disposition'] = 'attachment; filename=eclectric.xlsx'
    output_name = u'otchet po pokazaniyam za ' + electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    
    return response


def profil_30_min(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    meters_name         = request.session["obj_title"]
    electric_data_end   = request.session["electric_data_end"]
    
    ws.merge_cells('A2:F2')
    ws['A2'] = u'Профиль мощности по абоненту ' + unicode(request.session['obj_title']) + u' за ' + str(request.session["electric_data_end"])
    ws['A3'] = u'Ктн = ' + str(common_sql.get_k_t_n(meters_name))
    ws['B3'] = u'Ктт = ' + str(common_sql.get_k_t_t(meters_name))
    ws['B5'] = 'Дата'
    ws['B5'].style = ali_grey
    ws['C5'] = 'Время'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Наименование'
    ws['D5'].style = ali_grey
    ws['E5'] = 'A+ кВт'
    ws['E5'].style = ali_grey
    ws['F5'] = 'R+ кВАр'
    ws['F5'].style = ali_grey
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['D'].width = 30

    
    a_plus = connection.cursor()
    a_plus.execute("""SELECT 
                          various_values.date, 
                          various_values."time", 
                          various_values.value, 
                          meters.name, 
                          meters.address, 
                          names_params.name
                        FROM 
                          public.various_values, 
                          public.meters, 
                          public.params, 
                          public.taken_params, 
                          public.names_params
                        WHERE 
                          params.guid_names_params = names_params.guid AND
                          taken_params.guid_params = params.guid AND
                          taken_params.guid_meters = meters.guid AND
                          taken_params.id = various_values.id_taken_params AND
                          various_values.date = %s AND 
                          meters.name = %s AND 
                          names_params.name = 'A+ Профиль';""",[electric_data_end, meters_name])
    a_plus = a_plus.fetchall()
   
        
    r_plus = connection.cursor()
    r_plus.execute("""SELECT 
                          various_values.date, 
                          various_values."time", 
                          various_values.value, 
                          meters.name, 
                          meters.address, 
                          names_params.name
                        FROM 
                          public.various_values, 
                          public.meters, 
                          public.params, 
                          public.taken_params, 
                          public.names_params
                        WHERE 
                          params.guid_names_params = names_params.guid AND
                          taken_params.guid_params = params.guid AND
                          taken_params.guid_meters = meters.guid AND
                          taken_params.id = various_values.id_taken_params AND
                          various_values.date = %s AND 
                          meters.name = %s AND 
                          names_params.name = 'R+ Профиль';""",[electric_data_end, meters_name])
    r_plus = r_plus.fetchall()
          
    data_table = []
    for x in range(len(a_plus)):
        data_table_temp = []
        try:
            data_table_temp.append(a_plus[x][0])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(a_plus[x][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(a_plus[x][3])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(a_plus[x][2]*2*common_sql.get_k_t_n(meters_name)*common_sql.get_k_t_t(meters_name))
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(r_plus[x][2]*2*common_sql.get_k_t_n(meters_name)*common_sql.get_k_t_t(meters_name))
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        data_table.append(data_table_temp)
            
    for row in range(6, len(data_table)+6):
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][0]) # дата
        ws.cell('B%s'%(row)).style = ali_grey
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][1])  # время
        ws.cell('C%s'%(row)).style = ali_white
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # канал
        ws.cell('D%s'%(row)).style = ali_white
        
        ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3]) # значение A+
        ws.cell('E%s'%(row)).style = ali_white
        
        ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][4]) # значение R+
        ws.cell('F%s'%(row)).style = ali_white
           
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'profil 30 min za ' + electric_data_end # формируем имя для excel отчета 
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_hour_increment(request): # Выгрузка excel по часовым приращениям
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    meters_name         = request.session["obj_title"]
    electric_data_end   = request.session["electric_data_end"]    

# Шапка очета      
    ws.merge_cells('B2:F2')
    ws['B2'] = 'Почасовой учет электроэнергии за ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('B3:F3')
    ws['B3'] = u'Абонент: ' + unicode(request.session['obj_title'])
    
    ws['B5'] = 'Дата'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Время'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Абонент'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Серийный номер'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'A+ кВт*ч'
    ws['F5'].style = ali_grey
    
    ws['G5'] = '+ А+ кВт*ч'
    ws['G5'].style = ali_grey
    
    ws['H5'] = 'R+ кВар*ч'
    ws['H5'].style = ali_grey
    
    ws['I5'] = '+ R+ кВар*ч'
    ws['I5'].style = ali_grey
    
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20  
#-------------
    
#Запрашиваем данные для отчета
    time_list = ['00:00', '00:30','01:00', '01:30', '02:00', '02:30', '03:00', '03:30', '04:00', '04:30', '05:00', '05:30', '06:00', '06:30', '07:00', '07:30', '08:00', '08:30', '09:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00', '12:30', '13:00', '13:30', '14:00', '14:30', '15:00', '15:30', '16:00', '16:30', '17:00', '17:30', '18:00', '18:30', '19:00', '19:30', '20:00', '20:30', '21:00', '21:30', '22:00', '22:30', '23:00', '23:30']
    meters_name         = request.session["obj_title"]
    electric_data_end   = request.session["electric_data_end"]
    
    serial_number = common_sql.get_serial_number_by_meter_name(meters_name)
        
    data_table = []
    # Добавляем первую строку в таблицу данных. Делаем запрос показаний на начало суток.
    data_table.append([electric_data_end,u'00:00', meters_name, serial_number, common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ),common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 R+' ), u'0', u'0'])
    
    if common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ) != u'Нет данных':  # Если есть показания на начало суток выполняем почасовое приращение  
        for x in range(24):
            data_table_temp = []
            data_table_temp.append(electric_data_end)
            data_table_temp.append(time_list[(2*x)])
            data_table_temp.append(meters_name)
            data_table_temp.append(serial_number)
            data_table_temp.append(data_table[len(data_table)-1][4] + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))
            data_table_temp.append(common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 R+' ) + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))
            data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))
            data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))    
            if x == 0: # Убираем первую строку. Так как показания на 00:00 берем отдельным запросом
                next
            else:
                data_table.append(data_table_temp)    
#-----------------------------

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][0]) # дата
        ws.cell('B%s'%(row)).style = ali_grey
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][1])  # время
        ws.cell('C%s'%(row)).style = ali_white
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # канал
        ws.cell('D%s'%(row)).style = ali_white
        
        ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3]) # Заводской номер
        ws.cell('E%s'%(row)).style = ali_white
        
        ws.cell('F%s'%(row)).value = '%s' % (round(data_table[row-6][4],2)) # значение A+
        ws.cell('F%s'%(row)).style = ali_white
        
        ws.cell('G%s'%(row)).value = '%s' % (round(float(data_table[row-6][6]),2)) # значение + A+
        ws.cell('G%s'%(row)).style = ali_white
        
        ws.cell('H%s'%(row)).value = '%s' % (round(data_table[row-6][5],2)) # значение R+
        ws.cell('H%s'%(row)).style = ali_white

        ws.cell('I%s'%(row)).value = '%s' % (round(float(data_table[row-6][7]),2)) # значение + R+
        ws.cell('I%s'%(row)).style = ali_white
#---------------------------
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'chasovie prirasheniya'
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
    
def pokazania_period(request): # Показания по абоненту за период
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

    meters_name         = request.session["obj_title"]
    parent_name         = request.session['obj_parent_title']
    electric_data_end   = request.session["electric_data_end"]
    electric_data_start = request.session['electric_data_start']
    data_table = []
# Шапка отчета
    ws.merge_cells('B2:F2')
    ws['B2'] = u'Ежедневные показания за период с ' + str(request.session["electric_data_start"]) + u' по ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('B3:F3')
    ws['B3'] = u'Абонент: ' + unicode(request.session['obj_title'])
   
    ws['B5'] = 'Дата'
    ws['B5'].style = ali_grey
       
    ws['C5'] = 'Абонент'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Серийный номер'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'A+ кВт*ч'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'R+ кВар*ч'
    ws['F5'].style = ali_grey
       
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20          
# Конец шапки

#Запрашиваем данные для отчета
    end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
    start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
    dates = [x for x in common_sql.daterange(start_date,
                  end_date,
                  step=datetime.timedelta(days=1),
                  inclusive=True)]

    for x in range(len(dates)):
        data_table_temp = get_data_table_by_date_daily(meters_name, parent_name, datetime.datetime.strftime(dates[x], "%d.%m.%Y"))
        if data_table_temp:
            data_table.extend(data_table_temp)
        else:
            data_table.append([datetime.datetime.strftime(dates[x], "%d.%m.%Y"),meters_name,common_sql.get_serial_number_by_meter_name(meters_name), u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д'])
   
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][0]) # дата
        ws.cell('B%s'%(row)).style = ali_grey
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][1])  # абонент
        ws.cell('C%s'%(row)).style = ali_white
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
        ws.cell('D%s'%(row)).style = ali_white
        
        ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3]) # значение A+
        ws.cell('E%s'%(row)).style = ali_white
        
        ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][8]) # значение R+
        ws.cell('F%s'%(row)).style = ali_white
        
#---------------------------
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'pokazania za period' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_rejim_day(request): # Отчет по режимному дню
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    meters_name         = request.session["obj_title"]
#    parent_name         = request.session['obj_parent_title']
    electric_data_end   = request.session["electric_data_end"]
#    electric_data_start = request.session['electric_data_start']
    data_table = []
    general_k = common_sql.get_k_t_t(meters_name) * common_sql.get_k_t_n(meters_name)
# Шапка отчета
    ws['A1'] = u"ЗАО 'Кировская керамика'"
    ws['H1'] = u'Шифр'
    ws['H2'] = u'Питающий центр'
    ws['H3'] = u'№ фидера'
    
    ws['D5'] = u'Протокол (первичный)'
    ws['D5'].style = ali_white_size_18
    
    ws['E7'] = u'трансформаторного напряжения _____ вольт'
    
    ws['B6'] = u'записей показаний электросчетчиков и вольтметров, а также определения нагрузок'
    ws['B7'] = u"и тангенса 'фи' за " + str(electric_data_end) + u'г'
    ws['B9'] = u'Акт. Счетчик № '# + str(common_sql.get_serial_number_by_meter_name(meters_name)) 
    ws['E9'] = u'Реакт. Счетчик № '# + str(common_sql.get_serial_number_by_meter_name(meters_name))
    ws['B10'] = u'Расч. Коэфициент' 
    ws['E10'] = u'Расч. Коэфициент'

    ws['A11'] = u'Время' 
    ws['A12'] = u'записи, часы'
    ws['A39'] = u'суточный расход' 
    ws['A40'] = u'активной и реактивной'
    ws['A41'] = u'энергии' 
    ws['A42'] = u'Контрольная сумма'
    ws['A44'] = u'Запись показаний счетчиков производили'
    ws['A44'] = u'Запись показаний счетчиков производили'
    ws['A45'] = u'фамилия ______________ подпись ______________'
    ws['A46'] = u'фамилия ______________ подпись ______________'
    ws['A47'] = u'фамилия ______________ подпись ______________'

    ws['G44'] = u'Расчеты произвел:'
    ws['G46'] = u'фамилия ______________ подпись ______________'

    ws['D10'] = ws['G10'] = general_k # Общий коэффициент
    ws['D9'] = ws['G9'] = str(common_sql.get_serial_number_by_meter_name(meters_name)) #Серийный номер прибора
        
    ws['A13'] = u'0' 
    ws['A14'] = u'1' 
    ws['A15'] = u'2'
    ws['A16'] = u'3' 
    ws['A17'] = u'4'
    ws['A18'] = u'5' 
    ws['A19'] = u'6'
    ws['A20'] = u'7' 
    ws['A21'] = u'8'
    ws['A22'] = u'9' 
    ws['A23'] = u'10'
    ws['A24'] = u'11' 
    ws['A25'] = u'12'
    ws['A26'] = u'13' 
    ws['A27'] = u'14'
    ws['A28'] = u'15' 
    ws['A29'] = u'16'
    ws['A30'] = u'17' 
    ws['A31'] = u'18'
    ws['A32'] = u'19' 
    ws['A33'] = u'20'
    ws['A34'] = u'21' 
    ws['A35'] = u'22'
    ws['A36'] = u'23' 
    ws['A37'] = u'24'

    ws['B11'] = u'Показания' 
    ws['B12'] = u'счетчика'
    
    ws['C11'] = u'Разность' 
    ws['C12'] = u'показаний'
    
    ws['D11'] = u'расход за' 
    ws['D12'] = u'час(квт)'
    
    ws['E11'] = u'Показания' 
    ws['E12'] = u'счетчика'
    
    ws['F11'] = u'Разность' 
    ws['F12'] = u'показаний'
    
    ws['G11'] = u'расход за' 
    ws['G12'] = u'час(квт)'

    ws['H10'] = u'тангенс'
    ws['H11'] = u'фи'
    
    ws['I9'] = u'Показания'
    ws['I10'] = u'вольтметров на'
    ws['I11'] = u'стороне'
    ws['I12'] = u'в/н'
    ws['J12'] = u'н/н'
    
    ws['K9'] = u'мощность '
    ws['K10'] = u'включенных'
    ws['K11'] = u'компен.устр'
    ws['K12'] = u'кВар'
    
    for col_idx in range(1, 12):
        col = get_column_letter(col_idx)
        for row in range(13, 38):
            ws.cell('%s%s'%(col, row)).style = ali_white

    ws.column_dimensions['A'].width = 12            
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['I'].width = 7
    ws.column_dimensions['J'].width = 7    
    ws.row_dimensions[5].height = 30
    # Конец шапки

#Запрашиваем данные для отчета
    time_list = ['00:00', '00:30','01:00', '01:30', '02:00', '02:30', '03:00', '03:30', '04:00', '04:30', '05:00', '05:30', '06:00', '06:30', '07:00', '07:30', '08:00', '08:30', '09:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00', '12:30', '13:00', '13:30', '14:00', '14:30', '15:00', '15:30', '16:00', '16:30', '17:00', '17:30', '18:00', '18:30', '19:00', '19:30', '20:00', '20:30', '21:00', '21:30', '22:00', '22:30', '23:00', '23:30']
    
    serial_number = common_sql.get_serial_number_by_meter_name(meters_name)
    # Добавляем первую строку в таблицу данных. Делаем запрос показаний на начало суток.
    data_table.append([electric_data_end,u'00:00', meters_name, serial_number, common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ),common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 R+' ), u'0', u'0'])
    
    if common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ) != u'Нет данных':  # Если есть показания на начало суток выполняем почасовое приращение  
        for x in range(24):
            data_table_temp = []
            data_table_temp.append(electric_data_end) # Дата
            data_table_temp.append(time_list[(2*x)])  # Отчетный час
            data_table_temp.append(meters_name)       # Имя абонента
            data_table_temp.append(serial_number)     # Серийный номер
            data_table_temp.append(data_table[len(data_table)-1][4] + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))     # Показиние счётчика за предыдущий час + две получасовки А+           
            data_table_temp.append(data_table[len(data_table)-1][5] + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))     # Показиние счётчика за предыдущий час + две получасовки R+

            data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))                                        # Сумма двух получасовок: потребленная энергия за час А+
            data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))                                        # Сумма двух получасовок: потребленная энергия за час R+  
            if x == 0: # Убираем первую строку. Так как показания на 00:00 берем отдельным запросом 
                next
            else:
                data_table.append(data_table_temp)
    if data_table[23][4] and data_table[23][5]:
        data_table.append([(datetime.datetime.strptime(electric_data_end, u'%d.%m.%Y') + datetime.timedelta(days=1)).strftime(u'%d.%m.%Y'),u'00:00', meters_name, serial_number, common_sql.get_daily_value_by_meter_name(meters_name, (datetime.datetime.strptime(electric_data_end, u'%d.%m.%Y') + datetime.timedelta(days=1)).strftime(u'%d.%m.%Y'), 'T0 A+' ),common_sql.get_daily_value_by_meter_name(meters_name, (datetime.datetime.strptime(electric_data_end, u'%d.%m.%Y') + datetime.timedelta(days=1)).strftime(u'%d.%m.%Y'), 'T0 R+' ),common_sql.get_daily_value_by_meter_name(meters_name, (datetime.datetime.strptime(electric_data_end, u'%d.%m.%Y') + datetime.timedelta(days=1)).strftime(u'%d.%m.%Y'), 'T0 A+' ) - data_table[23][4],common_sql.get_daily_value_by_meter_name(meters_name, (datetime.datetime.strptime(electric_data_end, u'%d.%m.%Y') + datetime.timedelta(days=1)).strftime(u'%d.%m.%Y'), 'T0 R+' ) - data_table[23][5]])

    #------------

# Заполняем отчет значениями
    for row in range(13, len(data_table)+13):
        
        ws.cell('B%s'%(row)).value = '%s' % (round(data_table[row-13][4],4)) # значение A+
        ws.cell('B%s'%(row)).style = ali_white
        
        ws.cell('C%s'%(row)).value = '%s' % (round(float(data_table[row-13][6]),4)) # значение + A+
        ws.cell('C%s'%(row)).style = ali_white
        
        ws.cell('D%s'%(row)).value = '%s' % (round(float(data_table[row-13][6]),4)*general_k) # значение + A+
        ws.cell('D%s'%(row)).style = ali_white
        
        ws.cell('E%s'%(row)).value = '%s' % (round(data_table[row-13][5],4)) # значение R+
        ws.cell('E%s'%(row)).style = ali_white

        ws.cell('F%s'%(row)).value = '%s' % (round(float(data_table[row-13][7]),4)) # значение + R+
        ws.cell('F%s'%(row)).style = ali_white
        
        ws.cell('G%s'%(row)).value = '%s' % (round(float(data_table[row-13][7]),4)*general_k) # значение + R+
        ws.cell('G%s'%(row)).style = ali_white
        
        fi=0
        try:
            fi = round((float(data_table[row-13][7])*general_k)/(float(data_table[row-13][6])*general_k),2)
        except:
            fi = u''
        
        ws.cell('H%s'%(row)).value = '%s' % fi # значение + R+/+ A+
        ws.cell('H%s'%(row)).style = ali_white  
    #---------------------------      
    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'rejimniy den '+ str(electric_data_end) 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    

def report_economic_electric(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    #--------------------------------------------------------------------------------------------------------------------      
# Шапка отчета
    ws.merge_cells('B2:F2')
    ws['B2'] = u'Таблица расчета удельного коэффициента период с ' + str(request.session["electric_data_start"]) + u' по ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('B3:F3')
    ws['B3'] = u'Абонент: Литейный цех'
   
    ws['B5'] = 'Дата'
    ws['B5'].style = ali_grey
       
    ws['C5'] = 'Изготовленная продукция, кг'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Затраченная A+, кВт*ч'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Удельный расход A+, кВт*ч/кг'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Затраченная R+, кВар*ч'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Удельный расход R+, кВт*ч/кг'
    ws['G5'].style = ali_grey

    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18    
    ws.row_dimensions[5].height = 30    
# Конец шапки
    
#Запрашиваем данные для отчета---
    data_table = []
    
    electric_data_end   = request.session["electric_data_end"]
    electric_data_start   = request.session["electric_data_start"]

    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']    

    end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
    start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
    dates = [x for x in common_sql.daterange(start_date,
                  end_date,
                  step=datetime.timedelta(days=1),
                  inclusive=True)]
                  
    for x in range(len(dates)):
        try:
            data_table_temp = []
            delta_a_plus = 1
            delta_r_plus = 1

            try:
                delta_a_plus = common_sql.delta_sum_a_plus(dates[x+1])-common_sql.delta_sum_a_plus(dates[x])
                if delta_a_plus > 0:
				    delta_a_plus = delta_a_plus
                else:
				    delta_a_plus = u'Н/Д'
                delta_r_plus = common_sql.delta_sum_r_plus(dates[x+1])-common_sql.delta_sum_r_plus(dates[x])
                if delta_r_plus > 0:
				    delta_r_plus = delta_r_plus
                else:
                    delta_r_plus = u'Н/Д'

            except:
                delta_a_plus = u'Н/Д'
                delta_r_plus = u'Н/Д'

            data_table_temp.append(dates[x])
            data_table_temp.append(common_sql.product_sum(dates[x]))
            data_table_temp.append(delta_a_plus)
            data_table_temp.append(delta_a_plus/(common_sql.product_sum(dates[x])))
            data_table_temp.append(delta_r_plus)
            data_table_temp.append(delta_r_plus/(common_sql.product_sum(dates[x])))
        except:
            next
        data_table.append(data_table_temp)

#Конец запроса данных------------
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        ws.cell('B%s'%(row)).value = '%s' % ((data_table[row-6][0]).strftime("%d-%m-%Y")) # дата
        ws.cell('B%s'%(row)).style = ali_grey
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][1])  # изготовленная продукция, кг
        ws.cell('C%s'%(row)).style = ali_white
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # затраченная А+
        ws.cell('D%s'%(row)).style = ali_white
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3])  # удельных расход А+/кг
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][4])  # затраченная R+
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][5])  # удельный расход R+/кг
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
        
        
        
#---------------------------    
    #--------------------------------------------------------------------------------------------------------------------    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'udelniy_coefficient_liteiniy_ceh' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response


def report_pokazaniya_water_identificators(request): # Показания по воде за дату с идентификаторами
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

    meters_name         = request.session["obj_title"]
    parent_name         = request.session['obj_parent_title']
    electric_data_end   = request.session["electric_data_end"]
#    electric_data_start = request.session['electric_data_start']
    data_table = []

# Шапка отчета
    ws.merge_cells('B2:F2')
    ws['B2'] = u'Показания по воде с идентификаторами за ' + str(request.session["electric_data_end"])
       
    ws['B5'] = 'Абонент'
    ws['B5'].style = ali_grey
       
    ws['C5'] = 'Идентификатор'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания'
    ws['D5'].style = ali_grey
    
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20        
# Конец шапки

#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_channel(meters_name, electric_data_end)
    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []        
        for x in range(len(list_of_abonents_2)):
            data_table_temp = common_sql.get_daily_water_channel(list_of_abonents_2[x], electric_data_end)
            data_table.extend(data_table_temp)
    elif(bool(is_object_level_1.search(obj_key))):
        
        list_of_objects_2 = common_sql.list_of_objects(common_sql.return_parent_guid_by_abonent_name(meters_name)) #Список квартир для объекта с пульсарами
        data_table = []
        for x in range(len(list_of_objects_2)):
            data_table_temp = [(list_of_objects_2[x][0],)]
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(meters_name), list_of_objects_2[x][0])
            for y in range(len(list_of_abonents_2)):
                data_table_temp2 = common_sql.get_daily_water_channel(list_of_abonents_2[y], electric_data_end)

                data_table_temp.extend(data_table_temp2)                                
                      
            data_table.extend(data_table_temp)
              
    else:
        data_table = [1,1,1,1]

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][0]) # абонент
        ws.cell('B%s'%(row)).style = ali_grey
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][4])  # идентификатор
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # показания м3
            ws.cell('D%s'%(row)).style = ali_white

        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        
#---------------------------

    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'pokazania po vode' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_electric_simple_2_zones(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Срез показаний с коэффициентами на дату' + ' ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = ali_grey
    ws['G4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['F5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['G5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('H4:I4')
    ws['H4'] = 'Тариф 1'
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['H5'].style = ali_grey
    
    ws['I5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['I5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('J4:K4')
    ws['J4'] = 'Тариф 2'
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['J5'].style = ali_grey
    
    ws['K5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['K5'].style = ali_yellow
    
       
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    data_table = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
                        
#Запрашиваем данные для отчета конец
                  
    if (bool(is_abonent_level.search(obj_key))):
        if (is_electric_monthly == "1"):
            data_table = common_sql.get_data_table_by_date_monthly_2_zones(meters_name, parent_name, electric_data_end)
        elif (is_electric_daily == "1"):
            data_table = common_sql.get_data_table_by_date_daily_2_zones(meters_name, parent_name, electric_data_end)

    elif (bool(is_object_level_2.search(obj_key))):
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
            data_table = []
            for x in range(len(list_of_abonents_2)):
                if (is_electric_monthly == "1"):                
                    data_table_temp = common_sql.get_data_table_by_date_monthly_2_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                elif (is_electric_daily == "1"):                
                    data_table_temp = common_sql.get_data_table_by_date_daily_2_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                if data_table_temp:
                    data_table.extend(data_table_temp)
                else:
                    data_table.extend([[u'', list_of_abonents_2[x],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (get_k_t_n_by_serial_number(data_table[row-6][2]))  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (get_k_t_t_by_serial_number(data_table[row-6][2]))  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (get_k_a_by_serial_number(data_table[row-6][2]))  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][3])  # Сумма А+
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][3]*get_k_t_n_by_serial_number(data_table[row-6][2])*get_k_t_t_by_serial_number(data_table[row-6][2]))  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тариф 1 А+
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][4]*get_k_t_n_by_serial_number(data_table[row-6][2])*get_k_t_t_by_serial_number(data_table[row-6][2]))  # "Энергия Тариф 1 А+
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][5])  # Тариф 2 А+
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][5]*get_k_t_n_by_serial_number(data_table[row-6][2])*get_k_t_t_by_serial_number(data_table[row-6][2]))  # "Энергия Тариф 2 А+
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next           
# Сохраняем в ecxel    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'2_tariffa' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response
    

def report_electric_simple_2_zones_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Срез показаний с коэффициентами на дату' + ' ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = ali_grey
    ws['G4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['F5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['G5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('H4:I4')
    ws['H4'] = 'Тариф 1'
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['H5'].style = ali_grey
    
    ws['I5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['I5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('J4:K4')
    ws['J4'] = 'Тариф 2'
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['J5'].style = ali_grey
    
    ws['K5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['K5'].style = ali_yellow
    
       
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    is_electric_period  = request.session['is_electric_period']
    data_table = []
    if True:
        if True:            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'monthly')
                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'daily')

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # показания на начало месяца для объекта
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        

#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for abonents group
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
            elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
              
            elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][9])  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][8])  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][10])  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][3])  # Сумма А+
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % round((data_table[row-6][3]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тариф 1 А+
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Тариф 1 А+
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][5])  # Тариф 2 А+
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % round((data_table[row-6][5]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Тариф 2 А+
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next           
# Сохраняем в ecxel    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'2_tariffa' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response
    
def report_electric_simple_3_zones(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Срез показаний с коэффициентами на дату' + ' ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = ali_grey
    ws['G4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['F5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['G5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('H4:I4')
    ws['H4'] = 'Тариф 1'
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['H5'].style = ali_grey
    
    ws['I5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['I5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('J4:K4')
    ws['J4'] = 'Тариф 2'
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['J5'].style = ali_grey
    
    ws['K5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['K5'].style = ali_yellow
    
    # Тариф 3
    ws.merge_cells('L4:M4')
    ws['L4'] = 'Тариф 3'
    ws['L4'].style = ali_grey
    ws['M4'].style = ali_grey
    ws['L4'].style = ali_grey
    ws['M4'].style = ali_grey
    ws['L5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['L5'].style = ali_grey
    
    ws['M5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['M5'].style = ali_yellow
         
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    data_table = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
                     
#Запрашиваем данные для отчета конец                  
    if (bool(is_abonent_level.search(obj_key))):
        if (is_electric_monthly == "1"):
            data_table = common_sql.get_data_table_by_date_monthly_3_zones(meters_name, parent_name, electric_data_end)
        elif (is_electric_daily == "1"):
            data_table = common_sql.get_data_table_by_date_daily_3_zones(meters_name, parent_name, electric_data_end)

    elif (bool(is_object_level_2.search(obj_key))):
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
            data_table = []
            for x in range(len(list_of_abonents_2)):
                if (is_electric_monthly == "1"):                
                    data_table_temp = common_sql.get_data_table_by_date_monthly_3_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                elif (is_electric_daily == "1"):                
                    data_table_temp = common_sql.get_data_table_by_date_daily_3_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                if data_table_temp:
                    data_table.extend(data_table_temp)
                else:
                    data_table.extend([[u'', list_of_abonents_2[x],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (get_k_t_n_by_serial_number(data_table[row-6][2]))  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (get_k_t_t_by_serial_number(data_table[row-6][2]))  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (get_k_a_by_serial_number(data_table[row-6][2]))  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][3])  # Сумма А+
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][3]*get_k_t_n_by_serial_number(data_table[row-6][2])*get_k_t_t_by_serial_number(data_table[row-6][2]))  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тариф 1 А+
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][4]*get_k_t_n_by_serial_number(data_table[row-6][2])*get_k_t_t_by_serial_number(data_table[row-6][2]))  # "Энергия Тариф 1 А+
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][5])  # Тариф 2 А+
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][5]*get_k_t_n_by_serial_number(data_table[row-6][2])*get_k_t_t_by_serial_number(data_table[row-6][2])) # "Энергия Тариф 2 А+
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][6])  # Тариф 3 А+
            ws.cell('L%s'%(row)).style = ali_white

        except:
            ws.cell('L%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (data_table[row-6][6]*get_k_t_n_by_serial_number(data_table[row-6][2])*get_k_t_t_by_serial_number(data_table[row-6][2]))  # "Энергия Тариф 3 А+
            ws.cell('M%s'%(row)).style = ali_yellow
        except:
            ws.cell('M%s'%(row)).style = ali_yellow
            next

# Сохраняем в ecxel  
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'3_tariffa' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def report_electric_simple_3_zones_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']    
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+' .Срез показаний с коэффициентами на дату' + ' ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = ali_grey
    ws['G4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['F5'] = 'Показания A+ на ' + electric_data_end
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Энергия A+ на ' + electric_data_end
    ws['G5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('H4:I4')
    ws['H4'] = 'Тариф 1'
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H5'] = 'Показания A+ на ' +  electric_data_end
    ws['H5'].style = ali_grey
    
    ws['I5'] = 'Энергия A+ на ' +  electric_data_end
    ws['I5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('J4:K4')
    ws['J4'] = 'Тариф 2'
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J5'] = 'Показания A+ на ' +  electric_data_end
    ws['J5'].style = ali_grey
    
    ws['K5'] = 'Энергия A+ на ' +  electric_data_end
    ws['K5'].style = ali_yellow
    
    # Тариф 3
    ws.merge_cells('L4:M4')
    ws['L4'] = 'Тариф 3'
    ws['L4'].style = ali_grey
    ws['M4'].style = ali_grey
    ws['L4'].style = ali_grey
    ws['M4'].style = ali_grey
    ws['L5'] = 'Показания A+ на ' +  electric_data_end
    ws['L5'].style = ali_grey
    
    ws['M5'] = 'Энергия A+ на ' +  electric_data_end
    ws['M5'].style = ali_yellow
         
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
    #выборка данных из БД
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
            
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    is_electric_period  = request.session['is_electric_period']
    data_table = []
    if True:
        if True:            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'monthly')
                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'daily')

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # показания на начало месяца для объекта
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        

#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for abonents group
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
            elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
              
            elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][9])  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][8])  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][10])  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][3])  # Сумма А+
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % round((data_table[row-6][3]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тариф 1 А+
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Тариф 1 А+
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][5])  # Тариф 2 А+
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % round((data_table[row-6][5]*data_table[row-6][8]*data_table[row-6][9]),3) # "Энергия Тариф 2 А+
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][6])  # Тариф 3 А+
            ws.cell('L%s'%(row)).style = ali_white

        except:
            ws.cell('L%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % round((data_table[row-6][6]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Тариф 3 А+
            ws.cell('M%s'%(row)).style = ali_yellow
        except:
            ws.cell('M%s'%(row)).style = ali_yellow
            next

# Сохраняем в ecxel  
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'3_tariffa_'+translate(obj_title)+'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def electric_between_report(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Значения профиля показаний за период с' + ' '+str(request.session["electric_data_start"]) +' по '+ str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:C5')
    ws['C4'] = 'Дата'
    ws['C4'].style = ali_grey
    ws['C5'].style = ali_grey
    
    # Сумма
    ws.merge_cells('D4:D5')
    ws['D4'] = 'Сумма - Показания A+ '
    ws['D4'].style = ali_grey
    ws['D5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('E4:E5')
    ws['E4'] = 'Сумма - Расход за прошедшие сутки'
    ws['E4'].style = ali_grey
    ws['E5'].style = ali_grey
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_electric_daily    = request.session['is_electric_daily']
    obj_parent_title    = request.session['obj_parent_title']
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
    obj_key             = request.session['obj_key']

    data_table = []
    if True:
        if True:            
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end)
            else:
                pass
            
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][3])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][4])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # дата
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][5])  # сумма-показания
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][12])  # Расход за прошедшие сутки
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next

# Сохраняем в ecxel    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_electric' + str(electric_data_start) + u' - ' + str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response
    
def electric_between_2_zones_report(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Значения профиля показаний за период с' + ' '+str(request.session["electric_data_start"]) +' по '+ str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:C5')
    ws['C4'] = 'Дата'
    ws['C4'].style = ali_grey
    ws['C5'].style = ali_grey
    
    # Сумма
    ws.merge_cells('D4:D5')
    ws['D4'] = 'Сумма - Показания T0 A+ '
    ws['D4'].style = ali_grey
    ws['D5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('E4:E5')
    ws['E4'] = 'Сумма - Расход за прошедшие сутки T0'
    ws['E4'].style = ali_grey
    ws['E5'].style = ali_grey
    
        # Сумма
    ws.merge_cells('F4:F5')
    ws['F4'] = 'Сумма - Показания T1 A+ '
    ws['F4'].style = ali_grey
    ws['F5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('G4:G5')
    ws['G4'] = 'Сумма - Расход за прошедшие сутки T1'
    ws['G4'].style = ali_grey
    ws['G5'].style = ali_grey
    
        # Сумма
    ws.merge_cells('H4:H5')
    ws['H4'] = 'Сумма - Показания T2 A+ '
    ws['H4'].style = ali_grey
    ws['H5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('I4:I5')
    ws['I4'] = 'Сумма - Расход за прошедшие сутки T2'
    ws['I4'].style = ali_grey
    ws['I5'].style = ali_grey
    
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_electric_daily    = request.session['is_electric_daily']
    obj_parent_title    = request.session['obj_parent_title']
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
    obj_key             = request.session['obj_key']

    data_table = []
    if True:
        if True:            
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end)
            else:
                pass
            
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][3])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][4])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # дата
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][5])  # сумма-показания t0
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][12])  # Расход за прошедшие сутки t0
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # сумма-показанияt1
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][13])  # Расход за прошедшие суткиt1
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # сумма-показанияt2
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][14])  # Расход за прошедшие суткиt2
            ws.cell('I%s'%(row)).style = ali_white
        except:
            ws.cell('I%s'%(row)).style = ali_white
            next

# Сохраняем в ecxel    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_electric_2_zones_' + str(electric_data_start) + u' - ' + str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response
    
def electric_between_3_zones_report(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+'. Значения профиля показаний за период с' + ' '+electric_data_start +' по '+ electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:C5')
    ws['C4'] = 'Дата'
    ws['C4'].style = ali_grey
    ws['C5'].style = ali_grey
    
    # Сумма
    ws.merge_cells('D4:D5')
    ws['D4'] = 'Показания T0 A+ '
    ws['D4'].style = ali_grey
    ws['D5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('E4:E5')
    ws['E4'] = 'Расход за прошедшие сутки T0'
    ws['E4'].style = ali_grey
    ws['E5'].style = ali_grey
    
        # Сумма
    ws.merge_cells('F4:F5')
    ws['F4'] = 'Показания T1 A+ '
    ws['F4'].style = ali_grey
    ws['F5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('G4:G5')
    ws['G4'] = 'Расход за прошедшие сутки T1'
    ws['G4'].style = ali_grey
    ws['G5'].style = ali_grey
    
        # Сумма
    ws.merge_cells('H4:H5')
    ws['H4'] = 'Показания T2 A+ '
    ws['H4'].style = ali_grey
    ws['H5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('I4:I5')
    ws['I4'] = 'Расход за прошедшие сутки T2'
    ws['I4'].style = ali_grey
    ws['I5'].style = ali_grey
    
        # Сумма
    ws.merge_cells('J4:J5')
    ws['J4'] = 'Показания T3 A+ '
    ws['J4'].style = ali_grey
    ws['J5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('K4:K5')
    ws['K4'] = 'Расход за прошедшие сутки T3'
    ws['K4'].style = ali_grey
    ws['K5'].style = ali_grey
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_electric_daily    = request.session['is_electric_daily']
    obj_parent_title    = request.session['obj_parent_title']
    
    obj_key             = request.session['obj_key']

    data_table = []
    if True:
        if True:            
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end)
            else:
                pass
            
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][3])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][4])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # дата
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][5])  # сумма-показания t0
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][12])  # Расход за прошедшие сутки t0
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # сумма-показанияt1
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][13])  # Расход за прошедшие суткиt1
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # сумма-показанияt2
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][14])  # Расход за прошедшие суткиt2
            ws.cell('I%s'%(row)).style = ali_white
        except:
            ws.cell('I%s'%(row)).style = ali_white
            next
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][8])  # сумма-показанияt3
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][15])  # Расход за прошедшие суткиt3
            ws.cell('K%s'%(row)).style = ali_white
        except:
            ws.cell('K%s'%(row)).style = ali_white
            next

# Сохраняем в ecxel    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_electric_3_zones_'+translate(obj_title)+u'_' + electric_data_start + u' - ' +electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def report_electric_potreblenie_2_zones(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление электроэнергии в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F3:I3')
    ws['F3'] = 'Сумма A+, кВт*ч'
    ws['F3'].style = ali_grey
    ws['G3'].style = ali_grey
    ws['H3'].style = ali_grey
    ws['I3'].style = ali_grey
    
    ws.merge_cells('F4:G4')
    ws['F4'] = 'На ' + str(request.session["electric_data_start"])
    ws['F4'].style = ali_grey
    
    ws.merge_cells('H4:I4')
    ws['H4'] = 'На ' + str(request.session["electric_data_end"])
    ws['H4'].style = ali_grey

    ws['F5'] = 'Показания'
    ws['F5'].style = ali_grey     
    ws['G5'] = 'Энергия'
    ws['G5'].style = ali_yellow
    
    ws['H5'] = 'Показания'
    ws['H5'].style = ali_grey     
    ws['I5'] = 'Энергия'
    ws['I5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('J3:M3')
    ws['J3'] = 'Тариф 1 A+, кВт*ч'
    ws['J3'].style = ali_grey
    ws['K3'].style = ali_grey
    ws['L3'].style = ali_grey
    ws['M3'].style = ali_grey
    
    ws.merge_cells('J4:K4')
    ws['J4'] = 'На ' + str(request.session["electric_data_start"])
    ws['J4'].style = ali_grey
    
    ws.merge_cells('L4:M4')
    ws['L4'] = 'На ' + str(request.session["electric_data_end"])
    ws['L4'].style = ali_grey

    ws['J5'] = 'Показания'
    ws['J5'].style = ali_grey     
    ws['K5'] = 'Энергия'
    ws['K5'].style = ali_yellow
    
    ws['L5'] = 'Показания'
    ws['L5'].style = ali_grey     
    ws['M5'] = 'Энергия'
    ws['M5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('N3:Q3')
    ws['N3'] = 'Тариф 2 A+, кВт*ч'
    ws['N3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['Q3'].style = ali_grey
    
    ws.merge_cells('N4:O4')
    ws['N4'] = 'На ' + str(request.session["electric_data_start"])
    ws['N4'].style = ali_grey
    
    ws.merge_cells('P4:Q4')
    ws['P4'] = 'На ' + str(request.session["electric_data_end"])
    ws['P4'].style = ali_grey

    ws['N5'] = 'Показания'
    ws['N5'].style = ali_grey     
    ws['O5'] = 'Энергия'
    ws['O5'].style = ali_yellow
    
    ws['P5'] = 'Показания'
    ws['P5'].style = ali_grey     
    ws['Q5'] = 'Энергия'
    ws['Q5'].style = ali_yellow
         
    # Расход
    ws.merge_cells('R3:W3')
    ws['R3'] = 'Расход А+, кВт*ч'
    ws['R3'].style = ali_grey
    ws['W3'].style = ali_grey
        # Расход Т0
    ws.merge_cells('R4:S4')
    ws['R4'] = 'Сумма'
    ws['R4'].style = ali_grey
    ws['R5'] = 'Показания'
    ws['R5'].style = ali_grey
    ws['S5'] = 'Энергия'
    ws['S5'].style = ali_yellow
        # Расход Т1
    ws.merge_cells('T4:U4')
    ws['T4'] = 'Tариф 1'
    ws['T4'].style = ali_grey
    ws['T5'] = 'Показания'
    ws['T5'].style = ali_grey
    ws['U5'] = 'Энергия'
    ws['U5'].style = ali_yellow
        # Расход Т2
    ws.merge_cells('V4:W4')
    ws['V4'] = 'Tариф 2'
    ws['V4'].style = ali_grey
    ws['V5'] = 'Показания'
    ws['V5'].style = ali_grey
    ws['W5'] = 'Энергия'
    ws['W5'].style = ali_yellow
  
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']                        
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    data_table_end   = []
    data_table_start = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
                     
                 
    if (bool(is_abonent_level.search(obj_key))):
        if (is_electric_monthly == "1"):
            data_table_end   = common_sql.get_data_table_by_date_monthly_2_zones(meters_name, parent_name, electric_data_end)
            data_table_start = common_sql.get_data_table_by_date_monthly_2_zones(meters_name, parent_name, electric_data_start)
        elif (is_electric_daily == "1"):
            data_table_end   = common_sql.get_data_table_by_date_daily_2_zones(meters_name, parent_name, electric_data_end)
            data_table_start = common_sql.get_data_table_by_date_daily_2_zones(meters_name, parent_name, electric_data_start)


    elif (bool(is_object_level_2.search(obj_key))):
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
            data_table_end   = []
            data_table_start = []
            for x in range(len(list_of_abonents_2)):
                if (is_electric_monthly == "1"):                
                    data_table_temp_end = common_sql.get_data_table_by_date_monthly_2_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                    data_table_temp_start = common_sql.get_data_table_by_date_monthly_2_zones(list_of_abonents_2[x], meters_name, electric_data_start)

                elif (is_electric_daily == "1"):                
                    data_table_temp_end = common_sql.get_data_table_by_date_daily_2_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                    data_table_temp_start = common_sql.get_data_table_by_date_daily_2_zones(list_of_abonents_2[x], meters_name, electric_data_start)

                if data_table_temp_end and data_table_start:
                    data_table_end.extend(data_table_temp_end)
                    data_table_start.extend(data_table_temp_start)
                    
                else:
                    data_table_start.extend([[u'', list_of_abonents_2[x],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
                    data_table_end.extend([[u'', list_of_abonents_2[x],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
#Запрашиваем данные для отчета конец
                    
# Заполняем отчет значениями
    for row in range(6, len(data_table_end)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table_end[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table_end[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (get_k_t_n_by_serial_number(data_table_end[row-6][2]))  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (get_k_a_by_serial_number(data_table_end[row-6][2]))  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table_end[row-6][3])  # Сумма А+ на конец интервала
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table_end[row-6][3]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Энергия Сумма А+ на конец интервала
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table_start[row-6][3])  # Сумма А+ на начало интервала
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table_start[row-6][3]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # Энергия Сумма А+ на начало интервала
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table_end[row-6][4])  # Тариф 1 А+ на конец интервала
            ws.cell('L%s'%(row)).style = ali_white
        except:
            ws.cell('L%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (data_table_end[row-6][4]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Энергия Тариф 1 А+ на конец интервала
            ws.cell('M%s'%(row)).style = ali_yellow
        except:
            ws.cell('M%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table_start[row-6][4])  # Тариф 1 А+ на начало интервала
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table_start[row-6][4]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # Энергия Тариф 1 А+ на начало интервала
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table_end[row-6][5])  # Тариф 2 А+ на конец интервала
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('Q%s'%(row)).value = '%s' % (data_table_end[row-6][5]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Энергия Тариф 2 А+ на конец интервала
            ws.cell('Q%s'%(row)).style = ali_yellow
        except:
            ws.cell('Q%s'%(row)).style = ali_yellow
            next

        try:
            ws.cell('N%s'%(row)).value = '%s' % (data_table_start[row-6][5])  # Тариф 2 А+ на начало интервала
            ws.cell('N%s'%(row)).style = ali_white
        except:
            ws.cell('N%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table_start[row-6][5]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # Энергия Тариф 2 А+ на начало интервала
            ws.cell('O%s'%(row)).style = ali_yellow
        except:
            ws.cell('O%s'%(row)).style = ali_yellow
            next
            
        # Расход
        try:
            ws.cell('R%s'%(row)).value = '%s' % (data_table_end[row-6][3] - data_table_start[row-6][3] )  # Расход Сумма А+
            ws.cell('R%s'%(row)).style = ali_white
        except:
            ws.cell('R%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('S%s'%(row)).value = '%s' % ((data_table_end[row-6][3] - data_table_start[row-6][3])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Сумма Энергия А+
            ws.cell('S%s'%(row)).style = ali_yellow
        except:
            ws.cell('S%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('T%s'%(row)).value = '%s' % (data_table_end[row-6][4] - data_table_start[row-6][4] )  # Расход Тариф 1 А+
            ws.cell('T%s'%(row)).style = ali_white
        except:
            ws.cell('T%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('U%s'%(row)).value = '%s' % ((data_table_end[row-6][4] - data_table_start[row-6][4])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Тариф 1 Энергия А+
            ws.cell('U%s'%(row)).style = ali_yellow
        except:
            ws.cell('U%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('V%s'%(row)).value = '%s' % (data_table_end[row-6][5] - data_table_start[row-6][5] )  # Расход Тариф 2 А+
            ws.cell('V%s'%(row)).style = ali_white
        except:
            ws.cell('V%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('W%s'%(row)).value = '%s' % ((data_table_end[row-6][5] - data_table_start[row-6][5])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Тариф 2 Энергия А+
            ws.cell('W%s'%(row)).style = ali_yellow
        except:
            ws.cell('W%s'%(row)).style = ali_yellow
            next            
# Конец наполнения отчёта
            
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'rashod_2_zones ' + str(electric_data_start) + u' - ' + str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_electric_potreblenie_3_zones(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление электроэнергии в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F3:I3')
    ws['F3'] = 'Сумма A+, кВт*ч'
    ws['F3'].style = ali_grey
    ws['G3'].style = ali_grey
    ws['H3'].style = ali_grey
    ws['I3'].style = ali_grey
    
    ws.merge_cells('F4:G4')
    ws['F4'] = 'На ' + str(request.session["electric_data_start"])
    ws['F4'].style = ali_grey
    
    ws.merge_cells('H4:I4')
    ws['H4'] = 'На ' + str(request.session["electric_data_end"])
    ws['H4'].style = ali_grey

    ws['F5'] = 'Показания'
    ws['F5'].style = ali_grey     
    ws['G5'] = 'Энергия'
    ws['G5'].style = ali_yellow
    
    ws['H5'] = 'Показания'
    ws['H5'].style = ali_grey     
    ws['I5'] = 'Энергия'
    ws['I5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('J3:M3')
    ws['J3'] = 'Тариф 1 A+, кВт*ч'
    ws['J3'].style = ali_grey
    ws['K3'].style = ali_grey
    ws['L3'].style = ali_grey
    ws['M3'].style = ali_grey
    
    ws.merge_cells('J4:K4')
    ws['J4'] = 'На ' + str(request.session["electric_data_start"])
    ws['J4'].style = ali_grey
    
    ws.merge_cells('L4:M4')
    ws['L4'] = 'На ' + str(request.session["electric_data_end"])
    ws['L4'].style = ali_grey

    ws['J5'] = 'Показания'
    ws['J5'].style = ali_grey     
    ws['K5'] = 'Энергия'
    ws['K5'].style = ali_yellow
    
    ws['L5'] = 'Показания'
    ws['L5'].style = ali_grey     
    ws['M5'] = 'Энергия'
    ws['M5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('N3:Q3')
    ws['N3'] = 'Тариф 2 A+, кВт*ч'
    ws['N3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['Q3'].style = ali_grey
    
    ws.merge_cells('N4:O4')
    ws['N4'] = 'На ' + str(request.session["electric_data_start"])
    ws['N4'].style = ali_grey
    
    ws.merge_cells('P4:Q4')
    ws['P4'] = 'На ' + str(request.session["electric_data_end"])
    ws['P4'].style = ali_grey

    ws['N5'] = 'Показания'
    ws['N5'].style = ali_grey     
    ws['O5'] = 'Энергия'
    ws['O5'].style = ali_yellow
    
    ws['P5'] = 'Показания'
    ws['P5'].style = ali_grey     
    ws['Q5'] = 'Энергия'
    ws['Q5'].style = ali_yellow
    
    # Тариф 3
    ws.merge_cells('R3:U3')
    ws['R3'] = 'Тариф 3 A+, кВт*ч'
    ws['R3'].style = ali_grey
    ws['S3'].style = ali_grey
    ws['T3'].style = ali_grey
    ws['U3'].style = ali_grey
    
    ws.merge_cells('R4:S4')
    ws['R4'] = 'На ' + str(request.session["electric_data_start"])
    ws['R4'].style = ali_grey
    
    ws.merge_cells('T4:U4')
    ws['T4'] = 'На ' + str(request.session["electric_data_end"])
    ws['T4'].style = ali_grey

    ws['R5'] = 'Показания'
    ws['R5'].style = ali_grey     
    ws['S5'] = 'Энергия'
    ws['S5'].style = ali_yellow
    
    ws['T5'] = 'Показания'
    ws['T5'].style = ali_grey     
    ws['U5'] = 'Энергия'
    ws['U5'].style = ali_yellow
         
    # Расход
    ws.merge_cells('V3:AC3')
    ws['V3'] = 'Расход А+, кВт*ч'
    ws['V3'].style = ali_grey
    ws['AC3'].style = ali_grey
        # Расход Т0
    ws.merge_cells('V4:W4')
    ws['V4'] = 'Сумма'
    ws['V4'].style = ali_grey
    ws['V5'] = 'Показания'
    ws['V5'].style = ali_grey
    ws['W5'] = 'Энергия'
    ws['W5'].style = ali_yellow
        # Расход Т1
    ws.merge_cells('X4:Y4')
    ws['X4'] = 'Tариф 1'
    ws['X4'].style = ali_grey
    ws['X5'] = 'Показания'
    ws['X5'].style = ali_grey
    ws['Y5'] = 'Энергия'
    ws['Y5'].style = ali_yellow
        # Расход Т2
    ws.merge_cells('Z4:AA4')
    ws['Z4'] = 'Tариф 2'
    ws['Z4'].style = ali_grey
    ws['Z5'] = 'Показания'
    ws['Z5'].style = ali_grey
    ws['AA5'] = 'Энергия'
    ws['AA5'].style = ali_yellow
        # Расход Т3
    ws.merge_cells('AB4:AC4')
    ws['AB4'] = 'Tариф 3'
    ws['AB4'].style = ali_grey
    ws['AC4'].style = ali_grey
    ws['AB5'] = 'Показания'
    ws['AB5'].style = ali_grey
    ws['AC5'] = 'Энергия'
    ws['AC5'].style = ali_yellow
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']                        
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    data_table_end   = []
    data_table_start = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
                     
                 
    if (bool(is_abonent_level.search(obj_key))):
        if (is_electric_monthly == "1"):
            data_table_end   = common_sql.get_data_table_by_date_monthly_3_zones(meters_name, parent_name, electric_data_end)
            data_table_start = common_sql.get_data_table_by_date_monthly_3_zones(meters_name, parent_name, electric_data_start)
        elif (is_electric_daily == "1"):
            data_table_end   = common_sql.get_data_table_by_date_daily_3_zones(meters_name, parent_name, electric_data_end)
            data_table_start = common_sql.get_data_table_by_date_daily_3_zones(meters_name, parent_name, electric_data_start)


    elif (bool(is_object_level_2.search(obj_key))):
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
            data_table_end   = []
            data_table_start = []
            for x in range(len(list_of_abonents_2)):
                if (is_electric_monthly == "1"):                
                    data_table_temp_end = common_sql.get_data_table_by_date_monthly_3_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                    data_table_temp_start = common_sql.get_data_table_by_date_monthly_3_zones(list_of_abonents_2[x], meters_name, electric_data_start)

                elif (is_electric_daily == "1"):                
                    data_table_temp_end = common_sql.get_data_table_by_date_daily_3_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                    data_table_temp_start = common_sql.get_data_table_by_date_daily_3_zones(list_of_abonents_2[x], meters_name, electric_data_start)

                if data_table_temp_end and data_table_start:
                    data_table_end.extend(data_table_temp_end)
                    data_table_start.extend(data_table_temp_start)
                    
                else:
                    data_table_start.extend([[u'', list_of_abonents_2[x],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
                    data_table_end.extend([[u'', list_of_abonents_2[x],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
#Запрашиваем данные для отчета конец
                    
# Заполняем отчет значениями
    for row in range(6, len(data_table_end)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table_end[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table_end[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (get_k_t_n_by_serial_number(data_table_end[row-6][2]))  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (get_k_a_by_serial_number(data_table_end[row-6][2]))  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table_end[row-6][3])  # Сумма А+ на конец интервала
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table_end[row-6][3]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Энергия Сумма А+ на конец интервала
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table_start[row-6][3])  # Сумма А+ на начало интервала
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table_start[row-6][3]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # Энергия Сумма А+ на начало интервала
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table_end[row-6][4])  # Тариф 1 А+ на конец интервала
            ws.cell('L%s'%(row)).style = ali_white
        except:
            ws.cell('L%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (data_table_end[row-6][4]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Энергия Тариф 1 А+ на конец интервала
            ws.cell('M%s'%(row)).style = ali_yellow
        except:
            ws.cell('M%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table_start[row-6][4])  # Тариф 1 А+ на начало интервала
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table_start[row-6][4]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # Энергия Тариф 1 А+ на начало интервала
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table_end[row-6][5])  # Тариф 2 А+ на конец интервала
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('Q%s'%(row)).value = '%s' % (data_table_end[row-6][5]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Энергия Тариф 2 А+ на конец интервала
            ws.cell('Q%s'%(row)).style = ali_yellow
        except:
            ws.cell('Q%s'%(row)).style = ali_yellow
            next

        try:
            ws.cell('N%s'%(row)).value = '%s' % (data_table_start[row-6][5])  # Тариф 2 А+ на начало интервала
            ws.cell('N%s'%(row)).style = ali_white
        except:
            ws.cell('N%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table_start[row-6][5]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # Энергия Тариф 2 А+ на начало интервала
            ws.cell('O%s'%(row)).style = ali_yellow
        except:
            ws.cell('O%s'%(row)).style = ali_yellow
            next
            

            
        try:
            ws.cell('T%s'%(row)).value = '%s' % (data_table_end[row-6][6])  # Тариф 3 А+ на конец интервала
            ws.cell('T%s'%(row)).style = ali_white
        except:
            ws.cell('T%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('U%s'%(row)).value = '%s' % (data_table_end[row-6][6]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # "Энергия Тариф 3 А+ на конец интервала
            ws.cell('U%s'%(row)).style = ali_yellow
        except:
            ws.cell('U%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('R%s'%(row)).value = '%s' % (data_table_start[row-6][6])  # Тариф 3 А+ на начало интервала
            ws.cell('R%s'%(row)).style = ali_white
        except:
            ws.cell('R%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('S%s'%(row)).value = '%s' % (data_table_start[row-6][6]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # "Энергия Тариф 3 А+ на начало интервала
            ws.cell('S%s'%(row)).style = ali_yellow
        except:
            ws.cell('S%s'%(row)).style = ali_yellow
            next
        # Расход
        try:
            ws.cell('V%s'%(row)).value = '%s' % (data_table_end[row-6][3] - data_table_start[row-6][3] )  # Расход Сумма А+
            ws.cell('V%s'%(row)).style = ali_white
        except:
            ws.cell('V%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('W%s'%(row)).value = '%s' % ((data_table_end[row-6][3] - data_table_start[row-6][3])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Сумма Энергия А+
            ws.cell('W%s'%(row)).style = ali_yellow
        except:
            ws.cell('W%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('X%s'%(row)).value = '%s' % (data_table_end[row-6][4] - data_table_start[row-6][4] )  # Расход Тариф 1 А+
            ws.cell('X%s'%(row)).style = ali_white
        except:
            ws.cell('X%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('Y%s'%(row)).value = '%s' % ((data_table_end[row-6][4] - data_table_start[row-6][4])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Тариф 1 Энергия А+
            ws.cell('Y%s'%(row)).style = ali_yellow
        except:
            ws.cell('Y%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('Z%s'%(row)).value = '%s' % (data_table_end[row-6][5] - data_table_start[row-6][5] )  # Расход Тариф 2 А+
            ws.cell('Z%s'%(row)).style = ali_white
        except:
            ws.cell('Z%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('AA%s'%(row)).value = '%s' % ((data_table_end[row-6][5] - data_table_start[row-6][5])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Тариф 2 Энергия А+
            ws.cell('AA%s'%(row)).style = ali_yellow
        except:
            ws.cell('AA%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('AB%s'%(row)).value = '%s' % (data_table_end[row-6][6] - data_table_start[row-6][6] )  # Расход Тариф 3 А+
            ws.cell('AB%s'%(row)).style = ali_white
        except:
            ws.cell('AB%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('AC%s'%(row)).value = '%s' % ((data_table_end[row-6][6] - data_table_start[row-6][6])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Тариф 3 Энергия А+
            ws.cell('AC%s'%(row)).style = ali_yellow
        except:
            ws.cell('AC%s'%(row)).style = ali_yellow
            next
# Конец наполнения отчёта
            
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'rashod_3_zones ' + str(electric_data_start) + u' - ' + str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_electric_potreblenie_3_zones_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+'. Потребление электроэнергии в период с ' + electric_data_start + ' по ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F3:I3')
    ws['F3'] = 'Сумма A+, кВт*ч'
    ws['F3'].style = ali_grey
    ws['G3'].style = ali_grey
    ws['H3'].style = ali_grey
    ws['I3'].style = ali_grey
    
    ws.merge_cells('F4:G4')
    ws['F4'] = 'На ' + str(request.session["electric_data_start"])
    ws['F4'].style = ali_grey
    
    ws.merge_cells('H4:I4')
    ws['H4'] = 'На ' + str(request.session["electric_data_end"])
    ws['H4'].style = ali_grey

    ws['F5'] = 'Показания'
    ws['F5'].style = ali_grey     
    ws['G5'] = 'Энергия'
    ws['G5'].style = ali_yellow
    
    ws['H5'] = 'Показания'
    ws['H5'].style = ali_grey     
    ws['I5'] = 'Энергия'
    ws['I5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('J3:M3')
    ws['J3'] = 'Тариф 1 A+, кВт*ч'
    ws['J3'].style = ali_grey
    ws['K3'].style = ali_grey
    ws['L3'].style = ali_grey
    ws['M3'].style = ali_grey
    
    ws.merge_cells('J4:K4')
    ws['J4'] = 'На ' + str(request.session["electric_data_start"])
    ws['J4'].style = ali_grey
    
    ws.merge_cells('L4:M4')
    ws['L4'] = 'На ' + str(request.session["electric_data_end"])
    ws['L4'].style = ali_grey

    ws['J5'] = 'Показания'
    ws['J5'].style = ali_grey     
    ws['K5'] = 'Энергия'
    ws['K5'].style = ali_yellow
    
    ws['L5'] = 'Показания'
    ws['L5'].style = ali_grey     
    ws['M5'] = 'Энергия'
    ws['M5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('N3:Q3')
    ws['N3'] = 'Тариф 2 A+, кВт*ч'
    ws['N3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['Q3'].style = ali_grey
    
    ws.merge_cells('N4:O4')
    ws['N4'] = 'На ' + str(request.session["electric_data_start"])
    ws['N4'].style = ali_grey
    
    ws.merge_cells('P4:Q4')
    ws['P4'] = 'На ' + str(request.session["electric_data_end"])
    ws['P4'].style = ali_grey

    ws['N5'] = 'Показания'
    ws['N5'].style = ali_grey     
    ws['O5'] = 'Энергия'
    ws['O5'].style = ali_yellow
    
    ws['P5'] = 'Показания'
    ws['P5'].style = ali_grey     
    ws['Q5'] = 'Энергия'
    ws['Q5'].style = ali_yellow
    
    # Тариф 3
    ws.merge_cells('R3:U3')
    ws['R3'] = 'Тариф 3 A+, кВт*ч'
    ws['R3'].style = ali_grey
    ws['S3'].style = ali_grey
    ws['T3'].style = ali_grey
    ws['U3'].style = ali_grey
    
    ws.merge_cells('R4:S4')
    ws['R4'] = 'На ' + str(request.session["electric_data_start"])
    ws['R4'].style = ali_grey
    
    ws.merge_cells('T4:U4')
    ws['T4'] = 'На ' + str(request.session["electric_data_end"])
    ws['T4'].style = ali_grey

    ws['R5'] = 'Показания'
    ws['R5'].style = ali_grey     
    ws['S5'] = 'Энергия'
    ws['S5'].style = ali_yellow
    
    ws['T5'] = 'Показания'
    ws['T5'].style = ali_grey     
    ws['U5'] = 'Энергия'
    ws['U5'].style = ali_yellow
         
    # Расход
    ws.merge_cells('V3:AC3')
    ws['V3'] = 'Расход А+, кВт*ч'
    ws['V3'].style = ali_grey
    ws['AC3'].style = ali_grey
        # Расход Т0
    ws.merge_cells('V4:W4')
    ws['V4'] = 'Сумма'
    ws['V4'].style = ali_grey
    ws['V5'] = 'Показания'
    ws['V5'].style = ali_grey
    ws['W5'] = 'Энергия'
    ws['W5'].style = ali_yellow
        # Расход Т1
    ws.merge_cells('X4:Y4')
    ws['X4'] = 'Tариф 1'
    ws['X4'].style = ali_grey
    ws['X5'] = 'Показания'
    ws['X5'].style = ali_grey
    ws['Y5'] = 'Энергия'
    ws['Y5'].style = ali_yellow
        # Расход Т2
    ws.merge_cells('Z4:AA4')
    ws['Z4'] = 'Tариф 2'
    ws['Z4'].style = ali_grey
    ws['Z5'] = 'Показания'
    ws['Z5'].style = ali_grey
    ws['AA5'] = 'Энергия'
    ws['AA5'].style = ali_yellow
        # Расход Т3
    ws.merge_cells('AB4:AC4')
    ws['AB4'] = 'Tариф 3'
    ws['AB4'].style = ali_grey
    ws['AC4'].style = ali_grey
    ws['AB5'] = 'Показания'
    ws['AB5'].style = ali_grey
    ws['AC5'] = 'Энергия'
    ws['AC5'].style = ali_yellow
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
    
    obj_key             = request.session['obj_key']
    is_electric_delta  = request.session['is_electric_delta']
    is_electric_monthly=request.session['is_electric_monthly']
    data_table = []
    if True:
        if True:                        
            res=u'Электричество'
            
            if (is_electric_monthly=="1"):
                dm='monthly'
            else:
                dm='daily'
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                    isAbon=True                    
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
                
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                    isAbon=False
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
            #*********************************************************************************************************************************************************************
            elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
                    data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
                    request.session["data_table_export"] = data_table
#Запрашиваем данные для отчета конец
                    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][23])  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][20])  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][24])  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # Сумма А+ на конец интервала
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][7]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Сумма А+ на конец интервала
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][2])  # Сумма А+ на начало интервала
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][2]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Сумма А+ на начало интервала
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][8])  # Тариф 1 А+ на конец интервала
            ws.cell('L%s'%(row)).style = ali_white
        except:
            ws.cell('L%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (data_table[row-6][8]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 1 А+ на конец интервала
            ws.cell('M%s'%(row)).style = ali_yellow
        except:
            ws.cell('M%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][3])  # Тариф 1 А+ на начало интервала
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][3]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 1 А+ на начало интервала
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-6][9])  # Тариф 2 А+ на конец интервала
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('Q%s'%(row)).value = '%s' % (data_table[row-6][9]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 2 А+ на конец интервала
            ws.cell('Q%s'%(row)).style = ali_yellow
        except:
            ws.cell('Q%s'%(row)).style = ali_yellow
            next

        try:
            ws.cell('N%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тариф 2 А+ на начало интервала
            ws.cell('N%s'%(row)).style = ali_white
        except:
            ws.cell('N%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-6][4]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 2 А+ на начало интервала
            ws.cell('O%s'%(row)).style = ali_yellow
        except:
            ws.cell('O%s'%(row)).style = ali_yellow
            next
            

            
        try:
            ws.cell('T%s'%(row)).value = '%s' % (data_table[row-6][10])  # Тариф 3 А+ на конец интервала
            ws.cell('T%s'%(row)).style = ali_white
        except:
            ws.cell('T%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('U%s'%(row)).value = '%s' % (data_table[row-6][10]**data_table[row-6][20]*data_table[row-6][23])  # "Энергия Тариф 3 А+ на конец интервала
            ws.cell('U%s'%(row)).style = ali_yellow
        except:
            ws.cell('U%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('R%s'%(row)).value = '%s' % (data_table[row-6][5])  # Тариф 3 А+ на начало интервала
            ws.cell('R%s'%(row)).style = ali_white
        except:
            ws.cell('R%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('S%s'%(row)).value = '%s' % (data_table[row-6][5]**data_table[row-6][20]*data_table[row-6][23])  # "Энергия Тариф 3 А+ на начало интервала
            ws.cell('S%s'%(row)).style = ali_yellow
        except:
            ws.cell('S%s'%(row)).style = ali_yellow
            next
        # Расход
        try:
            ws.cell('V%s'%(row)).value = '%s' % (data_table[row-6][12])  # Расход Сумма А+
            ws.cell('V%s'%(row)).style = ali_white
        except:
            ws.cell('V%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('W%s'%(row)).value = '%s' % (data_table[row-6][12]*data_table[row-6][20]*data_table[row-6][23])  # Расход Сумма Энергия А+
            ws.cell('W%s'%(row)).style = ali_yellow
        except:
            ws.cell('W%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('X%s'%(row)).value = '%s' % (data_table[row-6][13])  # Расход Тариф 1 А+
            ws.cell('X%s'%(row)).style = ali_white
        except:
            ws.cell('X%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('Y%s'%(row)).value = '%s' % (data_table[row-6][13]*data_table[row-6][20]*data_table[row-6][23])  # Расход Тариф 1 Энергия А+
            ws.cell('Y%s'%(row)).style = ali_yellow
        except:
            ws.cell('Y%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('Z%s'%(row)).value = '%s' % (data_table[row-6][14])  # Расход Тариф 2 А+
            ws.cell('Z%s'%(row)).style = ali_white
        except:
            ws.cell('Z%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('AA%s'%(row)).value = '%s' % (data_table[row-6][14]*data_table[row-6][20]*data_table[row-6][23])  # Расход Тариф 2 Энергия А+
            ws.cell('AA%s'%(row)).style = ali_yellow
        except:
            ws.cell('AA%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('AB%s'%(row)).value = '%s' % (data_table[row-6][15])  # Расход Тариф 3 А+
            ws.cell('AB%s'%(row)).style = ali_white
        except:
            ws.cell('AB%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('AC%s'%(row)).value = '%s' % (data_table[row-6][15]*data_table[row-6][20]*data_table[row-6][23])  # Расход Тариф 3 Энергия А+
            ws.cell('AC%s'%(row)).style = ali_yellow
        except:
            ws.cell('AC%s'%(row)).style = ali_yellow
            next
# Конец наполнения отчёта
            
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'rashod_3_zones_'+translate(obj_title)+'_' + str(electric_data_start) + u'-' + str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_electric_potreblenie_2_zones_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление электроэнергии в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F3:I3')
    ws['F3'] = 'Сумма A+, кВт*ч'
    ws['F3'].style = ali_grey
    ws['G3'].style = ali_grey
    ws['H3'].style = ali_grey
    ws['I3'].style = ali_grey
    
    ws.merge_cells('F4:G4')
    ws['F4'] = 'На ' + str(request.session["electric_data_start"])
    ws['F4'].style = ali_grey
    
    ws.merge_cells('H4:I4')
    ws['H4'] = 'На ' + str(request.session["electric_data_end"])
    ws['H4'].style = ali_grey

    ws['F5'] = 'Показания'
    ws['F5'].style = ali_grey     
    ws['G5'] = 'Энергия'
    ws['G5'].style = ali_yellow
    
    ws['H5'] = 'Показания'
    ws['H5'].style = ali_grey     
    ws['I5'] = 'Энергия'
    ws['I5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('J3:M3')
    ws['J3'] = 'Тариф 1 A+, кВт*ч'
    ws['J3'].style = ali_grey
    ws['K3'].style = ali_grey
    ws['L3'].style = ali_grey
    ws['M3'].style = ali_grey
    
    ws.merge_cells('J4:K4')
    ws['J4'] = 'На ' + str(request.session["electric_data_start"])
    ws['J4'].style = ali_grey
    
    ws.merge_cells('L4:M4')
    ws['L4'] = 'На ' + str(request.session["electric_data_end"])
    ws['L4'].style = ali_grey

    ws['J5'] = 'Показания'
    ws['J5'].style = ali_grey     
    ws['K5'] = 'Энергия'
    ws['K5'].style = ali_yellow
    
    ws['L5'] = 'Показания'
    ws['L5'].style = ali_grey     
    ws['M5'] = 'Энергия'
    ws['M5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('N3:Q3')
    ws['N3'] = 'Тариф 2 A+, кВт*ч'
    ws['N3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['Q3'].style = ali_grey
    
    ws.merge_cells('N4:O4')
    ws['N4'] = 'На ' + str(request.session["electric_data_start"])
    ws['N4'].style = ali_grey
    
    ws.merge_cells('P4:Q4')
    ws['P4'] = 'На ' + str(request.session["electric_data_end"])
    ws['P4'].style = ali_grey

    ws['N5'] = 'Показания'
    ws['N5'].style = ali_grey     
    ws['O5'] = 'Энергия'
    ws['O5'].style = ali_yellow
    
    ws['P5'] = 'Показания'
    ws['P5'].style = ali_grey     
    ws['Q5'] = 'Энергия'
    ws['Q5'].style = ali_yellow
    

         
    # Расход
    ws.merge_cells('R3:W3')
    ws['R3'] = 'Расход А+, кВт*ч'
    ws['R3'].style = ali_grey
    ws['W3'].style = ali_grey
        # Расход Т0
    ws.merge_cells('R4:S4')
    ws['R4'] = 'Сумма'
    ws['R4'].style = ali_grey
    ws['R5'] = 'Показания'
    ws['R5'].style = ali_grey
    ws['S5'] = 'Энергия'
    ws['S5'].style = ali_yellow
        # Расход Т1
    ws.merge_cells('T4:U4')
    ws['T4'] = 'Tариф 1'
    ws['T4'].style = ali_grey
    ws['T5'] = 'Показания'
    ws['T5'].style = ali_grey
    ws['U5'] = 'Энергия'
    ws['U5'].style = ali_yellow
        # Расход Т2
    ws.merge_cells('V4:W4')
    ws['V4'] = 'Tариф 2'
    ws['V4'].style = ali_grey
    ws['V5'] = 'Показания'
    ws['V5'].style = ali_grey
    ws['W5'] = 'Энергия'
    ws['W5'].style = ali_yellow
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
    obj_key             = request.session['obj_key']
    is_electric_delta  = request.session['is_electric_delta']
    is_electric_monthly=request.session['is_electric_monthly']
    data_table = []
    if True:
        if True:                        
            res=u'Электричество'
            
            if (is_electric_monthly=="1"):
                dm='monthly'
            else:
                dm='daily'
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                    isAbon=True                    
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
                
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                    isAbon=False
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
            #*********************************************************************************************************************************************************************
            elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
                    data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
                    request.session["data_table_export"] = data_table
#Запрашиваем данные для отчета конец
                    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][23])  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][20])  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][24])  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # Сумма А+ на конец интервала
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][7]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Сумма А+ на конец интервала
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][2])  # Сумма А+ на начало интервала
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][2]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Сумма А+ на начало интервала
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][8])  # Тариф 1 А+ на конец интервала
            ws.cell('L%s'%(row)).style = ali_white
        except:
            ws.cell('L%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (data_table[row-6][8]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 1 А+ на конец интервала
            ws.cell('M%s'%(row)).style = ali_yellow
        except:
            ws.cell('M%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][3])  # Тариф 1 А+ на начало интервала
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][3]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 1 А+ на начало интервала
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-6][9])  # Тариф 2 А+ на конец интервала
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('Q%s'%(row)).value = '%s' % (data_table[row-6][9]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 2 А+ на конец интервала
            ws.cell('Q%s'%(row)).style = ali_yellow
        except:
            ws.cell('Q%s'%(row)).style = ali_yellow
            next

        try:
            ws.cell('N%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тариф 2 А+ на начало интервала
            ws.cell('N%s'%(row)).style = ali_white
        except:
            ws.cell('N%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-6][4]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 2 А+ на начало интервала
            ws.cell('O%s'%(row)).style = ali_yellow
        except:
            ws.cell('O%s'%(row)).style = ali_yellow
            next
            

        # Расход
        try:
            ws.cell('R%s'%(row)).value = '%s' % (data_table[row-6][12])  # Расход Сумма А+
            ws.cell('R%s'%(row)).style = ali_white
        except:
            ws.cell('R%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('S%s'%(row)).value = '%s' % (data_table[row-6][12]*data_table[row-6][20]*data_table[row-6][23])  # Расход Сумма Энергия А+
            ws.cell('S%s'%(row)).style = ali_yellow
        except:
            ws.cell('S%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('T%s'%(row)).value = '%s' % (data_table[row-6][13])  # Расход Тариф 1 А+
            ws.cell('T%s'%(row)).style = ali_white
        except:
            ws.cell('T%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('U%s'%(row)).value = '%s' % (data_table[row-6][13]*data_table[row-6][20]*data_table[row-6][23])  # Расход Тариф 1 Энергия А+
            ws.cell('U%s'%(row)).style = ali_yellow
        except:
            ws.cell('U%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('V%s'%(row)).value = '%s' % (data_table[row-6][14])  # Расход Тариф 2 А+
            ws.cell('V%s'%(row)).style = ali_white
        except:
            ws.cell('V%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('W%s'%(row)).value = '%s' % (data_table[row-6][14]*data_table[row-6][20]*data_table[row-6][23])  # Расход Тариф 2 Энергия А+
            ws.cell('W%s'%(row)).style = ali_yellow
        except:
            ws.cell('W%s'%(row)).style = ali_yellow
            next

# Конец наполнения отчёта
            
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'rashod_2_zones ' + str(electric_data_start) + u' - ' + str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response


def pokazaniya_heat_report(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания теплосчётчиков на ' + str(request.session["electric_data_end"])
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Заводской номер'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания, Гкал'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания, м3'
    ws['D5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']                        
    obj_key             = request.session['obj_key']

    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            #request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            #request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']  

    list_except = [u'ВРУ Счётчик01',u'ВРУ Счётчик02',u'ВРУ Счётчик03',u'ВРУ Счётчик04',u'ВРУ Счётчик05',u'ВРУ Счётчик06',u'ВРУ Счётчик07',u'ВРУ Счётчик08',u'ВРУ Счётчик09',u'ВРУ Счётчик10',u'ВРУ Счётчик11',u'ВРУ Счётчик12',u'ВРУ Счётчик13',u'ВРУ Счётчик14',u'ВРУ Счётчик15',u'ВРУ Счётчик16',u'ВРУ Счётчик17',u'ВРУ Счётчик18',u'ВРУ Счётчик19',u'ВРУ Счётчик20',u'ВРУ Счётчик21',u'ВРУ Счётчик22',u'ВРУ Счётчик23',u'Гараж Счётчик 1',u'Гараж Счётчик 2']
                     
    if (bool(is_abonent_level.search(obj_key))):     
        data_table = common_sql.get_data_table_by_date_heat(meters_name, parent_name, electric_data_end)

    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []
        for x in range(len(list_of_abonents_2)):
            data_table_temp = common_sql.get_data_table_by_date_heat(list_of_abonents_2[x], meters_name, electric_data_end)

            if list_of_abonents_2[x][0] in list_except:
                next
            elif data_table_temp:            
                data_table.extend(data_table_temp)
            else:
                data_table.extend([[0,list_of_abonents_2[x][0],u'Н/Д',u'Н/Д',u'Н/Д']])
                             
    else:
        data_table = []

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по расходу воды
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17 
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'heat_report' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def pokazaniya_heat_report_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания теплосчётчиков на ' + str(request.session["electric_data_end"])
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Заводской номер'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания, Гкал'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания, м3'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Время работы, ч'
    ws['E5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']

    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            #request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            #request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']  

    list_except = []
                     
    data_table = []
    list_except = []
    if (bool(is_abonent_level.search(obj_key))):     
        data_table = common_sql.get_data_table_by_date_heat_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_v2(meters_name, parent_name, electric_data_end, False)
        for row in data_table:
            for x in list_except:
                if x==row[2]:
                    data_table.remove(x)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по расходу воды
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # время работы
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17 
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'heat_report' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pokazaniya_sayany(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    meters_name           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = meters_name+'. Показания теплосчётчиков Саяны на ' + electric_data_end 
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Дата'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания Q1, Гкал'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания M1, т'
    ws['E5'].style = ali_grey
    
    ws['F5'] = u't1, C˚'
    ws['F5'].style = ali_grey
    
    ws['G5'] = u't2, C˚'
    ws['G5'].style = ali_grey

# ниже не переделывала
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    
    parent_name    = request.session['obj_parent_title']
    
    obj_key             = request.session['obj_key']

    
    #print parent_name,meters_name,electric_data_end, obj_key
    
    data_table = []
#    if request.is_ajax():
#        #if request.method == 'GET':
#            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
#            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
#            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
#            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # Дата
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу Q1
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу Q2
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания по теплу t1
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу t2
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17 
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'heat_sayany_report_'+translate(meters_name)+'_'+electric_data_end 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pokazaniya_sayany_archive(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания теплосчётчиков Саяны на ' + str(request.session["electric_data_end"])
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Дата'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания Q1'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания Q2'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 't1'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 't2'
    ws['G5'].style = ali_grey

# ниже не переделывала
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    
    parent_name    = request.session['obj_parent_title']
    meters_name           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']

    
    data_table = []

    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # Дата
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу Q1
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу Q2
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания по теплу t1
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу t2
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17 

    import zipfile

    o=StringIO.StringIO()
    
    zf = zipfile.ZipFile(response, mode='w', compression=zipfile.ZIP_DEFLATED)
    zf.writestr('README.txt', 'test msg')
    wb.save(o)
    zf.writestr('test.xlsx',o.getvalue())
    zf.close()
    response=HttpResponse(response.getvalue())
    response['Content-Type'] = 'application/x-zip-compressed'
    response['Content-Disposition'] = "attachment; filename=\"sayani_test.zip\""
    
    return response
    
def report_sayany_last(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

    meters_name           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = meters_name+'. Показания теплосчётчиков Саяны на ' + electric_data_end
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Дата'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания Q1, Гкал'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания М1, т'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 't1, C˚'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 't2, C˚'
    ws['G5'].style = ali_grey

# ниже не переделывала
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    
    parent_name    = request.session['obj_parent_title']
    meters_name           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']

    
    #print parent_name,meters_name,electric_data_end, obj_key
    
    data_table = []
#    if request.is_ajax():
#        #if request.method == 'GET':
#            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
#            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
#            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
#            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, False)
    
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        if (data_table[i][3] is None):
            #print data_table[i][1], meters_name
            data_table[i][0]=electric_data_end
            dt=common_sql.get_data_table_by_date_heat_sayany_v2(data_table[i][1], meters_name, None, True)
            if (len(dt)>0):
                data_table[i]=dt[0]
        data_table[i]=tuple(data_table[i])
    
    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # Дата
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу Q1
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу Q2
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания по теплу t1
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу t2
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17 
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'heat_sayany_report_'+translate(meters_name)+'_' +electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_heat_potreblenie_sayany(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
#Шапка
    ws.merge_cells('A2:G2')
    ws['A2'] = meters_name+'. Потребление по теплосчётчикам Sayany в период с ' + electric_data_start + ' по ' +electric_data_end
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания Q1 на '  + electric_data_start+u', Гкал'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания Q1 на '  + electric_data_end+u', Гкал'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление Q1, Гкал'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Показания M1 на '  + electric_data_start+u', т'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Показания M1 на '  + electric_data_end+u', т'
    ws['G5'].style = ali_grey
    
    ws['H5'] = 'Потребление M1, т'
    ws['H5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
                        
    obj_key             = request.session['obj_key']

    data_table = []
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_period_heat_sayany(meters_name, parent_name,electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_period_heat_sayany(meters_name, parent_name,electric_data_start, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания Q1 на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания Q1 на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление Q1
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания M1 на начало
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания M1 на конец
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # Потребление M1
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
#    ws.column_dimensions['C'].width = 35
#    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['H'].width = 18
#    ws.column_dimensions['F'].width = 18
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'potreblenie_heat_report_'+translate(meters_name)+u'_'+electric_data_start+u'-'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_water_potreblenie_pulsar(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
    
#Шапка
    ws.merge_cells('A2:G2')
    ws['A2'] = meters_name+'. Потребление по импульсным водосчётчикам в период с ' + electric_data_start + ' по ' +electric_data_end
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Тип ресурса '
    ws['C5'].style = ali_grey
    
    ws['d5'] = 'Показания на '  + electric_data_start+u', м3'
    ws['d5'].style = ali_grey
    
    ws['e5'] = 'Показания на '  + electric_data_end+u', м3'
    ws['e5'].style = ali_grey
    
    ws['f5'] = 'Потребление, м3'
    ws['f5'].style = ali_grey
    
    ws['g5'] = 'Лицевой номер '
    ws['g5'].style = ali_grey    
    



    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'level2')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level1')

    parent_name         = request.session['obj_parent_title']
    obj_key             = request.session['obj_key']
    data_table=[]
    
    if (bool(is_abonent_level.search(obj_key))): 
        #print 'test!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!'
        data_table = common_sql.get_data_table_water_period_pulsar(meters_name, parent_name,electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_period_pulsar(meters_name, parent_name,electric_data_start, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        num=data_table[i][3]
        if ('ХВС, №' in num) or ('ГВС, №' in num):
            num=num.replace(u'ХВС, №', ' ')
            num=num.replace(u'ГВС, №', ' ')
            data_table[i][3]=num
            #print num
        data_table[i]=tuple(data_table[i])

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][3])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тип ресурса
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания на начало
            ws.cell('d%s'%(row)).style = ali_white
        except:
            ws.cell('d%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания  на конец
            ws.cell('e%s'%(row)).style = ali_white
        except:
            ws.cell('e%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('f%s'%(row)).value = '%s' % (data_table[row-6][7])  # Потребление
            ws.cell('f%s'%(row)).style = ali_white
        except:
            ws.cell('f%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('g%s'%(row)).value = '%s' % (data_table[row-6][1])  # Лицевой номер
            ws.cell('g%s'%(row)).style = ali_white
        except:
            ws.cell('g%s'%(row)).style = ali_white
            next


    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    #ws.column_dimensions['H'].width = 25
    ws.column_dimensions['F'].width = 18
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'potreblenie_water_report_'+translate(meters_name)+u'_'+electric_data_start+u'-'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def pokazaniya_water_current_report(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Последние считанные показания по ХВС и ГВС на ' + str(datetime.now())
    
    ws['A5'] = 'Дата'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Время'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Абонент'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Заводской номер счётчика'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания ХВС, м3'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Показания ГВС, м3'
    ws['F5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']

    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    
    data_table=[]
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_current_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end,  True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table_temp=common_sql.get_current_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end,  False)
        for row in data_table_temp:
            if row[4]==u'Н/Д' and row[5]==u'Н/Д':
                row2=common_sql.get_current_water_gvs_hvs(unicode(row[2]), unicode(row[6]) , electric_data_end, True)
                #print row2
                #print unicode(row[2]), unicode(row[6]), electric_data_end, True
                if len(row2)==0:
                    r=[unicode(electric_data_end), u'Н/Д', unicode(row[2]),unicode(row[3]), u'Н/Д', u'Н/Д']
                    data_table.append(r)
                else:
                    data_table.append(row2[0])
            else:
                data_table.append(row)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Дата
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # время
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # абонент
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # заводской номер
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # хвс
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # гвс
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 17 
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'water_report_current' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def pokazaniya_water_daily_report(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания импульсные по ХВС и ГВС за ' + str(request.session["electric_data_end"])
    
    ws['A5'] = 'Дата'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Абонент'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Заводской номер счётчика'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания ХВС, м3'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания ГВС, м3'
    ws['E5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end, 'daily', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table=common_sql.get_daily_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end, 'daily', False)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Дата
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  #  абонент
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  #хвс
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # гвс
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next

    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25 
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'water_report' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

#_________________________________-

def pokazaniya_heat_current_report(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания теплосчётчиков. Последние считанные данные'
    

    ws['A5'] = 'Дата'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Время'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Абонент'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Заводской номер'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания, Гкал'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Показания, м3'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Время работы, ч'
    ws['G5'].style = ali_grey
    
    ws['H5'] = 'Твхода, С'
    ws['H5'].style = ali_grey

    ws['I5'] = 'Твыхода, С'
    ws['I5'].style = ali_grey
    
    ws['J5'] = 'Разница Т, С'
    ws['J5'].style = ali_grey
    
    ws['K5'] = 'Код ошибки'
    ws['K5'].style = ali_grey
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    #electric_data_end   = request.session['electric_data_end']
    #electric_data_start   = request.session['electric_data_start']                        
    obj_key             = request.session['obj_key']
    #is_electric_monthly = request.session['is_electric_monthly']
    #is_electric_daily   = request.session['is_electric_daily']
    #data_table_end   = []
    #data_table_start = []
    list_except = [u'ВРУ Счётчик01',u'ВРУ Счётчик02',u'ВРУ Счётчик03',u'ВРУ Счётчик04',u'ВРУ Счётчик05',u'ВРУ Счётчик06',u'ВРУ Счётчик07',u'ВРУ Счётчик08',u'ВРУ Счётчик09',u'ВРУ Счётчик10',u'ВРУ Счётчик11',u'ВРУ Счётчик12',u'ВРУ Счётчик13',u'ВРУ Счётчик14',u'ВРУ Счётчик15',u'ВРУ Счётчик16',u'ВРУ Счётчик17',u'ВРУ Счётчик18',u'ВРУ Счётчик19',u'ВРУ Счётчик20',u'ВРУ Счётчик21',u'ВРУ Счётчик22',u'ВРУ Счётчик23',u'Гараж Счётчик 1',u'Гараж Счётчик 2']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            #request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            #request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            #request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            #request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']  

    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_current_heat(meters_name, parent_name)
    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []
        for x in range(len(list_of_abonents_2)):
            data_table_temp = common_sql.get_data_table_current_heat(list_of_abonents_2[x], parent_name)
            
            if list_of_abonents_2[x][0] in list_except:
                next
            elif data_table_temp:            
                data_table.extend(data_table_temp)
            else:
                data_table.extend([[u'Н/Д',u'Н/Д',list_of_abonents_2[x][0],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
                             
    else:
        data_table = []

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по расходу воды
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Абонент
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # заводской номер
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # Показания по расходу воды
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][8])  # заводской номер
            ws.cell('I%s'%(row)).style = ali_white
        except:
            ws.cell('I%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][9])  # Показания по теплу
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][10])  # Показания по расходу воды
            ws.cell('K%s'%(row)).style = ali_white
        except:
            ws.cell('K%s'%(row)).style = ali_white

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'heat_current_report' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

#_________________________________________________________________________
def report_all_res_by_date(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]

#Шапка
    
    ws['A1'] = 'Лицевой счет'
    ws['A1'].style = ali_grey
    
    ws['B1'] = 'Дата начала работы'
    ws['B1'].style = ali_grey
    
    ws['C1'] = 'Номер прибора'
    ws['C1'].style = ali_grey
    
    ws['D1'] = 'тип прибора'
    ws['D1'].style = ali_grey
    
    ws['E1'] = 'Наименование ПУ'
    ws['E1'].style = ali_grey
    
    ws['F1'] = 'Показания'
    ws['F1'].style = ali_grey
    
    ws['G1'] = 'Дата Окончания' 
    ws['G1'].style = ali_grey
    
    ws['H1'] = 'Дата'
    ws['H1'].style = ali_grey

    ws['I1'] = 'Квартирный' 
    ws['I1'].style = ali_grey
    
    ws['J1'] = 'Адрес'
    ws['J1'].style = ali_grey
    
    ws['K1'] = 'Операция'
    ws['K1'].style = ali_grey
    
    ws['L1'] = 'Заменяемый счетчик'
    ws['L1'].style = ali_grey
    
    ws['M1'] = 'Показания заменяемого счетчика'
    ws['M1'].style = ali_grey

    ws['N1'] = 'Улица'
    ws['N1'].style = ali_grey

    ws['O1'] = 'Дом'
    ws['O1'].style = ali_grey

    ws['P1'] = 'Квартира'
    ws['P1'].style = ali_grey
#Запрашиваем данные для отчета

           
    
    data_table = []
    data_table = common_sql.get_data_table_report_all_res_by_date(electric_data_end)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)


# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  # лицевой счёт
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-2][1])  # начальная дата
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  # номер счётчика
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-2][3])  # тип энергоресурса
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-2][4])  # наименование ПУ
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-2][5])  # показания на начальную дату
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-2][6])  # дата снятия показаний
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-2][8])  # Дом-корпус
            ws.cell('O%s'%(row)).style = ali_white
        except:
            ws.cell('O%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-2][7])  # № квартиры
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            

    ws.row_dimensions[5].height = 41
#    ws.column_dimensions['A'].width = 10 
#    ws.column_dimensions['B'].width = 10 
#    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['j'].width = 15
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_all_resources_'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_electric_all_by_date(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]

#Шапка
    
    ws['A1'] = 'Лицевой счет'
    ws['A1'].style = ali_grey
    
    ws['B1'] = 'Дата начала работы'
    ws['B1'].style = ali_grey
    
    ws['C1'] = 'Номер прибора'
    ws['C1'].style = ali_grey
    
    ws['D1'] = 'тип прибора'
    ws['D1'].style = ali_grey
    
    ws['E1'] = 'Наименование ПУ'
    ws['E1'].style = ali_grey
    
    ws['F1'] = 'Показания'
    ws['F1'].style = ali_grey
    
    ws['G1'] = 'Дата Окончания' 
    ws['G1'].style = ali_grey
    
    ws['H1'] = 'Дата'
    ws['H1'].style = ali_grey

    ws['I1'] = 'Квартирный' 
    ws['I1'].style = ali_grey
    
    ws['J1'] = 'Адрес'
    ws['J1'].style = ali_grey
    
    ws['K1'] = 'Операция'
    ws['K1'].style = ali_grey
    
    ws['L1'] = 'Заменяемый счетчик'
    ws['L1'].style = ali_grey
    
    ws['M1'] = 'Показания заменяемого счетчика'
    ws['M1'].style = ali_grey

    ws['N1'] = 'Улица'
    ws['N1'].style = ali_grey

    ws['O1'] = 'Дом'
    ws['O1'].style = ali_grey

    ws['P1'] = 'Квартира'
    ws['P1'].style = ali_grey
#Запрашиваем данные для отчета

           
    
    data_table = []
    data_table = common_sql.get_data_table_report_electric_res_by_date(electric_data_end)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)


# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  # лицевой счёт
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-2][1])  # начальная дата
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  # номер счётчика
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-2][3])  # тип энергоресурса
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-2][4])  # наименование ПУ
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-2][5])  # показания на начальную дату
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-2][6])  # дата снятия показаний
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-2][8])  # Дом-корпус
            ws.cell('O%s'%(row)).style = ali_white
        except:
            ws.cell('O%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-2][7])  # № квартиры
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            

    ws.row_dimensions[5].height = 41
#    ws.column_dimensions['A'].width = 10 
#    ws.column_dimensions['B'].width = 10 
#    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['j'].width = 15
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_electric_resources_'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_water_all_by_date(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]

#Шапка
    
    ws['A1'] = 'Лицевой счет'
    ws['A1'].style = ali_grey
    
    ws['B1'] = 'Дата начала работы'
    ws['B1'].style = ali_grey
    
    ws['C1'] = 'Номер прибора'
    ws['C1'].style = ali_grey
    
    ws['D1'] = 'тип прибора'
    ws['D1'].style = ali_grey
    
    ws['E1'] = 'Наименование ПУ'
    ws['E1'].style = ali_grey
    
    ws['F1'] = 'Показания'
    ws['F1'].style = ali_grey
    
    ws['G1'] = 'Дата Окончания' 
    ws['G1'].style = ali_grey
    
    ws['H1'] = 'Дата'
    ws['H1'].style = ali_grey

    ws['I1'] = 'Квартирный' 
    ws['I1'].style = ali_grey
    
    ws['J1'] = 'Адрес'
    ws['J1'].style = ali_grey
    
    ws['K1'] = 'Операция'
    ws['K1'].style = ali_grey
    
    ws['L1'] = 'Заменяемый счетчик'
    ws['L1'].style = ali_grey
    
    ws['M1'] = 'Показания заменяемого счетчика'
    ws['M1'].style = ali_grey

    ws['N1'] = 'Улица'
    ws['N1'].style = ali_grey

    ws['O1'] = 'Дом'
    ws['O1'].style = ali_grey

    ws['P1'] = 'Квартира'
    ws['P1'].style = ali_grey
#Запрашиваем данные для отчета

           
    
    data_table = []
    data_table = common_sql.get_data_table_report_water_res_by_date(electric_data_end)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)


# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  # лицевой счёт
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-2][1])  # начальная дата
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  # номер счётчика
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-2][3])  # тип энергоресурса
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-2][4])  # наименование ПУ
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-2][5])  # показания на начальную дату
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-2][6])  # дата снятия показаний
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-2][8])  # Дом-корпус
            ws.cell('O%s'%(row)).style = ali_white
        except:
            ws.cell('O%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-2][7])  # № квартиры
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            

    ws.row_dimensions[5].height = 41
#    ws.column_dimensions['A'].width = 10 
#    ws.column_dimensions['B'].width = 10 
#    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['j'].width = 15
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_water_resources_'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_heat_all_by_date(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]

#Шапка
    
    ws['A1'] = 'Лицевой счет'
    ws['A1'].style = ali_grey
    
    ws['B1'] = 'Дата начала работы'
    ws['B1'].style = ali_grey
    
    ws['C1'] = 'Номер прибора'
    ws['C1'].style = ali_grey
    
    ws['D1'] = 'тип прибора'
    ws['D1'].style = ali_grey
    
    ws['E1'] = 'Наименование ПУ'
    ws['E1'].style = ali_grey
    
    ws['F1'] = 'Показания'
    ws['F1'].style = ali_grey
    
    ws['G1'] = 'Дата Окончания' 
    ws['G1'].style = ali_grey
    
    ws['H1'] = 'Дата'
    ws['H1'].style = ali_grey

    ws['I1'] = 'Квартирный' 
    ws['I1'].style = ali_grey
    
    ws['J1'] = 'Адрес'
    ws['J1'].style = ali_grey
    
    ws['K1'] = 'Операция'
    ws['K1'].style = ali_grey
    
    ws['L1'] = 'Заменяемый счетчик'
    ws['L1'].style = ali_grey
    
    ws['M1'] = 'Показания заменяемого счетчика'
    ws['M1'].style = ali_grey

    ws['N1'] = 'Улица'
    ws['N1'].style = ali_grey

    ws['O1'] = 'Дом'
    ws['O1'].style = ali_grey

    ws['P1'] = 'Квартира'
    ws['P1'].style = ali_grey
#Запрашиваем данные для отчета

           
    
    data_table = []
    data_table = common_sql.get_data_table_report_heat_res_by_date(electric_data_end)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)


# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  # лицевой счёт
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-2][1])  # начальная дата
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  # номер счётчика
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-2][3])  # тип энергоресурса
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-2][4])  # наименование ПУ
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-2][5])  # показания на начальную дату
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-2][6])  # дата снятия показаний
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-2][8])  # Дом-корпус
            ws.cell('O%s'%(row)).style = ali_white
        except:
            ws.cell('O%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-2][7])  # № квартиры
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            

    ws.row_dimensions[5].height = 41
#    ws.column_dimensions['A'].width = 10 
#    ws.column_dimensions['B'].width = 10 
#    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['j'].width = 15
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_heat_resources_'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_resources_all(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_start = request.session["electric_data_start"]
    electric_data_end   = request.session["electric_data_end"]
#Шапка
    ws.merge_cells('A2:G2')
    ws['A2'] = 'Филиград: 1, 2, 3 корпуса. Показания по энергоресурсам за период c '+electric_data_start +' по '+electric_data_end
    
    ws['A5'] = 'Лицевой счёт'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Дата начальная'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Номер прибора'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Тип прибора'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания на '+electric_data_end
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Показания на '+electric_data_start
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Потребление' 
    ws['G5'].style = ali_grey
    
    ws['H5'] = 'Дата установки'
    ws['H5'].style = ali_grey

    ws['I5'] = 'Дата конечная' 
    ws['I5'].style = ali_grey
    
    ws['J5'] = 'Абонент'
    ws['J5'].style = ali_grey
    
#Запрашиваем данные для отчета

           
    
    data_table = []
    data_table = common_sql.get_data_table_report_all_res_period3(electric_data_start, electric_data_end)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
#        
#    #удаляем из номеров счётчиков лишнее
#    for i in range(len(data_table)):
#        data_table[i]=list(data_table[i])
#        num=data_table[i][3]
#        if ('ХВС, №' in num) or ('ГВС, №' in num):
#            num=num.replace(u'ХВС, №', ' ')
#            num=num.replace(u'ГВС, №', ' ')
#            data_table[i][3]=num
#            #print num
#        data_table[i]=tuple(data_table[i])


# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # лицевой счёт
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # начальная дата
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # номер счётчика
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # тип энергоресурса
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # показания на конченую дату
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # показания на начальную дату
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][7])  # Дельта
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][8])  # Дата установки приборов
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][9])  # Дата конечная
            ws.cell('I%s'%(row)).style = ali_white
        except:
            ws.cell('I%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][10])  # Абонент
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
#    ws.column_dimensions['A'].width = 10 
#    ws.column_dimensions['B'].width = 10 
#    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['j'].width = 15
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_all_resources_'+ str(electric_data_start)+'_'+str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def pokazaniya_heat_current_report_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания теплосчётчиков. Последние считанные данные'
    
    ws['A5'] = 'Дата'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Время'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Абонент'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Заводской номер'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания, Гкал'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Показания, м3'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Время работы, ч'
    ws['G5'].style = ali_grey
    
    ws['H5'] = 'Твхода, С'
    ws['H5'].style = ali_grey

    ws['I5'] = 'Твыхода, С'
    ws['I5'].style = ali_grey
    
    ws['J5'] = 'Разница Т, С'
    ws['J5'].style = ali_grey
    
    ws['K5'] = 'Код ошибки'
    ws['K5'].style = ali_grey
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    obj_key             = request.session['obj_key']

    list_except = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']

    data_table=[]
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_current_heat_v2(obj_title, parent_name, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_current_heat_v2(obj_title, parent_name, False)
        for row in data_table:
            for x in list_except:
                if x==row[2]:
                    data_table.remove(x)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по расходу воды
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Абонент
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # заводской номер
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # Показания по расходу воды
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][8])  # заводской номер
            ws.cell('I%s'%(row)).style = ali_white
        except:
            ws.cell('I%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][9])  # Показания по теплу
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][10])  # Показания по расходу воды
            ws.cell('K%s'%(row)).style = ali_white
        except:
            ws.cell('K%s'%(row)).style = ali_white

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'heat_current_report' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_potreblenie_heat_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление тепловой энергии в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_start"])
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление, Гкал'
    ws['E5'].style = ali_grey
    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']                        
    obj_key             = request.session['obj_key']
    #is_electric_monthly = request.session['is_electric_monthly']
    #is_electric_daily   = request.session['is_electric_daily']
    #data_table_end   = []
    #data_table_start = []
    list_except = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_for_period_for_abon_heat_v2(meters_name, parent_name, electric_data_start, electric_data_end)

    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_for_period_heat_v2(meters_name, parent_name, electric_data_start, electric_data_end)
        for row in data_table:
            for x in list_except:
                if x==row[2]:
                    data_table.remove(x)
    else:
        data_table = []

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
#        try:
#            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # Время работы
#            ws.cell('F%s'%(row)).style = ali_white
#        except:
#            ws.cell('F%s'%(row)).style = ali_white
#            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_heat_report'
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_water_potreblenie_tekon_hvs(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ХВС в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_start"])
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление, м3'
    ws['E5'].style = ali_grey
    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_start = request.session['electric_data_start']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name, electric_data_start, electric_data_end, u'Канал 1', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name,electric_data_start, electric_data_end, u'Канал 1', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
#        try:
#            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # Время работы
#            ws.cell('F%s'%(row)).style = ali_white
#        except:
#            ws.cell('F%s'%(row)).style = ali_white
#            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_water_tekon_hvs_report'
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_water_tekon_hvs(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ХВС на ' +str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['C5'].style = ali_grey

    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']    
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, u'Канал 1', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, u'Канал 1', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35

#    ws.column_dimensions['F'].width = 18
#____________
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'pokazaniya_na_datu_water_tekon_hvs_report'
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_water_potreblenie_tekon_gvs(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ГВС в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_start"])
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление, м3'
    ws['E5'].style = ali_grey
    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_start = request.session['electric_data_start']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name, electric_data_start, electric_data_end, u'Канал 2', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name,electric_data_start, electric_data_end, u'Канал 2', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
#        try:
#            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # Время работы
#            ws.cell('F%s'%(row)).style = ali_white
#        except:
#            ws.cell('F%s'%(row)).style = ali_white
#            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_water_tekon_gvs_report'
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_water_tekon_gvs(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ГВС на ' +str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['C5'].style = ali_grey

    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']    
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, u'Канал 2', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, u'Канал 2', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35

#    ws.column_dimensions['F'].width = 18
#____________
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'pokazaniya_na_datu_water_tekon_gvs_report'
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_potreblenie_heat(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление тепловой энергии в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_start"])
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление, Гкал'
    ws['E5'].style = ali_grey
    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']                        
    obj_key             = request.session['obj_key']
    #is_electric_monthly = request.session['is_electric_monthly']
    #is_electric_daily   = request.session['is_electric_daily']
    #data_table_end   = []
    #data_table_start = []
    list_except = [u'ВРУ Счётчик01',u'ВРУ Счётчик02',u'ВРУ Счётчик03',u'ВРУ Счётчик04',u'ВРУ Счётчик05',u'ВРУ Счётчик06',u'ВРУ Счётчик07',u'ВРУ Счётчик08',u'ВРУ Счётчик09',u'ВРУ Счётчик10',u'ВРУ Счётчик11',u'ВРУ Счётчик12',u'ВРУ Счётчик13',u'ВРУ Счётчик14',u'ВРУ Счётчик15',u'ВРУ Счётчик16',u'ВРУ Счётчик17',u'ВРУ Счётчик18',u'ВРУ Счётчик19',u'ВРУ Счётчик20',u'ВРУ Счётчик21',u'ВРУ Счётчик22',u'ВРУ Счётчик23',u'Гараж Счётчик 1',u'Гараж Счётчик 2']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table_end = common_sql.get_data_table_by_date_heat(meters_name, parent_name, electric_data_end)
        #print data_table_end
        data_table_start = common_sql.get_data_table_by_date_heat(meters_name, parent_name, electric_data_start)
        #print data_table_start
        data_table = []
        for x in range(len(data_table_end)):
            try:
                data_table_temp=[data_table_end[x][0], data_table_end[x][1], data_table_end[x][2], data_table_start[x][3], data_table_end[x][3], data_table_end[x][3]-data_table_start[x][3], data_table_end[x][5] - data_table_start[x][5]]
                data_table.append(data_table_temp)
            except:
                data_table = []
            

    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []
        for x in range(len(list_of_abonents_2)):
            data_table_end_temp = common_sql.get_data_table_by_date_heat(list_of_abonents_2[x][0], meters_name, electric_data_end)
            data_table_start_temp = common_sql.get_data_table_by_date_heat(list_of_abonents_2[x][0], meters_name, electric_data_start)
            data_table_temp = []
            for x in range(len(data_table_end_temp)):

                data_table_temp_2 = []
                try:
                    data_table_temp_2.append(data_table_end_temp[x][0])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_end_temp[x][1])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_end_temp[x][2])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_start_temp[x][3])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_end_temp[x][3])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_end_temp[x][3]-data_table_start_temp[x][3])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                #try:
                #    data_table_temp_2.append(data_table_end_temp[x][5]-data_table_start_temp[x][5])
                #except IndexError:
                #    data_table_temp_2.append(u"Н/Д")
                #except TypeError:
                #    data_table_temp_2.append(u"Н/Д")

                data_table_temp.append(data_table_temp_2)
                
            data_table_end_temp = []
            data_table_start_temp = []
            
            if list_of_abonents_2[x][0] in list_except:
                next
            elif data_table_temp:            
                data_table.extend(data_table_temp)
            else:
                data_table.extend([[0,list_of_abonents_2[x][0],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
                
              
    else:
        data_table = []

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
#        try:
#            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # Время работы
#            ws.cell('F%s'%(row)).style = ali_white
#        except:
#            ws.cell('F%s'%(row)).style = ali_white
#            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_heat_report'
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_water_by_date(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = meters_name+'. Показания по воде на ' + electric_data_end
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Пульсар'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Канал'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания'
    ws['E5'].style = ali_grey
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'level2')
    is_object_level_2 = re.compile(r'level1')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_by_date(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_by_date(meters_name, parent_name, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # пульсар
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # канал
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
            
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 25
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'heat_water_report_'+translate(meters_name)+'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

    
def report_forma_80020(request):
    import zipfile
    response = StringIO.StringIO()
    
    
    #Запрашиваем данные для отчета
    
    group_80020_name    = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']                        
   

    # Формируем список дат на основе начальной и конечной даты полученной от web-календарей
    end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
    start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
    list_of_dates = [x for x in common_sql.daterange(start_date,
                  end_date,
                  step=datetime.timedelta(days=1),
                  inclusive=True)]
    
    info_group_80020 = common_sql.get_info_group_80020(group_80020_name)
    inn_sender_from_base      = info_group_80020[0][0]
    name_sender_from_base     = info_group_80020[0][1]
    inn_postavshik_from_base  = info_group_80020[0][2]
    name_postavshik_from_base = info_group_80020[0][3]
    dogovor_number_from_base  = info_group_80020[0][4]
    
    #Узнаем GUID счётчиков, которые входят в группу 80020
    meters_guid_list = common_sql.get_meters_guid_list_by_group_name(group_80020_name)

   
    #Создаем архив
    zf = zipfile.ZipFile(response, mode='w', compression=zipfile.ZIP_DEFLATED)
    
    for dates in range(len(list_of_dates)):
    #Формируем файл xml по форме Мосэнергосбыт 80020   
        
        # Создание корневого элемента message
        root = etree.Element('message', {'class': '80020', 'version': '2', 'number': '1' })
        
        # Добавление дочерних элементов - <datetime> <sender> <area> в <root>
        datetimeElt = etree.SubElement(root, 'datetime')
        senderElt = etree.SubElement(root, 'sender')
        areaElt = etree.SubElement(root, 'area')
        
        # Присваиваем значения в <day> <timestamp> <daylightsavingtime>
        day = etree.SubElement(datetimeElt, 'day')
        timestamp = etree.SubElement(datetimeElt, 'timestamp')
        daylightsavingtime = etree.SubElement(datetimeElt, 'daylightsavingtime')
        
        day.text = unicode(list_of_dates[0].strftime('%Y%m%d'))
        timestamp.text = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        daylightsavingtime.text = u'0'
        
        # Присваиваем значения <name> <inn> для <sender>
        inn_sender = etree.SubElement(senderElt, 'inn')
        name_sender = etree.SubElement(senderElt, 'name')
        
        inn_sender.text  = inn_sender_from_base
        name_sender.text = name_sender_from_base
        
        # Присваиваем значения для <area>
        inn_area = etree.SubElement(areaElt, 'inn')
        name_abonent_area = etree.SubElement(areaElt, 'name_abonent')
        
        inn_area.text = inn_postavshik_from_base
        name_abonent_area.text = name_postavshik_from_base
        
        # Добавление дочерних элементов
        for x in range(len(meters_guid_list)):
            info_measuring_point = common_sql.get_info_measuring_point_in_group_80020(meters_guid_list[x])
            my_measure_code = unicode(info_measuring_point[0][0])
            my_measure_name = unicode(info_measuring_point[0][1])
            measurepointElt = etree.SubElement(areaElt, 'measuringpoint', code = my_measure_code , name = my_measure_name )
        
            list_of_taken_params = []
            #Получаем считываемые параметры по guid счётчика.
             #A+
            guid_params = u'6af9ddce-437a-4e07-bd70-6cf9dcc10b31'
            result = common_sql.get_taken_param_by_guid_meters_and_guid_params(meters_guid_list[x], guid_params)
            result_list = []
            result_list.append(unicode(result[0][0]))
            result_list.append(unicode(result[0][1]))
            list_of_taken_params.append(result_list)
            #R+
            guid_params = u'66e997c0-8128-40a7-ae65-7e8993fbea61'
            result = common_sql.get_taken_param_by_guid_meters_and_guid_params(meters_guid_list[x], guid_params)
            result_list = []
            result_list.append(unicode(result[0][0]))
            result_list.append(unicode(result[0][1]))
            list_of_taken_params.append(result_list)
            #print list_of_taken_params
            
            for y in range(len(list_of_taken_params)):
                my_dict_of_profil = {0: [u'0000', u'0030', 0, 1],
                                     1: [u'0030', u'0100', 0, 1],
                                     2: [u'0100', u'0130', 0, 1],
                                     3: [u'0130', u'0200', 0, 1],
                                     4: [u'0200', u'0230', 0, 1],
                                     5: [u'0230', u'0300', 0, 1],
                                     6: [u'0300', u'0330', 0, 1],
                                     7: [u'0330', u'0400', 0, 1],
                                     8: [u'0400', u'0430', 0, 1],
                                     9: [u'0430', u'0500', 0, 1],
                                     10: [u'0500', u'0530', 0, 1],
                                     11: [u'0530', u'0600', 0, 1],
                                     12: [u'0600', u'0630', 0, 1],
                                     13: [u'0630', u'0700', 0, 1],
                                     14: [u'0700', u'0730', 0, 1],
                                     15: [u'0730', u'0800', 0, 1],
                                     16: [u'0800', u'0830', 0, 1],
                                     17: [u'0830', u'0900', 0, 1],
                                     18: [u'0900', u'0930', 0, 1],
                                     19: [u'0930', u'1000', 0, 1],
                                     20: [u'1000', u'1030', 0, 1],
                                     21: [u'1030', u'1100', 0, 1],
                                     22: [u'1100', u'1130', 0, 1],                                
                                     23: [u'1130', u'1200', 0, 1],
                                     24: [u'1200', u'1230', 0, 1],
                                     25: [u'1230', u'1300', 0, 1],
                                     26: [u'1300', u'1330', 0, 1],
                                     27: [u'1330', u'1400', 0, 1],
                                     28: [u'1400', u'1430', 0, 1],
                                     29: [u'1430', u'1500', 0, 1],
                                     30: [u'1500', u'1530', 0, 1],
                                     31: [u'1530', u'1600', 0, 1],
                                     32: [u'1600', u'1630', 0, 1],
                                     33: [u'1630', u'1700', 0, 1],
                                     34: [u'1700', u'1730', 0, 1],
                                     35: [u'1730', u'1800', 0, 1],
                                     36: [u'1800', u'1830', 0, 1],
                                     37: [u'1830', u'1900', 0, 1],
                                     38: [u'1900', u'1930', 0, 1],
                                     39: [u'1930', u'2000', 0, 1],
                                     40: [u'2000', u'2030', 0, 1],                                
                                     41: [u'2030', u'2100', 0, 1],
                                     42: [u'2100', u'2130', 0, 1],
                                     43: [u'2130', u'2200', 0, 1],
                                     44: [u'2200', u'2230', 0, 1],
                                     45: [u'2230', u'2300', 0, 1],
                                     46: [u'2300', u'2330', 0, 1],
                                     47: [u'2330', u'0000', 0, 1] }
                time_table = ['00:00:00', '00:30:00',
                              '01:00:00', '01:30:00',
                              '02:00:00', '02:30:00',
                              '03:00:00', '03:30:00',
                              '04:00:00', '04:30:00',
                              '05:00:00', '05:30:00',
                              '06:00:00', '06:30:00',
                              '07:00:00', '07:30:00',
                              '08:00:00', '08:30:00',
                              '09:00:00', '09:30:00',
                              '10:00:00', '10:30:00',
                              '11:00:00', '11:30:00',
                              '12:00:00', '12:30:00',
                              '13:00:00', '13:30:00',
                              '14:00:00', '14:30:00',
                              '15:00:00', '15:30:00',
                              '16:00:00', '16:30:00',
                              '17:00:00', '17:30:00',
                              '18:00:00', '18:30:00',
                              '19:00:00', '19:30:00',
                              '20:00:00', '20:30:00',
                              '21:00:00', '21:30:00',
                              '22:00:00', '22:30:00',
                              '23:00:00', '23:30:00']
                if list_of_taken_params[y][1] == u'A+ Профиль':
                    my_measuring_channel_code = u'01'
                    for time in range(len(time_table)):
                        result_30_min = common_sql.get_30_min_value_by_meters_number_param_names_and_datetime(list_of_taken_params[y][0], list_of_taken_params[y][1], list_of_dates[dates].strftime('%Y-%m-%d'), time_table[time])
                        if result_30_min:
                            my_dict_of_profil[time]=[my_dict_of_profil[time][0], my_dict_of_profil[time][1], float(result_30_min[0][4])*common_sql.get_k_t_t_by_factory_number_manual(list_of_taken_params[y][0])*0.5, 0] #Квт.ч
                            #print u'------ Есть значение',list_of_taken_params[y][1] , my_dict_of_profil[time][0], my_dict_of_profil[time][1]
                            
                        else:
                            #print u'Оставляем по умолчанию', list_of_taken_params[y][1], my_dict_of_profil[time][0], my_dict_of_profil[time][1]
                            pass
                    
                elif list_of_taken_params[y][1] == u'R+ Профиль':
                    my_measuring_channel_code = u'03'
                    for time in range(len(time_table)):
                        result_30_min = common_sql.get_30_min_value_by_meters_number_param_names_and_datetime(list_of_taken_params[y][0], list_of_taken_params[y][1], list_of_dates[dates].strftime('%Y-%m-%d'), time_table[time])
                        if result_30_min:
                            my_dict_of_profil[time]=[my_dict_of_profil[time][0], my_dict_of_profil[time][1], float(result_30_min[0][4])*common_sql.get_k_t_t_by_factory_number_manual(list_of_taken_params[y][0]), 0] # Квар.ч
                            #print u'------ Есть значение',list_of_taken_params[y][1], my_dict_of_profil[time][0], my_dict_of_profil[time][1]
                            
                        else:
                           # print u'Оставляем по умолчанию', list_of_taken_params[y][1], my_dict_of_profil[time][0], my_dict_of_profil[time][1]
                           pass
                else:
                    #print u'Нет совпадений параметров'
                    my_measuring_channel_code = u''
                       
                my_measuring_channel_desc = unicode(list_of_taken_params[y][0]) + u' ' + unicode(list_of_taken_params[y][1])            
                measuringchannelElt = etree.SubElement(measurepointElt, 'measuringchannel', code = my_measuring_channel_code, desc = my_measuring_channel_desc )
                
                                                  
                for z in range (len(my_dict_of_profil)):
                    periodElt = etree.SubElement(measuringchannelElt, 'period', start = unicode(my_dict_of_profil[z][0]), end = unicode(my_dict_of_profil[z][1]))
                    value  = etree.SubElement(periodElt, 'value', status = unicode(my_dict_of_profil[z][3]))
                    value.text = unicode(my_dict_of_profil[z][2])
            
        
        # Создание и сохранение документа
        doc = etree.ElementTree(root) 
        myxml_IO=StringIO.StringIO()   
        doc.write(myxml_IO, xml_declaration=True, encoding='UTF-8')
        # Формируем имя документа
        name_of_document = u'80020'
        name_of_file_80020 =name_of_document + u'_'+ inn_sender_from_base + u'_' + unicode(list_of_dates[dates].strftime('%Y%m%d')) + u'_1' + u'.xml'
        zf.writestr(name_of_file_80020, myxml_IO.getvalue())
       
    zf.close()
    
    response=HttpResponse(response.getvalue())
    response['Content-Type'] = 'application/x-zip-compressed'
    
    output_name = u'80020_' + unicode(dogovor_number_from_base) + u'_' + start_date.strftime('%Y%m%d') + u'-'+ end_date.strftime('%Y%m%d')
    file_ext = u'zip'
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   

   
    return response
